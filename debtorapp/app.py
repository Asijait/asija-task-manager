try:
    from . import db_compat as sqlite3
except ImportError:
    import db_compat as sqlite3
import csv
import os
import shutil
import json
import tempfile
import zipfile
import html
import re
from urllib.parse import urlencode
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
import pandas as pd
import io
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
app.secret_key = "dev-secret-key"

# Add the Indian currency formatting function
def format_indian_currency(amount, decimals=True):
    if amount is None:
        return ''
    amount = float(amount)
    if not decimals:
        amount = round(amount)

    is_negative = amount < 0
    amount = abs(amount)

    integer_part = int(amount)

    s = str(integer_part)
    if len(s) <= 3:
        formatted_integer = s
    else:
        last_three = s[-3:]
        other_digits = s[:-3]
        formatted_integer_parts = []
        while other_digits:
            if len(other_digits) > 2:
                formatted_integer_parts.insert(0, other_digits[-2:])
                other_digits = other_digits[:-2]
            else:
                formatted_integer_parts.insert(0, other_digits)
                other_digits = ''
        formatted_integer = ','.join(formatted_integer_parts) + ',' + last_three

    if decimals:
        decimal_part = round((amount - integer_part) * 100)
        result = formatted_integer + f'.{decimal_part:02d}'
    else:
        result = formatted_integer

    if is_negative:
        result = '-' + result
    return result

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'database.db')
MAIN_DB_PATH = os.path.join(os.path.dirname(BASE_DIR), 'tasks.db')
CSV_PATH = os.path.join(BASE_DIR, 'data.csv')
STATIC_DIR = os.path.join(BASE_DIR, 'static')
JS_DIR = os.path.join(STATIC_DIR, 'js')
TEMPLATE_DIR = os.path.join(BASE_DIR, 'templates')
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
BACKUP_DIR = os.path.join(BASE_DIR, 'backups')
BACKUP_LOG_PATH = os.path.join(BACKUP_DIR, 'update_log.json')
TALLY_RECEIPT_TRIAL_PATH = os.path.join(BASE_DIR, 'fcRece.xlsx')
BACKUP_EXCLUDE_DIRS = {'backups', '__pycache__'}
BACKUP_EXCLUDE_FILES = {'server.out.log', 'server.err.log'}
DB_BUSY_TIMEOUT_MS = 30000

def connect_debtor_db():
    conn = sqlite3.connect(DB_PATH, timeout=DB_BUSY_TIMEOUT_MS / 1000)
    conn.row_factory = sqlite3.Row
    conn.execute(f'PRAGMA busy_timeout = {DB_BUSY_TIMEOUT_MS}')
    conn.execute('PRAGMA journal_mode = WAL')
    return conn

DEBTOR_NAV_ACCESS_ITEMS = [
    {'id': 'import_tally_debtors', 'label': 'Import Tally Debtors', 'group': 'Import / Export', 'view_key': '', 'full_key': 'import_tally_debtors'},
    {'id': 'import_tally_receipts', 'label': 'Import Tally Receipts', 'group': 'Import / Export', 'view_key': 'import_tally_receipts_view', 'full_key': 'import_tally_receipts'},
    {'id': 'download_report_excel', 'label': 'Reports in Excel Format', 'group': 'Import / Export', 'view_key': 'download_report_excel', 'full_key': 'download_report_excel'},
    {'id': 'client_master', 'label': 'Client Master', 'group': 'Masters', 'view_key': 'client_master_view', 'full_key': 'client_master_edit'},
    {'id': 'client_group_master', 'label': 'Client Group Master', 'group': 'Masters', 'view_key': 'client_group_master_view', 'full_key': 'client_group_master_edit'},
    {'id': 'crp_master', 'label': 'CRP Master', 'group': 'Masters', 'view_key': 'crp_master_view', 'full_key': 'crp_master_edit'},
    {'id': 'firm_master', 'label': 'Firm Master', 'group': 'Masters', 'view_key': 'firm_master_view', 'full_key': 'firm_master_edit'},
    {'id': 'executive_partner_master', 'label': 'Executive Partner', 'group': 'Masters', 'view_key': 'executive_partner_master_view', 'full_key': 'executive_partner_master_edit'},
    {'id': 'run_followup_logic', 'label': 'Run Followup Logic', 'group': 'Masters', 'view_key': '', 'full_key': 'run_followup_logic'},
    {'id': 'clear_report', 'label': 'Clear Report', 'group': 'Masters', 'view_key': '', 'full_key': 'clear_report'},
]
DEBTOR_NAV_ACCESS_KEYS = {
    key
    for item in DEBTOR_NAV_ACCESS_ITEMS
    for key in (item.get('view_key'), item.get('full_key'))
    if key
}
DEBTOR_NAV_EDIT_TO_VIEW = {
    'import_tally_receipts': 'import_tally_receipts_view',
    'client_master_edit': 'client_master_view',
    'client_group_master_edit': 'client_group_master_view',
    'crp_master_edit': 'crp_master_view',
    'firm_master_edit': 'firm_master_view',
    'executive_partner_master_edit': 'executive_partner_master_view',
}

# Register the filter
app.jinja_env.filters['indian_currency'] = format_indian_currency

@app.context_processor
def debtor_report_template_helpers():
    def debtor_url(path='/'):
        if not path.startswith('/'):
            path = '/' + path
        return f"{request.script_root}{path}"
    user_email = str(session.get('user_email', '')).lower()
    is_arif = user_email == 'arif.siddiqui@asija.in'
    debtor_nav_access = get_debtor_nav_access_for_user(user_email)
    return {
        'debtor_url': debtor_url,
        'is_arif_user': is_arif,
        'debtor_nav_access': debtor_nav_access,
        'report_as_on_label': get_report_as_on_label(),
        'report_as_on_lines': get_report_as_on_lines(),
    }

def get_app_meta(cursor, key, default=''):
    try:
        cursor.execute('SELECT value FROM app_meta WHERE key = ? LIMIT 1', (key,))
        row = cursor.fetchone()
        if row:
            return row['value'] if isinstance(row, sqlite3.Row) else row[0]
    except sqlite3.Error:
        return default
    return default

def set_app_meta(cursor, key, value):
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS app_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    cursor.execute('''
        INSERT INTO app_meta (key, value)
        VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    ''', (key, value))

def ensure_debtor_nav_access_table(cursor):
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS debtor_nav_access (
            user_email TEXT NOT NULL,
            access_key TEXT NOT NULL,
            PRIMARY KEY (user_email, access_key)
        )
    ''')

def ensure_deleted_records_log_table(cursor):
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS deleted_records_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_table TEXT NOT NULL,
            source_pk TEXT,
            display_type TEXT,
            display_label TEXT,
            display_summary TEXT,
            payload_json TEXT NOT NULL,
            deleted_at TEXT,
            deleted_by TEXT,
            restore_status TEXT DEFAULT 'deleted',
            restored_at TEXT,
            restored_by TEXT
        )
    ''')

def sqlite_record_to_dict(row):
    if not row:
        return {}
    if isinstance(row, sqlite3.Row):
        return {key: row[key] for key in row.keys()}
    return dict(row)

def log_deleted_record(cursor, source_table, source_pk, display_type, display_label, payload, display_summary=''):
    ensure_deleted_records_log_table(cursor)
    cursor.execute('''
        INSERT INTO deleted_records_log (
            source_table, source_pk, display_type, display_label, display_summary,
            payload_json, deleted_at, deleted_by, restore_status
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'deleted')
    ''', (
        source_table,
        str(source_pk or ''),
        display_type,
        display_label,
        display_summary,
        json.dumps(payload or {}, default=str),
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        session.get('user_email', ''),
    ))

def insert_deleted_payload(cursor, table_name, payload):
    allowed_tables = {'billing_report', 'client_group_master', 'crp_master', 'firm_master'}
    if table_name not in allowed_tables:
        raise ValueError('Recall is not available for this record type.')

    cursor.execute(f'PRAGMA table_info({table_name})')
    table_columns = [row[1] for row in cursor.fetchall()]
    columns = [column for column in table_columns if column in payload]
    if not columns:
        raise ValueError('No saved data found for recall.')

    if 'id' in columns and payload.get('id') not in (None, ''):
        cursor.execute(f'SELECT id FROM {table_name} WHERE id = ?', (payload.get('id'),))
        if cursor.fetchone():
            columns = [column for column in columns if column != 'id']

    placeholders = ', '.join('?' for _ in columns)
    column_sql = ', '.join(columns)
    cursor.execute(
        f'INSERT INTO {table_name} ({column_sql}) VALUES ({placeholders})',
        [payload.get(column) for column in columns]
    )

def restore_deleted_log_record(cursor, delete_log_id):
    ensure_deleted_records_log_table(cursor)
    cursor.execute('''
        SELECT *
        FROM deleted_records_log
        WHERE id = ? AND COALESCE(restore_status, 'deleted') = 'deleted'
        LIMIT 1
    ''', (delete_log_id,))
    log_row = cursor.fetchone()
    if not log_row:
        return False, 'Deleted record not found or already recalled.'

    source_table = log_row['source_table'] if isinstance(log_row, sqlite3.Row) else log_row[1]
    payload_json = log_row['payload_json'] if isinstance(log_row, sqlite3.Row) else log_row[6]
    payload = json.loads(payload_json or '{}')

    if source_table == 'billing_report':
        row_id = payload.get('id') or (log_row['source_pk'] if isinstance(log_row, sqlite3.Row) else log_row[2])
        cursor.execute('SELECT id FROM billing_report WHERE id = ?', (row_id,))
        if cursor.fetchone():
            cursor.execute('''
                UPDATE billing_report
                SET deleted_at = NULL, deleted_by = NULL, delete_reason = NULL
                WHERE id = ?
            ''', (row_id,))
        else:
            payload['deleted_at'] = None
            payload['deleted_by'] = None
            payload['delete_reason'] = None
            insert_deleted_payload(cursor, 'billing_report', payload)
    else:
        insert_deleted_payload(cursor, source_table, payload)
        if source_table == 'client_group_master':
            sync_client_master_from_group_master(cursor)

    cursor.execute('''
        UPDATE deleted_records_log
        SET restore_status = 'recalled', restored_at = ?, restored_by = ?
        WHERE id = ?
    ''', (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get('user_email', ''), delete_log_id))
    return True, 'Deleted record recalled successfully.'

def build_deleted_records_rows(cursor):
    ensure_deleted_records_log_table(cursor)
    rows = []
    cursor.execute('''
        SELECT *
        FROM deleted_records_log
        WHERE COALESCE(restore_status, 'deleted') = 'deleted'
        ORDER BY deleted_at DESC, id DESC
    ''')
    log_rows = cursor.fetchall()
    logged_billing_ids = set()
    for row in log_rows:
        row_dict = sqlite_record_to_dict(row)
        if row_dict.get('source_table') == 'billing_report':
            logged_billing_ids.add(str(row_dict.get('source_pk') or ''))
        rows.append({
            'delete_log_id': row_dict.get('id'),
            'legacy_billing_id': '',
            'deleted_at': row_dict.get('deleted_at') or '',
            'deleted_by': row_dict.get('deleted_by') or '',
            'record_type': row_dict.get('display_type') or row_dict.get('source_table') or '',
            'record_name': row_dict.get('display_label') or row_dict.get('source_pk') or '',
            'details': row_dict.get('display_summary') or '',
        })

    cursor.execute('''
        SELECT *
        FROM billing_report
        WHERE deleted_at IS NOT NULL
        ORDER BY deleted_at DESC, id DESC
    ''')
    for row in cursor.fetchall():
        row_dict = sqlite_record_to_dict(row)
        if str(row_dict.get('id') or '') in logged_billing_ids:
            continue
        details = ' | '.join(part for part in [
            row_dict.get('short_name') or row_dict.get('firm_name') or '',
            row_dict.get('bill_date') or '',
            row_dict.get('ref_no') or '',
            f"Rs {format_indian_currency(row_dict.get('amount') or 0, decimals=False)}",
        ] if part)
        rows.append({
            'delete_log_id': '',
            'legacy_billing_id': row_dict.get('id'),
            'deleted_at': row_dict.get('deleted_at') or '',
            'deleted_by': row_dict.get('deleted_by') or '',
            'record_type': 'Billing Report',
            'record_name': row_dict.get('party_name') or row_dict.get('ref_no') or row_dict.get('id'),
            'details': details,
        })

    rows.sort(key=lambda item: (item.get('deleted_at') or '', str(item.get('delete_log_id') or item.get('legacy_billing_id') or '')), reverse=True)
    return rows

def soft_delete_billing_rows(cursor, row_ids, delete_reason=''):
    deleted_count = 0
    deleted_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    deleted_by = session.get('user_email', '')
    for row_id in row_ids:
        cursor.execute('SELECT * FROM billing_report WHERE id = ? AND deleted_at IS NULL', (row_id,))
        row = cursor.fetchone()
        if not row:
            continue
        row_dict = sqlite_record_to_dict(row)
        log_deleted_record(
            cursor,
            'billing_report',
            row_id,
            'Billing Report',
            row_dict.get('party_name') or row_dict.get('ref_no') or f"Bill #{row_id}",
            row_dict,
            ' | '.join(part for part in [
                row_dict.get('short_name') or row_dict.get('firm_name') or '',
                row_dict.get('bill_date') or '',
                row_dict.get('ref_no') or '',
                f"Rs {format_indian_currency(row_dict.get('amount') or 0, decimals=False)}",
            ] if part),
        )
        cursor.execute(
            '''
            UPDATE billing_report
            SET deleted_at = ?, deleted_by = ?, delete_reason = ?
            WHERE id = ? AND deleted_at IS NULL
            ''',
            (deleted_at, deleted_by, delete_reason, row_id)
        )
        deleted_count += cursor.rowcount
    return deleted_count

def ensure_client_group_master_seed(cursor):
    cursor.execute('''
        SELECT
            client_group,
            COALESCE(crp_of_group, '') AS crp_of_group,
            COALESCE(reffered_by, '') AS reffered_by,
            COUNT(*) AS use_count
        FROM client_master
        WHERE client_group IS NOT NULL AND trim(client_group) != ''
        GROUP BY lower(trim(client_group)), client_group, crp_of_group, reffered_by
        ORDER BY lower(trim(client_group)), use_count DESC, client_group
    ''')
    seeded_groups = set()
    for row in cursor.fetchall():
        group_name = (row['client_group'] if isinstance(row, sqlite3.Row) else row[0]) or ''
        crp_name = (row['crp_of_group'] if isinstance(row, sqlite3.Row) else row[1]) or ''
        reffered_by = (row['reffered_by'] if isinstance(row, sqlite3.Row) else row[2]) or ''
        group_name = group_name.strip()
        group_key = group_name.lower()
        if not group_name or group_key in seeded_groups:
            continue
        cursor.execute('''
            INSERT OR IGNORE INTO client_group_master (group_name, crp_name, reffered_by)
            VALUES (?, ?, ?)
        ''', (group_name, crp_name.strip(), reffered_by.strip()))
        seeded_groups.add(group_key)

def ensure_crp_master_seed(cursor):
    cursor.execute('''
        INSERT OR IGNORE INTO crp_master (crp_name)
        SELECT DISTINCT trim(crp_name)
        FROM client_group_master
        WHERE crp_name IS NOT NULL AND trim(crp_name) != ''
    ''')
    cursor.execute('''
        INSERT OR IGNORE INTO crp_master (crp_name)
        SELECT DISTINCT trim(crp_of_group)
        FROM client_master
        WHERE crp_of_group IS NOT NULL AND trim(crp_of_group) != ''
    ''')

def get_crp_options(cursor):
    cursor.execute('''
        SELECT crp_name
        FROM crp_master
        WHERE crp_name IS NOT NULL AND trim(crp_name) != ''
        ORDER BY lower(crp_name)
    ''')
    return [row['crp_name'] if isinstance(row, sqlite3.Row) else row[0] for row in cursor.fetchall()]

def resolve_client_group_parent(cursor, client_group, crp_of_group='', reffered_by=''):
    group_name = (client_group or '').strip()
    crp_of_group = (crp_of_group or '').strip()
    reffered_by = (reffered_by or '').strip()
    if not group_name:
        return group_name, crp_of_group, reffered_by

    cursor.execute('''
        SELECT id, group_name, COALESCE(crp_name, '') AS crp_name, COALESCE(reffered_by, '') AS reffered_by
        FROM client_group_master
        WHERE lower(trim(group_name)) = lower(trim(?))
        LIMIT 1
    ''', (group_name,))
    row = cursor.fetchone()
    if row:
        parent_id = row['id'] if isinstance(row, sqlite3.Row) else row[0]
        parent_group = (row['group_name'] if isinstance(row, sqlite3.Row) else row[1]) or group_name
        parent_crp = (row['crp_name'] if isinstance(row, sqlite3.Row) else row[2]) or ''
        parent_referred = (row['reffered_by'] if isinstance(row, sqlite3.Row) else row[3]) or ''

        updated_crp = parent_crp or crp_of_group
        updated_referred = parent_referred or reffered_by
        if updated_crp != parent_crp or updated_referred != parent_referred:
            cursor.execute('''
                UPDATE client_group_master
                SET crp_name = ?, reffered_by = ?, updated_at = ?
                WHERE id = ?
            ''', (updated_crp, updated_referred, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), parent_id))

        return parent_group.strip(), updated_crp.strip(), updated_referred.strip()

    cursor.execute('''
        INSERT INTO client_group_master (group_name, crp_name, reffered_by, updated_at)
        VALUES (?, ?, ?, ?)
    ''', (group_name, crp_of_group, reffered_by, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    return group_name, crp_of_group, reffered_by

def sync_client_master_from_group_master(cursor):
    cursor.execute('''
        SELECT group_name, COALESCE(crp_name, '') AS crp_name, COALESCE(reffered_by, '') AS reffered_by
        FROM client_group_master
        WHERE group_name IS NOT NULL AND trim(group_name) != ''
    ''')
    updated_count = 0
    for row in cursor.fetchall():
        group_name = (row['group_name'] if isinstance(row, sqlite3.Row) else row[0]) or ''
        crp_name = (row['crp_name'] if isinstance(row, sqlite3.Row) else row[1]) or ''
        referred_by = (row['reffered_by'] if isinstance(row, sqlite3.Row) else row[2]) or ''
        cursor.execute('''
            UPDATE client_master
            SET client_group = ?, crp_of_group = ?, reffered_by = ?
            WHERE lower(trim(client_group)) = lower(trim(?))
              AND (
                  COALESCE(client_group, '') != ?
                  OR COALESCE(crp_of_group, '') != ?
                  OR COALESCE(reffered_by, '') != ?
              )
        ''', (group_name, crp_name, referred_by, group_name, group_name, crp_name, referred_by))
        updated_count += cursor.rowcount
    return updated_count

def get_debtor_users():
    try:
        conn = sqlite3.connect(MAIN_DB_PATH)
        conn.row_factory = sqlite3.Row
        users = conn.execute('''
            SELECT u.id, u.email,
                   CASE WHEN up.user_id IS NULL THEN 0 ELSE 1 END AS has_debtor_report
            FROM users u
            LEFT JOIN user_permissions up
              ON up.user_id = u.id
             AND up.permission_key = 'daily_debtor_report'
            ORDER BY u.email COLLATE NOCASE
        ''').fetchall()
        conn.close()
        return [dict(user) for user in users]
    except sqlite3.Error:
        return []

def get_all_debtor_nav_access():
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        ensure_debtor_nav_access_table(cursor)
        conn.commit()
        rows = cursor.execute('SELECT user_email, access_key FROM debtor_nav_access').fetchall()
        conn.close()
    except sqlite3.Error:
        return {}

    access = {}
    for row in rows:
        user_access = access.setdefault(str(row['user_email']).lower(), set())
        access_key = row['access_key']
        legacy_map = {
            'client_master': 'client_master_view',
            'client_group_master': 'client_group_master_view',
            'crp_master': 'crp_master_view',
            'firm_master': 'firm_master_view',
            'executive_partner_master': 'executive_partner_master_view',
        }
        access_key = legacy_map.get(access_key, access_key)
        user_access.add(access_key)
        if access_key in DEBTOR_NAV_EDIT_TO_VIEW:
            user_access.add(DEBTOR_NAV_EDIT_TO_VIEW[access_key])
    for user_access in access.values():
        if 'client_group_master_view' in user_access:
            user_access.add('crp_master_view')
        if 'client_group_master_edit' in user_access:
            user_access.update({'crp_master_view', 'crp_master_edit'})
    return access

def get_debtor_nav_access_for_user(user_email):
    user_email = str(user_email or '').lower()
    if user_email == 'arif.siddiqui@asija.in':
        return set(DEBTOR_NAV_ACCESS_KEYS)
    return get_all_debtor_nav_access().get(user_email, set())

def has_debtor_nav_access(user_email, access_key):
    return access_key in get_debtor_nav_access_for_user(user_email)

def get_debtor_access_key_for_path(path):
    path = path or '/'
    if path in ('/upload', '/import-followup-choice') or path == '/download/billing-template':
        return 'import_tally_debtors'
    if path == '/import-errors' or path == '/download/missing-clients-report':
        return 'import_tally_debtors'
    if path in ('/import-tally-receipts', '/tally-receipts/import-summary') or path == '/api/import-tally-receipts/preview':
        return 'import_tally_receipts_view'
    if path == '/import-tally-receipts/upload' or path == '/api/import-tally-receipts/post':
        return 'import_tally_receipts'
    if path == '/download/report-excel':
        return 'download_report_excel'
    if path == '/download/current-report-excel':
        return 'download_report_excel'
    if path in ('/report/update', '/report/add', '/report/delete'):
        return ''
    if path in ('/client-master', '/download/client-list'):
        return 'client_master_view'
    if path in ('/client-master/add', '/client-master/update', '/client-master/upload', '/download/client-template'):
        return 'client_master_edit'
    if path == '/client-group-master' or path == '/download/client-group-master-excel' or (path.startswith('/client-group-master/') and (path.endswith('/clients') or path.endswith('/clients-data') or path.endswith('/clients-excel'))):
        return 'client_group_master_view'
    if path in ('/client-group-master/add', '/client-group-master/update', '/client-group-master/delete'):
        return 'client_group_master_edit'
    if path == '/crp-master' or path == '/download/crp-master-excel' or (path.startswith('/crp-master/') and (path.endswith('/groups-data') or path.endswith('/groups-excel'))):
        return 'crp_master_view'
    if path in ('/crp-master/add', '/crp-master/update', '/crp-master/delete'):
        return 'crp_master_edit'
    if path == '/firm-master':
        return 'firm_master_view'
    if path in ('/firm-master/add', '/firm-master/update', '/firm-master/delete'):
        return 'firm_master_edit'
    if path == '/executive-partner-master':
        return 'executive_partner_master_view'
    if path == '/executive-partner-master/add':
        return 'executive_partner_master_edit'
    if path == '/report/run-followup-logic':
        return 'run_followup_logic'
    if path == '/clear':
        return 'clear_report'
    return ''

def get_manual_report_update_dates():
    defaults = {
        'report_as_on_date': '',
        'last_bill_update_date': '',
        'last_receipt_update_date': '',
    }
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        defaults['report_as_on_date'] = get_app_meta(cursor, 'manual_report_as_on_date')
        defaults['last_bill_update_date'] = get_app_meta(cursor, 'manual_last_bill_update_date')
        defaults['last_receipt_update_date'] = get_app_meta(cursor, 'manual_last_receipt_update_date')
        conn.close()
    except sqlite3.Error:
        pass
    return defaults

def get_report_as_on_lines():
    manual_dates = get_manual_report_update_dates()
    manual_report_as_on = manual_dates.get('report_as_on_date')
    manual_bill_date = manual_dates.get('last_bill_update_date')
    manual_receipt_date = manual_dates.get('last_receipt_update_date')
    if manual_report_as_on or manual_bill_date or manual_receipt_date:
        return [
            f"Report as on - {format_header_date(manual_report_as_on)}",
            f"Last Bill added - {format_header_date(manual_bill_date)}",
            f"Receipt Added - {format_header_date(manual_receipt_date)}",
        ]

    label = get_report_as_on_label()
    return [label] if label else []

def get_report_as_on_label():
    manual_dates = get_manual_report_update_dates()
    manual_report_as_on = manual_dates.get('report_as_on_date')
    manual_bill_date = manual_dates.get('last_bill_update_date')
    manual_receipt_date = manual_dates.get('last_receipt_update_date')
    if manual_report_as_on or manual_bill_date or manual_receipt_date:
        return (
            f"Report Update - Report as on - {format_header_date(manual_report_as_on)}, "
            f"Last Bill added - {format_header_date(manual_bill_date)}, "
            f"Receipt Added - {format_header_date(manual_receipt_date)}"
        )

    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT MAX(bill_date) AS latest_bill_date FROM billing_report')
        latest_row = cursor.fetchone()
        latest_bill_date = latest_row['latest_bill_date'] if latest_row else ''
        cursor.execute('SELECT MAX(receipt_date) AS latest_receipt_date FROM receipt_register')
        receipt_row = cursor.fetchone()
        latest_receipt_date = receipt_row['latest_receipt_date'] if receipt_row else ''
        last_receipt_activity_at = get_app_meta(cursor, 'last_receipt_activity_at')
        last_receipt_activity_type = get_app_meta(cursor, 'last_receipt_activity_type')
        conn.close()
    except sqlite3.Error:
        return ''

    labels = []
    if latest_bill_date:
        labels.append(f"Bill Report As On (Bill Import): {format_header_date(latest_bill_date)}")

    receipt_label_date = latest_receipt_date or last_receipt_activity_at
    if receipt_label_date:
        label_type = last_receipt_activity_type or 'Import/Manual'
        labels.append(f"Receipt Last {label_type}: {format_header_date(receipt_label_date)}")

    return ' | '.join(labels)

@app.before_request
def enforce_debtor_nav_access():
    if request.endpoint == 'static':
        return None
    user_email = str(session.get('user_email', '')).lower()
    if user_email == 'arif.siddiqui@asija.in':
        return None

    access_key = get_debtor_access_key_for_path(request.path)
    if access_key and not has_debtor_nav_access(user_email, access_key):
        flash('You do not have access to this debtor report option.')
        return redirect(url_for('dashboard'))
    return None

def read_backup_log():
    try:
        with open(BACKUP_LOG_PATH, 'r', encoding='utf-8') as file:
            log = json.load(file)
            return log if isinstance(log, list) else []
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def write_backup_log(log):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    with open(BACKUP_LOG_PATH, 'w', encoding='utf-8') as file:
        json.dump(log, file, indent=2)

def add_backup_log_entry(entry):
    log = read_backup_log()
    log.insert(0, entry)
    write_backup_log(log)

def make_backup_stamp():
    base_stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    stamp = base_stamp
    counter = 1
    while os.path.exists(os.path.join(BACKUP_DIR, f'backup_{stamp}.zip')):
        stamp = f'{base_stamp}_{counter}'
        counter += 1
    return stamp

def create_project_backup(note='', backup_type='manual'):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = make_backup_stamp()
    zip_name = f'backup_{stamp}.zip'
    zip_path = os.path.join(BACKUP_DIR, zip_name)
    temp_zip_path = os.path.join(BACKUP_DIR, f'.{zip_name}.tmp')
    skipped = []
    temp_db_copy = None

    try:
        with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as backup_zip:
            for current_dir, dirs, files in os.walk(BASE_DIR):
                dirs[:] = [name for name in dirs if name not in BACKUP_EXCLUDE_DIRS]
                for filename in files:
                    if filename in BACKUP_EXCLUDE_FILES or filename.startswith('~$'):
                        continue

                    file_path = os.path.join(current_dir, filename)
                    arcname = os.path.relpath(file_path, BASE_DIR)
                    try:
                        if os.path.abspath(file_path) == os.path.abspath(DB_PATH) and sqlite3.use_postgres(DB_PATH):
                            backup_zip.writestr(
                                'POSTGRES_DATABASE.txt',
                                'Debtor app is configured to use PostgreSQL via DEBTOR_DATABASE_URL. '
                                'Use pg_dump to back up the live debtor database.\n'
                            )
                        elif os.path.abspath(file_path) == os.path.abspath(DB_PATH):
                            with tempfile.NamedTemporaryFile(delete=False, suffix='.db', dir=BACKUP_DIR) as temp_file:
                                temp_db_copy = temp_file.name
                            source_conn = sqlite3.connect(DB_PATH)
                            backup_conn = sqlite3.connect(temp_db_copy)
                            try:
                                source_conn.backup(backup_conn)
                            finally:
                                backup_conn.close()
                                source_conn.close()
                            backup_zip.write(temp_db_copy, arcname)
                        else:
                            backup_zip.write(file_path, arcname)
                    except Exception as exc:
                        skipped.append({'file': arcname, 'reason': str(exc)})
                    finally:
                        if temp_db_copy and os.path.exists(temp_db_copy):
                            try:
                                os.remove(temp_db_copy)
                            except OSError:
                                pass
                            temp_db_copy = None
        os.replace(temp_zip_path, zip_path)
    except Exception:
        if os.path.exists(temp_zip_path):
            try:
                os.remove(temp_zip_path)
            except OSError:
                pass
        raise

    add_backup_log_entry({
        'id': stamp,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'type': backup_type,
        'note': note or 'Project backup',
        'file': zip_name,
        'skipped': skipped,
    })
    return stamp, skipped

def restore_project_backup(backup_id):
    log = read_backup_log()
    backup_entry = next((entry for entry in log if entry.get('id') == backup_id), None)
    if not backup_entry:
        return False, 'Backup not found.'

    zip_path = os.path.join(BACKUP_DIR, backup_entry.get('file', ''))
    if not os.path.exists(zip_path):
        return False, 'Backup file is missing.'

    create_project_backup(
        note=f'Automatic backup before rollback to {backup_entry.get("created_at", backup_id)}',
        backup_type='pre-rollback',
    )

    with tempfile.TemporaryDirectory(dir=BACKUP_DIR) as temp_dir:
        with zipfile.ZipFile(zip_path, 'r') as backup_zip:
            backup_zip.extractall(temp_dir)
            backup_members = {
                name.replace('/', os.sep)
                for name in backup_zip.namelist()
                if not name.endswith('/')
            }

        for current_dir, dirs, files in os.walk(BASE_DIR, topdown=False):
            relative_dir = os.path.relpath(current_dir, BASE_DIR)
            if relative_dir == '.':
                relative_dir = ''

            if relative_dir.split(os.sep)[0] in BACKUP_EXCLUDE_DIRS:
                continue

            for filename in files:
                if filename in BACKUP_EXCLUDE_FILES or filename.startswith('~$'):
                    continue

                file_path = os.path.join(current_dir, filename)
                relative_path = os.path.join(relative_dir, filename) if relative_dir else filename
                if relative_path not in backup_members:
                    try:
                        os.remove(file_path)
                    except OSError:
                        pass

            if relative_dir and relative_dir not in BACKUP_EXCLUDE_DIRS:
                try:
                    if not os.listdir(current_dir):
                        os.rmdir(current_dir)
                except OSError:
                    pass

        for current_dir, dirs, files in os.walk(temp_dir):
            for filename in files:
                source_path = os.path.join(current_dir, filename)
                relative_path = os.path.relpath(source_path, temp_dir)
                target_path = os.path.join(BASE_DIR, relative_path)
                os.makedirs(os.path.dirname(target_path), exist_ok=True)
                shutil.copy2(source_path, target_path)

    add_backup_log_entry({
        'id': datetime.now().strftime('%Y%m%d_%H%M%S'),
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'type': 'rollback',
        'note': f'Rolled back to backup {backup_id}',
        'file': backup_entry.get('file', ''),
        'skipped': [],
    })
    return True, 'Rollback completed. Restart the app if code files were restored.'

def get_short_name(name):
    """Converts a name like 'Asija And Associates LLP' to 'AAA LLP'."""
    parts = name.replace('.', '').split()
    if len(parts) == 1:
        return parts[0].upper()
    short = ""
    for p in parts:
        if p.upper() in ["LLP", "LTD"]:
            short += " " + p.upper()
        else:
            short += p[0].upper()
    return short.strip()

def get_firm_short_name(cursor, firm_name):
    firm_name = (firm_name or '').strip()
    if not firm_name:
        return ''

    cursor.execute(
        'SELECT short_name FROM firm_master WHERE lower(trim(firm_name)) = lower(trim(?)) LIMIT 1',
        (firm_name,)
    )
    row = cursor.fetchone()
    if row and row[0]:
        return row[0]
    return get_short_name(firm_name)

def ensure_firm_master(cursor, firm_name):
    firm_name = (firm_name or '').strip()
    if not firm_name or firm_name == 'No Data Found':
        return

    cursor.execute(
        'SELECT id FROM firm_master WHERE lower(trim(firm_name)) = lower(trim(?)) LIMIT 1',
        (firm_name,)
    )
    if cursor.fetchone():
        return

    cursor.execute(
        'INSERT INTO firm_master (firm_name, short_name) VALUES (?, ?)',
        (firm_name, get_short_name(firm_name))
    )

def format_display_date(value):
    if not value:
        return ''

    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime('%d-%m-%y')

    value = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y', '%d-%m-%y'):
        try:
            return datetime.strptime(value, fmt).strftime('%d-%m-%y')
        except ValueError:
            continue
    return value

def format_display_datetime(value):
    if not value:
        return ''

    if isinstance(value, (datetime, pd.Timestamp)):
        date_obj = value.to_pydatetime() if isinstance(value, pd.Timestamp) else value
        return date_obj.strftime('%d-%m-%y %H:%M')

    value = str(value).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%d-%m-%Y %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
        try:
            return datetime.strptime(value, fmt).strftime('%d-%m-%y %H:%M')
        except ValueError:
            continue
    return value

def format_header_date(value):
    date_obj = parse_input_date(value)
    if date_obj:
        return date_obj.strftime('%d.%m.%Y')
    return format_display_date(value)

def get_current_fy_string():
    """Dynamically determines the current FY string based on today's date."""
    now = datetime.now()
    start_year = now.year if (now.month, now.day) >= (4, 2) else now.year - 1
    return f'02.04.{start_year} - 01.04.{start_year + 1}'

def parse_input_date(value):
    if not value:
        return None

    if isinstance(value, (datetime, pd.Timestamp)):
        return value.to_pydatetime() if isinstance(value, pd.Timestamp) else value

    value = str(value).strip()
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%y', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%b-%y', '%d-%b-%Y'):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None

def format_financial_year(value):
    bill_date = parse_input_date(value)
    if not bill_date:
        return ''

    start_year = bill_date.year if (bill_date.month, bill_date.day) >= (4, 2) else bill_date.year - 1
    return f'02.04.{start_year} - 01.04.{start_year + 1}'

def is_blank(value):
    return value is None or pd.isna(value) or str(value).strip() == ''

def first_nonblank(values):
    for value in values:
        if not is_blank(value):
            return str(value).strip()
    return ''

def normalize_tally_cell(value):
    if value is None:
        return ''
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()

def parse_tally_receipt_trial(search_date='', file_path=None, file_name=None):
    """Preview-only parser for Tally receipt Excel files. It does not write to DB."""
    receipt_path = file_path or TALLY_RECEIPT_TRIAL_PATH
    receipt_name = file_name or (os.path.basename(receipt_path) if isinstance(receipt_path, (str, bytes, os.PathLike)) else 'fcRece.xlsx')
    if isinstance(receipt_path, (str, bytes, os.PathLike)) and not os.path.exists(receipt_path):
        raise FileNotFoundError('fcRece.xlsx file not found.')

    requested_date = parse_input_date(search_date) if search_date else None
    workbook = load_workbook(receipt_path, data_only=True, read_only=True)
    worksheet = None
    rows = []
    header_index = None

    for candidate in workbook.worksheets:
        candidate_rows = [tuple(row) for row in candidate.iter_rows(values_only=True)]
        candidate_header_index = None
        for index, row in enumerate(candidate_rows):
            if any(str(cell).strip().lower() == 'date' for cell in row if cell is not None):
                candidate_header_index = index
                break
        if candidate_header_index is not None:
            worksheet = candidate
            rows = candidate_rows
            header_index = candidate_header_index
            break

    if worksheet is None:
        worksheet = workbook[workbook.sheetnames[0]]

    if header_index is None:
        workbook.close()
        return {
            'file_name': receipt_name,
            'sheet_name': worksheet.title,
            'header_row': None,
            'data_start_row': None,
            'rows': [],
            'message': 'Date header not found in Tally receipt file.'
        }

    parsed_rows = []
    current_voucher = None
    current_ledger = None
    data_start_index = header_index + 1

    for zero_index, row in enumerate(rows[data_start_index:], start=data_start_index):
        cells = list(row) + [None] * max(0, 10 - len(row))
        first_cell_text = normalize_tally_cell(cells[0])

        if first_cell_text.lower().startswith('total'):
            break

        row_date = parse_input_date(cells[0])
        if row_date:
            current_voucher = {
                'excel_row': zero_index + 1,
                'receipt_date': row_date.strftime('%Y-%m-%d'),
                'receipt_date_display': format_display_date(row_date),
                'voucher_type': normalize_tally_cell(cells[6]),
                'voucher_no': normalize_tally_cell(cells[7]),
                'debit_amount': float(cells[8] or 0) if str(cells[8] or '').strip() else 0,
                'bank_name': '',
                'items': [],
            }
            current_ledger = {
                'party_name': normalize_tally_cell(cells[1]),
                'credit_amount': float(cells[9] or 0) if str(cells[9] or '').strip() else 0,
                'raw_rows': [zero_index + 1],
            }
            continue

        if current_voucher is None:
            continue

        label = normalize_tally_cell(cells[1]).lower()
        if label in {'agst ref', 'new ref'}:
            receipt_item = {
                'excel_row': zero_index + 1,
                'receipt_date': current_voucher['receipt_date'],
                'receipt_date_display': current_voucher['receipt_date_display'],
                'party_name': (current_ledger or {}).get('party_name', ''),
                'voucher_type': current_voucher['voucher_type'],
                'voucher_no': current_voucher['voucher_no'],
                'debit_amount': current_voucher['debit_amount'],
                'credit_amount': (current_ledger or {}).get('credit_amount', 0),
                'reference_type': normalize_tally_cell(cells[1]),
                'reference_no': normalize_tally_cell(cells[2]),
                'credit_days': normalize_tally_cell(cells[3]),
                'reference_amount': float(cells[4] or 0) if str(cells[4] or '').strip() else 0,
                'reference_dr_cr': normalize_tally_cell(cells[5]),
                'bank_name': current_voucher['bank_name'],
                'raw_rows': list((current_ledger or {}).get('raw_rows', [])) + [zero_index + 1],
            }
            current_voucher['items'].append(receipt_item)
            parsed_rows.append(receipt_item)
        elif normalize_tally_cell(cells[1]):
            credit_amount = float(cells[9] or 0) if str(cells[9] or '').strip() else 0
            if credit_amount > 0:
                current_ledger = {
                    'party_name': normalize_tally_cell(cells[1]),
                    'credit_amount': credit_amount,
                    'raw_rows': [zero_index + 1],
                }
                continue

            current_voucher['bank_name'] = normalize_tally_cell(cells[1])
            if str(cells[8] or '').strip():
                current_voucher['debit_amount'] = float(cells[8] or 0)
            for receipt_item in current_voucher['items']:
                receipt_item['bank_name'] = current_voucher['bank_name']
                receipt_item['debit_amount'] = current_voucher['debit_amount']
                receipt_item['raw_rows'].append(zero_index + 1)

    if requested_date:
        requested = requested_date.strftime('%Y-%m-%d')
        parsed_rows = [row for row in parsed_rows if row['receipt_date'] == requested]

    result = {
        'file_name': receipt_name,
        'sheet_name': worksheet.title,
        'header_row': header_index + 1,
        'data_start_row': data_start_index + 1,
        'rows': parsed_rows,
        'message': 'Preview only. No receipt data has been posted or saved.'
    }
    workbook.close()
    return result

def normalize_reference_key(value):
    return ''.join(ch for ch in str(value or '').upper() if ch.isalnum())

def get_tally_receipt_amount(row):
    reference_amount = float(row.get('reference_amount') or 0)
    credit_amount = float(row.get('credit_amount') or 0)
    debit_amount = float(row.get('debit_amount') or 0)
    if reference_amount > 0:
        return reference_amount
    if credit_amount > 0:
        return credit_amount
    return debit_amount

def ensure_receipt_register_tally_columns(cursor):
    cursor.execute('PRAGMA table_info(receipt_register)')
    existing_columns = {row[1] for row in cursor.fetchall()}
    required_columns = {
        'import_source': 'TEXT',
        'tally_voucher_no': 'TEXT',
        'tally_reference_no': 'TEXT',
        'tally_bank_name': 'TEXT',
        'tally_excel_rows': 'TEXT',
    }
    for column_name, column_type in required_columns.items():
        if column_name not in existing_columns:
            cursor.execute(f'ALTER TABLE receipt_register ADD COLUMN {column_name} {column_type}')

def build_current_bill_matches(cursor):
    matches = {}
    for row in get_report_rows(cursor):
        key = normalize_reference_key(row.get('ref_no'))
        if not key:
            continue
        matches.setdefault(key, []).append(row)
    return matches

def enrich_tally_receipt_rows(tally_rows, cursor):
    bill_matches = build_current_bill_matches(cursor)
    cursor.execute('''
        SELECT tally_voucher_no, tally_reference_no
        FROM receipt_register
        WHERE import_source = 'Tally'
          AND tally_voucher_no IS NOT NULL AND tally_reference_no IS NOT NULL
    ''')
    posted_keys = {
        (row['tally_voucher_no'] or '', normalize_reference_key(row['tally_reference_no']))
        for row in cursor.fetchall()
    }
    cursor.execute('''
        SELECT id, ref_no, tally_reference_no, party_name, received_amount, receipt_date,
               tally_voucher_no
        FROM receipt_register
        WHERE (ref_no IS NOT NULL AND TRIM(ref_no) != '')
           OR (tally_reference_no IS NOT NULL AND TRIM(tally_reference_no) != '')
    ''')
    posted_receipts_by_ref = {}
    for row in cursor.fetchall():
        for ref_value in (row['ref_no'], row['tally_reference_no']):
            key = normalize_reference_key(ref_value)
            if key:
                posted_receipts_by_ref.setdefault(key, row)

    enriched_rows = []
    seen_preview_keys = set()
    for row in tally_rows:
        item = row.to_dict() if hasattr(row, 'to_dict') else dict(row)
        amount = round(get_tally_receipt_amount(item), 2)
        ref_key = normalize_reference_key(item.get('reference_no'))
        voucher_no = item.get('voucher_no') or ''
        candidates = bill_matches.get(ref_key, [])
        posted_key = (voucher_no, ref_key)

        item['post_amount'] = amount
        item['matched_bill_id'] = ''
        item['matched_ref_no'] = ''
        item['matched_party_name'] = ''
        item['matched_bill_amount'] = 0
        item['post_status'] = 'Not Matched'
        item['post_message'] = 'No open bill found for reference number.'
        item['is_postable'] = False
        item['already_posted_from_receipt_register'] = False

        if posted_key in seen_preview_keys:
            item['post_status'] = 'Duplicate Preview'
            item['post_message'] = 'This voucher/reference appears more than once in the preview.'
        elif posted_key in posted_keys:
            item['post_status'] = 'Already Posted'
            item['post_message'] = 'This Tally voucher/reference is already in receipt register.'
            item['already_posted_from_receipt_register'] = True
            posted_receipt = posted_receipts_by_ref.get(ref_key)
            if posted_receipt:
                item['matched_ref_no'] = posted_receipt['ref_no'] or posted_receipt['tally_reference_no'] or ''
                item['matched_party_name'] = posted_receipt['party_name'] or ''
                item['matched_bill_amount'] = round(float(posted_receipt['received_amount'] or 0), 2)
        elif ref_key in posted_receipts_by_ref:
            posted_receipt = posted_receipts_by_ref[ref_key]
            item['post_status'] = 'Already Posted'
            item['post_message'] = 'This reference number is already available in receipt register.'
            item['matched_ref_no'] = posted_receipt['ref_no'] or posted_receipt['tally_reference_no'] or ''
            item['matched_party_name'] = posted_receipt['party_name'] or ''
            item['matched_bill_amount'] = round(float(posted_receipt['received_amount'] or 0), 2)
            item['already_posted_from_receipt_register'] = True
        elif not ref_key:
            item['post_status'] = 'Missing Ref'
            item['post_message'] = 'Reference number is blank.'
        elif len(candidates) > 1:
            item['post_status'] = 'Multiple Match'
            item['post_message'] = 'More than one open bill matched this reference.'
        elif len(candidates) == 1:
            bill = candidates[0]
            bill_amount = round(float(bill.get('amount') or 0), 2)
            item['matched_bill_id'] = bill.get('id')
            item['matched_ref_no'] = bill.get('ref_no') or ''
            item['matched_party_name'] = bill.get('party_name') or ''
            item['matched_bill_amount'] = bill_amount
            if amount <= 0:
                item['post_status'] = 'Invalid Amount'
                item['post_message'] = 'Receipt amount is zero or blank.'
            elif amount > bill_amount:
                item['post_status'] = 'Excess Amount'
                item['post_message'] = 'Receipt amount is greater than current bill balance.'
            else:
                item['post_status'] = 'Ready'
                item['post_message'] = 'Ready to post.'
                item['is_postable'] = True

        if ref_key:
            seen_preview_keys.add(posted_key)
        enriched_rows.append(item)

    return enriched_rows

def post_receipt_rows(cursor, receipt_rows, receipt_mode, receipt_date, posted_at, import_meta_by_bill=None):
    import_meta_by_bill = import_meta_by_bill or {}
    current_rows = {int(row['id']): row for row in get_report_rows(cursor)}

    posted_count = 0
    posted_total = 0
    for item in receipt_rows:
        try:
            bill_id = int(item.get('row_id'))
            received_amount = float(item.get('received_amount'))
        except (TypeError, ValueError):
            raise ValueError('Please check actual received amount.')

        if received_amount <= 0:
            raise ValueError('Actual received amount must be greater than zero.')

        row = current_rows.get(bill_id)
        if not row:
            raise ValueError('One selected bill is no longer available in the debtor report.')

        bill_amount = float(row.get('amount') or 0)
        if received_amount > bill_amount:
            raise ValueError('Actual received amount cannot be greater than bill amount.')

        balance_amount = round(bill_amount - received_amount, 2)
        existing_paid_amount = float(row.get('paid_amount') or 0)
        total_paid_amount = round(existing_paid_amount + received_amount, 2)
        meta = import_meta_by_bill.get(bill_id, {})

        cursor.execute('''
            INSERT INTO receipt_register (
                source_bill_id, firm_name, short_name, bill_date, ref_no, party_name,
                bill_amount, received_amount, balance_amount, due_date, overdue_days,
                client_group, followup_partner, final_ep, crp_of_group, client_category,
                financial_year, receipt_mode, receipt_date, posted_at,
                import_source, tally_voucher_no, tally_reference_no, tally_bank_name, tally_excel_rows
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            bill_id,
            row.get('firm_name') or '',
            row.get('short_name') or '',
            row.get('bill_date') or '',
            row.get('ref_no') or '',
            row.get('party_name') or '',
            bill_amount,
            received_amount,
            balance_amount,
            row.get('due_date') or '',
            row.get('overdue_days') or 0,
            row.get('client_group') or '',
            row.get('followup_partner') or '',
            row.get('final_ep') or '',
            row.get('crp_of_group') or '',
            row.get('client_category') or '',
            row.get('financial_year') or '',
            receipt_mode,
            receipt_date,
            posted_at,
            meta.get('import_source') or '',
            meta.get('tally_voucher_no') or '',
            meta.get('tally_reference_no') or '',
            meta.get('tally_bank_name') or '',
            meta.get('tally_excel_rows') or '',
        ))

        if balance_amount <= 0:
            cursor.execute('''
                UPDATE billing_report
                SET receipt_status = ?, paid_amount = ?, closed_at = ?
                WHERE id = ?
            ''', ('full_paid', total_paid_amount, posted_at, bill_id))
        else:
            cursor.execute('''
                UPDATE billing_report
                SET amount = ?, receipt_status = ?, paid_amount = ?, closed_at = NULL
                WHERE id = ?
            ''', (balance_amount, 'part_paid', total_paid_amount, bill_id))

        current_rows[bill_id]['amount'] = balance_amount
        current_rows[bill_id]['paid_amount'] = total_paid_amount
        posted_count += 1
        posted_total += received_amount

    return posted_count, posted_total

def post_adjustment_rows(cursor, adjustment_rows, adjustment_type, adjustment_date, posted_at):
    current_rows = {int(row['id']): row for row in get_report_rows(cursor)}

    posted_count = 0
    posted_total = 0
    for item in adjustment_rows:
        try:
            bill_id = int(item.get('row_id'))
            adjustment_amount = float(item.get('received_amount'))
        except (TypeError, ValueError):
            raise ValueError('Please check adjustment amount.')

        if adjustment_amount <= 0:
            raise ValueError('Adjustment amount must be greater than zero.')

        row = current_rows.get(bill_id)
        if not row:
            raise ValueError('One selected bill is no longer available in the debtor report.')

        bill_amount = float(row.get('amount') or 0)
        if adjustment_amount > bill_amount:
            raise ValueError('Adjustment amount cannot be greater than bill amount.')

        balance_amount = round(bill_amount - adjustment_amount, 2)
        existing_paid_amount = float(row.get('paid_amount') or 0)
        total_paid_amount = round(existing_paid_amount + adjustment_amount, 2)

        cursor.execute('''
            INSERT INTO receipt_adjustment_register (
                source_bill_id, firm_name, short_name, bill_date, ref_no, party_name,
                bill_amount, adjustment_amount, balance_amount, due_date, overdue_days,
                client_group, followup_partner, final_ep, crp_of_group, client_category,
                financial_year, adjustment_type, adjustment_date, posted_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            bill_id,
            row.get('firm_name') or '',
            row.get('short_name') or '',
            row.get('bill_date') or '',
            row.get('ref_no') or '',
            row.get('party_name') or '',
            bill_amount,
            adjustment_amount,
            balance_amount,
            row.get('due_date') or '',
            row.get('overdue_days') or 0,
            row.get('client_group') or '',
            row.get('followup_partner') or '',
            row.get('final_ep') or '',
            row.get('crp_of_group') or '',
            row.get('client_category') or '',
            row.get('financial_year') or '',
            adjustment_type,
            adjustment_date,
            posted_at,
        ))

        if balance_amount <= 0:
            cursor.execute('''
                UPDATE billing_report
                SET receipt_status = ?, paid_amount = ?, closed_at = ?
                WHERE id = ?
            ''', ('full_paid', total_paid_amount, posted_at, bill_id))
        else:
            cursor.execute('''
                UPDATE billing_report
                SET amount = ?, receipt_status = ?, paid_amount = ?, closed_at = NULL
                WHERE id = ?
            ''', (balance_amount, 'part_paid', total_paid_amount, bill_id))

        current_rows[bill_id]['amount'] = balance_amount
        current_rows[bill_id]['paid_amount'] = total_paid_amount
        posted_count += 1
        posted_total += adjustment_amount

    return posted_count, posted_total

def normalize_cell(value):
    if is_blank(value):
        return ''
    return ''.join(ch for ch in str(value).lower() if ch.isalnum())

def to_float(value):
    if is_blank(value):
        return None
    try:
        return float(str(value).replace(',', '').strip())
    except (ValueError, TypeError):
        return None

def insert_billing_row(cursor, firm_name, bill_date_obj, ref_no, party_name, amount, due_date_obj=None, overdue_days=None, import_batch_id=''):
    if not firm_name or not bill_date_obj or not ref_no or not party_name or amount is None:
        return False

    ensure_firm_master(cursor, firm_name)
    due_date_obj = bill_date_obj + timedelta(days=30)
    if overdue_days is None:
        overdue_days = (datetime.now() - bill_date_obj).days

    cursor.execute('''
        INSERT INTO billing_report 
        (firm_name, short_name, bill_date, ref_no, party_name, amount, due_date, overdue_days, import_batch_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        firm_name,
        get_firm_short_name(cursor, firm_name),
        bill_date_obj.strftime('%Y-%m-%d'),
        str(ref_no).strip(),
        str(party_name).strip(),
        amount,
        due_date_obj.strftime('%Y-%m-%d'),
        int(overdue_days) if overdue_days is not None else 0,
        import_batch_id,
    ))
    return True

def client_exists_in_master(cursor, party_name):
    cursor.execute('''
        SELECT 1
        FROM client_master
        WHERE lower(trim(client_name)) = lower(trim(?))
        LIMIT 1
    ''', (party_name,))
    return cursor.fetchone() is not None

def save_import_client_errors(cursor, batch_id, imported_rows):
    imported_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    grouped = {}

    for row in imported_rows:
        party_name = (row.get('party_name') or '').strip()
        if not party_name or client_exists_in_master(cursor, party_name):
            continue

        key = party_name.lower()
        item = grouped.setdefault(key, {
            'party_name': party_name,
            'firm_names': set(),
            'ref_nos': [],
            'bill_count': 0,
            'total_amount': 0,
        })
        firm_name = (row.get('firm_name') or '').strip()
        if firm_name:
            item['firm_names'].add(firm_name)
        ref_no = (row.get('ref_no') or '').strip()
        if ref_no and len(item['ref_nos']) < 5:
            item['ref_nos'].append(ref_no)
        item['bill_count'] += 1
        item['total_amount'] += float(row.get('amount') or 0)

    for item in grouped.values():
        cursor.execute('''
            INSERT INTO debtor_import_client_errors
            (batch_id, imported_at, party_name, firm_names, bill_count, total_amount, sample_ref_nos)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            batch_id,
            imported_at,
            item['party_name'],
            ', '.join(sorted(item['firm_names'])),
            item['bill_count'],
            item['total_amount'],
            ', '.join(item['ref_nos']),
        ))

    set_app_meta(cursor, 'last_debtor_import_batch_id', batch_id)
    return len(grouped)

def normalize_ep_code(value):
    return re.sub(r'[^a-z0-9]', '', str(value or '').lower())

def load_executive_partner_lookup(cursor):
    cursor.execute('''
        SELECT partner_name, COALESCE(NULLIF(final_ep, ''), partner_name) AS final_ep
        FROM executive_partner_master
        WHERE partner_name IS NOT NULL AND trim(partner_name) != ''
    ''')
    lookup = {}
    for row in cursor.fetchall():
        partner_name = row['partner_name'] if isinstance(row, sqlite3.Row) else row[0]
        final_ep = row['final_ep'] if isinstance(row, sqlite3.Row) else row[1]
        code = normalize_ep_code(partner_name)
        if code:
            lookup[code] = (final_ep or partner_name or '').strip()
    return lookup

def infer_final_ep_from_ref(ref_no, ep_lookup):
    tokens = [
        normalize_ep_code(token)
        for token in re.split(r'[^A-Za-z0-9]+', str(ref_no or ''))
    ]
    for token in tokens:
        if token in ep_lookup:
            return ep_lookup[token]
    return ''

def run_followup_partner_logic(cursor):
    cursor.execute('''
        SELECT client_group, crp_of_group, COUNT(*) AS use_count
        FROM client_master
        WHERE client_group IS NOT NULL AND trim(client_group) != ''
          AND crp_of_group IS NOT NULL AND trim(crp_of_group) != ''
        GROUP BY client_group, crp_of_group
        ORDER BY client_group, use_count DESC, crp_of_group
    ''')
    crp_by_group = {}
    for row in cursor.fetchall():
        crp_by_group.setdefault(row['client_group'] if isinstance(row, sqlite3.Row) else row[0], row['crp_of_group'] if isinstance(row, sqlite3.Row) else row[1])
    ep_lookup = load_executive_partner_lookup(cursor)

    cursor.execute('''
        SELECT
            billing_report.id,
            billing_report.party_name,
            billing_report.amount,
            billing_report.ref_no,
            billing_report.ep_override,
            COALESCE(NULLIF(billing_report.group_override, ''), client_master.client_group) AS client_group,
            client_master.crp_of_group
        FROM billing_report
        LEFT JOIN client_master
            ON lower(trim(billing_report.party_name)) = lower(trim(client_master.client_name))
        WHERE COALESCE(billing_report.receipt_status, 'open') != 'full_paid'
          AND billing_report.deleted_at IS NULL
    ''')
    columns = [description[0] for description in cursor.description]
    rows = [
        dict(row) if isinstance(row, sqlite3.Row) else dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

    for row in rows:
        group_name = (row.get('client_group') or '').strip()
        if group_name and crp_by_group.get(group_name):
            row['crp_of_group'] = crp_by_group[group_name]
        row['final_ep'] = (row.get('ep_override') or '').strip() or infer_final_ep_from_ref(row.get('ref_no'), ep_lookup)

    group_data = {}
    for row in rows:
        group_name = (row.get('client_group') or row.get('party_name') or '').strip()
        ep_name = (row.get('final_ep') or '').strip()
        crp_name = (row.get('crp_of_group') or '').strip()
        amount = row.get('amount') or 0

        if not group_name:
            continue

        if group_name not in group_data:
            group_data[group_name] = {
                'crp': crp_name,
                'ep_totals': {},
            }
        elif not group_data[group_name]['crp'] and crp_name:
            group_data[group_name]['crp'] = crp_name

        if ep_name:
            group_data[group_name]['ep_totals'][ep_name] = (
                group_data[group_name]['ep_totals'].get(ep_name, 0) + amount
            )

    followup_by_group = {}
    for group_name, data in group_data.items():
        crp_name = data['crp']
        ep_totals = data['ep_totals']
        ep_names = set(ep_totals)
        non_crp_ep_totals = {
            ep_name: total
            for ep_name, total in ep_totals.items()
            if ep_name != crp_name
        }

        if not ep_names:
            followup_by_group[group_name] = crp_name
        elif crp_name and crp_name in ep_names:
            followup_by_group[group_name] = crp_name
        elif len(non_crp_ep_totals) == 1:
            followup_by_group[group_name] = next(iter(non_crp_ep_totals))
        elif non_crp_ep_totals:
            followup_by_group[group_name] = max(non_crp_ep_totals, key=non_crp_ep_totals.get)
        else:
            followup_by_group[group_name] = crp_name

    updated_count = 0
    for row in rows:
        group_name = (row.get('client_group') or row.get('party_name') or '').strip()
        followup_partner = followup_by_group.get(group_name, '')
        cursor.execute(
            'UPDATE billing_report SET followup_partner = ? WHERE id = ?',
            (followup_partner, row.get('id'))
        )
        updated_count += cursor.rowcount

    set_app_meta(cursor, 'last_followup_logic_run_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    return updated_count

def resolve_followup_partner_for_row(cursor, billing_row_id):
    rows = get_report_rows(cursor)
    target_row = next((row for row in rows if int(row.get('id') or 0) == int(billing_row_id)), None)
    if not target_row:
        return ''

    target_group = (target_row.get('client_group') or target_row.get('party_name') or '').strip()
    if not target_group:
        return ''

    group_rows = [
        row for row in rows
        if (row.get('client_group') or row.get('party_name') or '').strip() == target_group
    ]
    crp_name = ''
    ep_totals = {}
    for row in group_rows:
        if not crp_name and (row.get('crp_of_group') or '').strip():
            crp_name = (row.get('crp_of_group') or '').strip()
        ep_name = (row.get('final_ep') or '').strip()
        if ep_name:
            ep_totals[ep_name] = ep_totals.get(ep_name, 0) + float(row.get('amount') or 0)

    ep_names = set(ep_totals)
    non_crp_ep_totals = {
        ep_name: total
        for ep_name, total in ep_totals.items()
        if ep_name != crp_name
    }

    if not ep_names:
        return crp_name
    if crp_name and crp_name in ep_names:
        return crp_name
    if len(non_crp_ep_totals) == 1:
        return next(iter(non_crp_ep_totals))
    if non_crp_ep_totals:
        return max(non_crp_ep_totals, key=non_crp_ep_totals.get)
    return crp_name

def get_existing_followup_partner_for_row(cursor, billing_row_id):
    cursor.execute('''
        SELECT
            COALESCE(NULLIF(billing_report.group_override, ''), client_master.client_group, billing_report.party_name) AS group_name
        FROM billing_report
        LEFT JOIN client_master
            ON lower(trim(billing_report.party_name)) = lower(trim(client_master.client_name))
        WHERE billing_report.id = ?
        LIMIT 1
    ''', (billing_row_id,))
    target_row = cursor.fetchone()
    if not target_row:
        return ''

    target_group = (target_row['group_name'] if isinstance(target_row, sqlite3.Row) else target_row[0]) or ''
    target_group = target_group.strip()
    if not target_group:
        return ''

    cursor.execute('''
        SELECT billing_report.followup_partner, COUNT(*) AS use_count
        FROM billing_report
        LEFT JOIN client_master
            ON lower(trim(billing_report.party_name)) = lower(trim(client_master.client_name))
        WHERE billing_report.deleted_at IS NULL
          AND COALESCE(billing_report.receipt_status, 'open') != 'full_paid'
          AND billing_report.id != ?
          AND COALESCE(billing_report.followup_partner, '') != ''
          AND COALESCE(NULLIF(billing_report.group_override, ''), client_master.client_group, billing_report.party_name) = ?
        GROUP BY billing_report.followup_partner
        ORDER BY use_count DESC, billing_report.followup_partner
        LIMIT 1
    ''', (billing_row_id, target_group))
    existing_row = cursor.fetchone()
    if not existing_row:
        return ''

    followup_partner = existing_row['followup_partner'] if isinstance(existing_row, sqlite3.Row) else existing_row[0]
    return (followup_partner or '').strip()

def resolve_or_copy_followup_partner_for_row(cursor, billing_row_id):
    return (
        get_existing_followup_partner_for_row(cursor, billing_row_id)
        or resolve_followup_partner_for_row(cursor, billing_row_id)
    )

def ensure_cheque_bounce_register_table(cursor):
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS cheque_bounce_register (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            receipt_register_id INTEGER,
            source_bill_id INTEGER,
            readded_bill_id INTEGER,
            bounced_at TEXT,
            bounce_date TEXT,
            firm_name TEXT,
            short_name TEXT,
            bill_date TEXT,
            ref_no TEXT,
            party_name TEXT,
            bill_amount REAL,
            bounced_amount REAL,
            receipt_date TEXT,
            receipt_mode TEXT,
            posted_at TEXT,
            due_date TEXT,
            overdue_days INTEGER,
            client_group TEXT,
            followup_partner TEXT,
            final_ep TEXT,
            crp_of_group TEXT,
            client_category TEXT,
            financial_year TEXT,
            reason TEXT
        )
    ''')
    cursor.execute('PRAGMA table_info(cheque_bounce_register)')
    existing_columns = {row[1] for row in cursor.fetchall()}
    if 'bounce_date' not in existing_columns:
        cursor.execute('ALTER TABLE cheque_bounce_register ADD COLUMN bounce_date TEXT')

def readd_receipts_as_cheque_bounce(cursor, receipt_ids, reason='', bounce_date=''):
    ensure_cheque_bounce_register_table(cursor)
    clean_ids = []
    for receipt_id in receipt_ids:
        try:
            clean_ids.append(int(receipt_id))
        except (TypeError, ValueError):
            continue

    if not clean_ids:
        return {'added_count': 0, 'skipped_count': 0}

    added_count = 0
    skipped_count = 0
    bounced_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    bounce_date = (bounce_date or datetime.now().strftime('%Y-%m-%d')).strip()
    import_batch_id = 'cheque_bounce_' + datetime.now().strftime('%Y%m%d%H%M%S')
    placeholders = ','.join('?' for _ in clean_ids)
    cursor.execute(f'''
        SELECT *
        FROM receipt_register
        WHERE id IN ({placeholders})
        ORDER BY receipt_date DESC, id DESC
    ''', clean_ids)
    receipt_rows = [dict(row) for row in cursor.fetchall()]

    for row in receipt_rows:
        receipt_id = int(row.get('id') or 0)
        cursor.execute(
            'SELECT id FROM cheque_bounce_register WHERE receipt_register_id = ? LIMIT 1',
            (receipt_id,)
        )
        if cursor.fetchone():
            skipped_count += 1
            continue

        bounced_amount = float(row.get('received_amount') or 0)
        if bounced_amount <= 0:
            skipped_count += 1
            continue

        bill_date = row.get('bill_date') or ''
        due_date = row.get('due_date') or ''
        if not due_date:
            bill_date_obj = parse_input_date(bill_date)
            due_date = (bill_date_obj + timedelta(days=30)).strftime('%Y-%m-%d') if bill_date_obj else ''

        cursor.execute('''
            INSERT INTO billing_report (
                firm_name, short_name, bill_date, ref_no, party_name, amount,
                due_date, overdue_days, import_batch_id, ep_override, receipt_status,
                paid_amount, closed_at, group_override, followup_partner
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'open', 0, NULL, ?, '')
        ''', (
            row.get('firm_name') or '',
            row.get('short_name') or '',
            bill_date,
            row.get('ref_no') or '',
            row.get('party_name') or '',
            bounced_amount,
            due_date,
            row.get('overdue_days') or 0,
            import_batch_id,
            row.get('final_ep') or '',
            row.get('client_group') or '',
        ))
        readded_bill_id = cursor.lastrowid
        followup_partner = resolve_or_copy_followup_partner_for_row(cursor, readded_bill_id)
        cursor.execute(
            'UPDATE billing_report SET followup_partner = ? WHERE id = ?',
            (followup_partner, readded_bill_id)
        )

        cursor.execute('''
            INSERT INTO cheque_bounce_register (
                receipt_register_id, source_bill_id, readded_bill_id, bounced_at, bounce_date,
                firm_name, short_name, bill_date, ref_no, party_name, bill_amount,
                bounced_amount, receipt_date, receipt_mode, posted_at, due_date,
                overdue_days, client_group, followup_partner, final_ep, crp_of_group,
                client_category, financial_year, reason
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            receipt_id,
            row.get('source_bill_id'),
            readded_bill_id,
            bounced_at,
            bounce_date,
            row.get('firm_name') or '',
            row.get('short_name') or '',
            bill_date,
            row.get('ref_no') or '',
            row.get('party_name') or '',
            row.get('bill_amount') or 0,
            bounced_amount,
            row.get('receipt_date') or '',
            row.get('receipt_mode') or '',
            row.get('posted_at') or '',
            due_date,
            row.get('overdue_days') or 0,
            row.get('client_group') or '',
            followup_partner,
            row.get('final_ep') or '',
            row.get('crp_of_group') or '',
            row.get('client_category') or '',
            row.get('financial_year') or '',
            reason,
        ))
        added_count += 1

    return {'added_count': added_count, 'skipped_count': skipped_count}

def copy_existing_followup_for_import_batch(cursor, batch_id):
    if not batch_id:
        return 0

    cursor.execute('''
        SELECT
            COALESCE(NULLIF(billing_report.group_override, ''), client_master.client_group, billing_report.party_name) AS group_name,
            billing_report.followup_partner,
            COUNT(*) AS use_count
        FROM billing_report
        LEFT JOIN client_master
            ON lower(trim(billing_report.party_name)) = lower(trim(client_master.client_name))
        WHERE billing_report.deleted_at IS NULL
          AND COALESCE(billing_report.receipt_status, 'open') != 'full_paid'
          AND COALESCE(billing_report.followup_partner, '') != ''
          AND COALESCE(billing_report.import_batch_id, '') != ?
        GROUP BY group_name, billing_report.followup_partner
        ORDER BY group_name, use_count DESC, billing_report.followup_partner
    ''', (batch_id,))
    followup_by_group = {}
    for row in cursor.fetchall():
        group_name = (row['group_name'] if isinstance(row, sqlite3.Row) else row[0]) or ''
        followup_partner = (row['followup_partner'] if isinstance(row, sqlite3.Row) else row[1]) or ''
        group_name = group_name.strip()
        followup_partner = followup_partner.strip()
        if group_name and followup_partner and group_name not in followup_by_group:
            followup_by_group[group_name] = followup_partner

    cursor.execute('''
        SELECT
            billing_report.id,
            COALESCE(NULLIF(billing_report.group_override, ''), client_master.client_group, billing_report.party_name) AS group_name
        FROM billing_report
        LEFT JOIN client_master
            ON lower(trim(billing_report.party_name)) = lower(trim(client_master.client_name))
        WHERE COALESCE(billing_report.import_batch_id, '') = ?
          AND billing_report.deleted_at IS NULL
          AND COALESCE(billing_report.receipt_status, 'open') != 'full_paid'
    ''', (batch_id,))
    rows = cursor.fetchall()

    updated_count = 0
    for row in rows:
        row_id = row['id'] if isinstance(row, sqlite3.Row) else row[0]
        group_name = ((row['group_name'] if isinstance(row, sqlite3.Row) else row[1]) or '').strip()
        followup_partner = followup_by_group.get(group_name)
        if not followup_partner:
            continue
        cursor.execute(
            'UPDATE billing_report SET followup_partner = ? WHERE id = ?',
            (followup_partner, row_id)
        )
        updated_count += cursor.rowcount

    return updated_count

def find_report_header(df):
    for row_index in range(min(len(df), 30)):
        labels = [normalize_cell(value) for value in df.iloc[row_index].tolist()]
        next_labels = [normalize_cell(value) for value in df.iloc[row_index + 1].tolist()] if row_index + 1 < len(df) else []
        combined = [
            f'{labels[index]}{next_labels[index] if index < len(next_labels) else ""}'
            for index in range(len(labels))
        ]

        date_idx = next((index for index, label in enumerate(combined) if label == 'date'), None)
        ref_idx = next((index for index, label in enumerate(combined) if label in ('refno', 'refnumber', 'referenceno')), None)
        party_idx = next((index for index, label in enumerate(combined) if 'party' in label and 'name' in label), None)
        amount_idx = next((index for index, label in enumerate(combined) if label in ('amount', 'pendingamount') or ('pending' in label and 'amount' in label)), None)
        due_idx = next((index for index, label in enumerate(combined) if label in ('dueon', 'duedate')), None)
        overdue_idx = next((index for index, label in enumerate(combined) if 'overdue' in label), None)

        if date_idx is not None and ref_idx is not None and party_idx is not None and amount_idx is not None:
            return {
                'row_index': row_index,
                'date_idx': date_idx,
                'ref_idx': ref_idx,
                'party_idx': party_idx,
                'amount_idx': amount_idx,
                'due_idx': due_idx,
                'overdue_idx': overdue_idx,
            }
    return None

def process_data_file(data_file, manual_firm_name=''):
    """Parses a file and appends rows to the database."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    batch_id = datetime.now().strftime('%Y%m%d%H%M%S%f')
    try:
        imported_count = 0
        imported_rows = []
        firm_name = manual_firm_name.strip() or "No Data Found"
        if data_file.endswith('.csv'):
            with open(data_file, 'r', encoding='utf-8', errors='ignore') as f:
                reader = list(csv.reader(f))
                if not reader:
                    return {'imported_count': 0, 'batch_id': batch_id, 'missing_client_count': 0}
                firm_name = manual_firm_name.strip() or first_nonblank(reader[0]) or firm_name
                rows_data = reader[1:]
                for row in rows_data:
                    try:
                        if len(row) < 4:
                            continue
                        bill_date_obj = parse_input_date(row[0])
                        amount = to_float(row[3])
                        if insert_billing_row(cursor, firm_name, bill_date_obj, row[1], row[2], amount, import_batch_id=batch_id):
                            imported_count += 1
                            imported_rows.append({
                                'firm_name': firm_name,
                                'ref_no': str(row[1]).strip(),
                                'party_name': str(row[2]).strip(),
                                'amount': amount,
                            })
                    except Exception:
                        continue
        else:
            with pd.ExcelFile(data_file) as xls:
                df_firm = pd.read_excel(xls, header=None, nrows=1)
                if not df_firm.empty:
                    firm_name = manual_firm_name.strip() or first_nonblank(df_firm.iloc[0].tolist()) or firm_name

                df = pd.read_excel(xls, header=None)
                header = find_report_header(df)
                if header:
                    start_index = header['row_index'] + 1
                    for row_index in range(start_index, len(df)):
                        row = df.iloc[row_index]
                        try:
                            bill_date_obj = parse_input_date(row.iloc[header['date_idx']])
                            ref_no = row.iloc[header['ref_idx']]
                            party_name = row.iloc[header['party_idx']]
                            amount = to_float(row.iloc[header['amount_idx']])
                            due_date_obj = parse_input_date(row.iloc[header['due_idx']]) if header['due_idx'] is not None else None
                            overdue_days = to_float(row.iloc[header['overdue_idx']]) if header['overdue_idx'] is not None else None
                            if insert_billing_row(cursor, firm_name, bill_date_obj, ref_no, party_name, amount, due_date_obj, overdue_days, batch_id):
                                imported_count += 1
                                imported_rows.append({
                                    'firm_name': firm_name,
                                    'ref_no': str(ref_no).strip(),
                                    'party_name': str(party_name).strip(),
                                    'amount': amount,
                                })
                        except Exception:
                            continue
                else:
                    df_main = pd.read_excel(xls, skiprows=1, header=None)
                    for row in df_main.values.tolist():
                        try:
                            if len(row) < 4:
                                continue
                            bill_date_obj = parse_input_date(row[0])
                            amount = to_float(row[3])
                            if insert_billing_row(cursor, firm_name, bill_date_obj, row[1], row[2], amount, import_batch_id=batch_id):
                                imported_count += 1
                                imported_rows.append({
                                    'firm_name': firm_name,
                                    'ref_no': str(row[1]).strip(),
                                    'party_name': str(row[2]).strip(),
                                    'amount': amount,
                                })
                        except Exception:
                            continue
        missing_client_count = save_import_client_errors(cursor, batch_id, imported_rows)
        if imported_count > 0:
            set_app_meta(cursor, 'last_debtor_imported_at', datetime.now().strftime('%Y-%m-%d'))
        conn.commit()
        return {
            'imported_count': imported_count,
            'batch_id': batch_id,
            'missing_client_count': missing_client_count,
        }
    except Exception as e:
        print(f"Error processing file: {e}")
        return {'imported_count': 0, 'batch_id': batch_id, 'missing_client_count': 0}
    finally:
        conn.close()

def init_db():
    """Initializes the database environment and handles discovery load."""
    os.makedirs(STATIC_DIR, exist_ok=True)
    os.makedirs(JS_DIR, exist_ok=True)
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)

    # Move static files if in root
    for src, dst_dir in [(os.path.join(BASE_DIR, 'style.css'), STATIC_DIR), 
                          (os.path.join(BASE_DIR, 'main.js'), JS_DIR)]:
        if os.path.exists(src) and not os.path.exists(os.path.join(dst_dir, os.path.basename(src))):
            shutil.move(src, os.path.join(dst_dir, os.path.basename(src)))

    conn = connect_debtor_db()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS billing_report (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            firm_name TEXT,
            short_name TEXT,
            bill_date TEXT,
            ref_no TEXT,
            party_name TEXT,
            amount REAL,
            due_date TEXT,
            overdue_days INTEGER,
            import_batch_id TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_name TEXT NOT NULL,
            phone TEXT,
            email TEXT,
            gstin TEXT,
            client_group TEXT,
            crp_of_group TEXT,
            reffered_by TEXT,
            whatapp_group TEXT,
            client_category TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_group_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_name TEXT NOT NULL UNIQUE COLLATE NOCASE,
            crp_name TEXT,
            reffered_by TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS crp_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            crp_name TEXT NOT NULL UNIQUE COLLATE NOCASE,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS executive_partner_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            partner_name TEXT NOT NULL UNIQUE,
            final_ep TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS firm_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            firm_name TEXT NOT NULL UNIQUE COLLATE NOCASE,
            short_name TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS receipt_register (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_bill_id INTEGER,
            firm_name TEXT,
            short_name TEXT,
            bill_date TEXT,
            ref_no TEXT,
            party_name TEXT,
            bill_amount REAL,
            received_amount REAL,
            balance_amount REAL,
            due_date TEXT,
            overdue_days INTEGER,
            client_group TEXT,
            followup_partner TEXT,
            final_ep TEXT,
            crp_of_group TEXT,
            client_category TEXT,
            financial_year TEXT,
            receipt_mode TEXT,
            receipt_date TEXT,
            posted_at TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS receipt_adjustment_register (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_bill_id INTEGER,
            firm_name TEXT,
            short_name TEXT,
            bill_date TEXT,
            ref_no TEXT,
            party_name TEXT,
            bill_amount REAL,
            adjustment_amount REAL,
            balance_amount REAL,
            due_date TEXT,
            overdue_days INTEGER,
            client_group TEXT,
            followup_partner TEXT,
            final_ep TEXT,
            crp_of_group TEXT,
            client_category TEXT,
            financial_year TEXT,
            adjustment_type TEXT,
            adjustment_date TEXT,
            posted_at TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS app_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS debtor_import_client_errors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id TEXT NOT NULL,
            imported_at TEXT,
            party_name TEXT,
            firm_names TEXT,
            bill_count INTEGER DEFAULT 0,
            total_amount REAL DEFAULT 0,
            sample_ref_nos TEXT
        )
    ''')
    ensure_debtor_nav_access_table(cursor)
    ensure_deleted_records_log_table(cursor)

    ensure_receipt_register_tally_columns(cursor)
    ensure_cheque_bounce_register_table(cursor)

    cursor.execute('PRAGMA table_info(client_master)')
    existing_client_columns = {row[1] for row in cursor.fetchall()}
    required_client_columns = {
        'client_group': 'TEXT',
        'crp_of_group': 'TEXT',
        'reffered_by': 'TEXT',
        'whatapp_group': 'TEXT',
        'client_category': 'TEXT',
    }
    for column_name, column_type in required_client_columns.items():
        if column_name not in existing_client_columns:
            cursor.execute(f'ALTER TABLE client_master ADD COLUMN {column_name} {column_type}')

    cursor.execute('PRAGMA table_info(client_group_master)')
    existing_group_columns = {row[1] for row in cursor.fetchall()}
    required_group_columns = {
        'crp_name': 'TEXT',
        'reffered_by': 'TEXT',
        'created_at': 'TEXT',
        'updated_at': 'TEXT',
    }
    for column_name, column_type in required_group_columns.items():
        if column_name not in existing_group_columns:
            cursor.execute(f'ALTER TABLE client_group_master ADD COLUMN {column_name} {column_type}')
    ensure_client_group_master_seed(cursor)
    ensure_crp_master_seed(cursor)
    sync_client_master_from_group_master(cursor)

    cursor.execute('PRAGMA table_info(billing_report)')
    existing_billing_columns = {row[1] for row in cursor.fetchall()}
    if 'ep_override' not in existing_billing_columns:
        cursor.execute('ALTER TABLE billing_report ADD COLUMN ep_override TEXT')
    if 'receipt_status' not in existing_billing_columns:
        cursor.execute("ALTER TABLE billing_report ADD COLUMN receipt_status TEXT DEFAULT 'open'")
    if 'paid_amount' not in existing_billing_columns:
        cursor.execute('ALTER TABLE billing_report ADD COLUMN paid_amount REAL DEFAULT 0')
    if 'closed_at' not in existing_billing_columns:
        cursor.execute('ALTER TABLE billing_report ADD COLUMN closed_at TEXT')
    if 'group_override' not in existing_billing_columns:
        cursor.execute('ALTER TABLE billing_report ADD COLUMN group_override TEXT')
    if 'followup_partner' not in existing_billing_columns:
        cursor.execute('ALTER TABLE billing_report ADD COLUMN followup_partner TEXT')
    if 'deleted_at' not in existing_billing_columns:
        cursor.execute('ALTER TABLE billing_report ADD COLUMN deleted_at TEXT')
    if 'deleted_by' not in existing_billing_columns:
        cursor.execute('ALTER TABLE billing_report ADD COLUMN deleted_by TEXT')
    if 'delete_reason' not in existing_billing_columns:
        cursor.execute('ALTER TABLE billing_report ADD COLUMN delete_reason TEXT')
    if 'import_batch_id' not in existing_billing_columns:
        cursor.execute('ALTER TABLE billing_report ADD COLUMN import_batch_id TEXT')

    cursor.execute('PRAGMA table_info(executive_partner_master)')
    existing_partner_columns = {row[1] for row in cursor.fetchall()}
    if 'final_ep' not in existing_partner_columns:
        cursor.execute('ALTER TABLE executive_partner_master ADD COLUMN final_ep TEXT')

    cursor.execute('''
        INSERT OR IGNORE INTO firm_master (firm_name, short_name)
        SELECT DISTINCT firm_name, short_name
        FROM billing_report
        WHERE firm_name IS NOT NULL AND trim(firm_name) != ''
    ''')
    cursor.execute('''
        UPDATE billing_report
        SET due_date = date(bill_date, '+30 days')
        WHERE bill_date IS NOT NULL AND trim(bill_date) != ''
    ''')
    cursor.execute('''
        SELECT COUNT(*)
        FROM billing_report
        WHERE COALESCE(followup_partner, '') = ''
          AND COALESCE(receipt_status, 'open') != 'full_paid'
          AND COALESCE(import_batch_id, '') = ''
    ''')
    if cursor.fetchone()[0] > 0:
        run_followup_partner_logic(cursor)
    conn.commit()
    
    # Load data only if the table is empty to prevent duplicates on restart
    cursor.execute('SELECT COUNT(*) FROM billing_report')
    is_empty = cursor.fetchone()[0] == 0
    conn.close()

    if is_empty:
        data_file = None
        for name in ['data', 'input']:
            for ext in ['.csv', '.xlsx']:
                p = os.path.join(BASE_DIR, f'{name}{ext}')
                if os.path.exists(p):
                    data_file = p
                    break
            if data_file: break
        if data_file:
            process_data_file(data_file)

def get_report_rows(cursor):
    cursor.execute('''
        SELECT client_group, crp_of_group, COUNT(*) AS use_count
        FROM client_master
        WHERE client_group IS NOT NULL AND trim(client_group) != ''
          AND crp_of_group IS NOT NULL AND trim(crp_of_group) != ''
        GROUP BY client_group, crp_of_group
        ORDER BY client_group, use_count DESC, crp_of_group
    ''')
    crp_by_group = {}
    for row in cursor.fetchall():
        crp_by_group.setdefault(row['client_group'], row['crp_of_group'])
    ep_lookup = load_executive_partner_lookup(cursor)

    cursor.execute('''
        SELECT
            billing_report.*,
            COALESCE(NULLIF(billing_report.group_override, ''), client_master.client_group) AS client_group,
            client_master.client_category,
            client_master.crp_of_group
        FROM billing_report
        LEFT JOIN client_master
            ON lower(trim(billing_report.party_name)) = lower(trim(client_master.client_name))
        WHERE COALESCE(billing_report.receipt_status, 'open') != 'full_paid'
          AND billing_report.deleted_at IS NULL
    ''')
    rows = [dict(row) for row in cursor.fetchall()]
    for row in rows:
        group_name = (row.get('client_group') or '').strip()
        if group_name and crp_by_group.get(group_name):
            row['crp_of_group'] = crp_by_group[group_name]
        row['final_ep'] = (row.get('ep_override') or '').strip() or infer_final_ep_from_ref(row.get('ref_no'), ep_lookup)

    for row in rows:
        row['followup_partner'] = row.get('followup_partner') or ''
        row['bill_date_display'] = format_display_date(row.get('bill_date'))
        bill_date_obj = parse_input_date(row.get('bill_date'))
        if bill_date_obj:
            due_date_obj = bill_date_obj + timedelta(days=30)
            row['due_date'] = due_date_obj.strftime('%Y-%m-%d')
            row['due_date_display'] = format_display_date(row['due_date'])
        else:
            row['due_date_display'] = format_display_date(row.get('due_date'))
        row['financial_year'] = format_financial_year(row.get('bill_date'))

    rows.sort(key=lambda row: (
        (row.get('client_group') or row.get('party_name') or '').lower(),
        (row.get('party_name') or '').lower(),
        row.get('bill_date') or '',
        row.get('ref_no') or '',
    ))
    return rows

def build_report_page_context(active_page='report', page_title='Firm Billing Report'):
    conn = connect_debtor_db()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_report_rows(cursor)

    cursor.execute('''
        SELECT source_bill_id, receipt_date, bill_amount, received_amount, balance_amount, receipt_mode
        FROM receipt_register
        WHERE source_bill_id IS NOT NULL
        ORDER BY receipt_date, id
    ''')
    receipts_by_bill = {}
    for receipt in cursor.fetchall():
        receipts_by_bill.setdefault(receipt['source_bill_id'], []).append(receipt)

    for row in rows:
        receipts = receipts_by_bill.get(row.get('id'), [])
        if not receipts:
            continue

        receipt_lines = []
        for receipt in receipts:
            received_amount = float(receipt['received_amount'] or 0)
            bill_amount = float(receipt['bill_amount'] or 0) if 'bill_amount' in receipt.keys() else 0
            balance_amount = float(receipt['balance_amount'] or 0) if 'balance_amount' in receipt.keys() else 0
            receipt_lines.append(
                f"{format_display_date(receipt['receipt_date'])} | Original Bill: Rs {format_indian_currency(bill_amount)} | Receipt: Rs {format_indian_currency(received_amount)} | Balance: Rs {format_indian_currency(balance_amount)} | {receipt['receipt_mode'] or 'Receipt'}"
            )
        total_received = sum(float(receipt['received_amount'] or 0) for receipt in receipts)
        latest_receipt = receipts[-1]
        original_bill_amount = float(latest_receipt['bill_amount'] or 0)
        balance_amount = float(latest_receipt['balance_amount'] or 0)
        row['part_payment_summary'] = (
            f"Original Rs {format_indian_currency(original_bill_amount)} | "
            f"Paid Rs {format_indian_currency(total_received)} | "
            f"Balance Rs {format_indian_currency(balance_amount)}"
        )
        row['receipt_hover_note'] = (
            "Part payment received\n"
            + "\n".join(receipt_lines)
            + f"\nTotal received: Rs {format_indian_currency(total_received)}"
        )
    
    # Calculate firm-wise summary
    cursor.execute('''
        SELECT short_name, COUNT(*) as bill_count, SUM(amount) as total_amount 
        FROM billing_report
        WHERE deleted_at IS NULL
        GROUP BY short_name
    ''')
    summary = cursor.fetchall()

    cursor.execute('''
        SELECT DISTINCT COALESCE(NULLIF(final_ep, ''), partner_name) AS ep_name
        FROM executive_partner_master
        WHERE COALESCE(NULLIF(final_ep, ''), partner_name) != ''
        ORDER BY ep_name
    ''')
    ep_options = [row['ep_name'] for row in cursor.fetchall()]

    cursor.execute('''
        SELECT DISTINCT client_group
        FROM client_master
        WHERE client_group IS NOT NULL AND trim(client_group) != ''
        ORDER BY client_group
    ''')
    group_options = [row['client_group'] for row in cursor.fetchall()]

    cursor.execute('''
        SELECT client_group, crp_of_group, COUNT(*) AS use_count
        FROM client_master
        WHERE client_group IS NOT NULL AND trim(client_group) != ''
          AND crp_of_group IS NOT NULL AND trim(crp_of_group) != ''
        GROUP BY client_group, crp_of_group
        ORDER BY client_group, use_count DESC, crp_of_group
    ''')
    group_crp_map = {}
    for row in cursor.fetchall():
        group_crp_map.setdefault(row['client_group'], row['crp_of_group'])

    cursor.execute('''
        SELECT client_name, client_group
        FROM client_master
        WHERE client_name IS NOT NULL AND trim(client_name) != ''
        ORDER BY client_name
    ''')
    master_clients = [
        {
            'client_name': row['client_name'],
            'client_group': row['client_group'] or '',
        }
        for row in cursor.fetchall()
    ]
    master_client_names = [client['client_name'] for client in master_clients]

    cursor.execute('SELECT firm_name, short_name FROM firm_master ORDER BY firm_name')
    firm_options = cursor.fetchall()
    
    firm_name = rows[0]['firm_name'] if rows else "No Data"
    conn.close()
    return {
        'rows': rows,
        'firm_name': firm_name,
        'summary': summary,
        'ep_options': ep_options,
        'group_options': group_options,
        'group_crp_map': group_crp_map,
        'master_client_names': master_client_names,
        'master_clients': master_clients,
        'firm_options': firm_options,
        'active_page': active_page,
        'page_title': page_title,
    }

def empty_report_page_context(active_page='report', page_title='Firm Billing Report'):
    return {
        'rows': [],
        'firm_name': 'No Data',
        'summary': [],
        'ep_options': [],
        'group_options': [],
        'group_crp_map': {},
        'master_client_names': [],
        'master_clients': [],
        'firm_options': [],
        'active_page': active_page,
        'page_title': page_title,
    }

@app.route('/')
def index():
    return dashboard()

@app.route('/report')
def report():
    try:
        context = build_report_page_context()
    except sqlite3.OperationalError as exc:
        if 'database is locked' in str(exc).lower():
            flash('Debtor report database is busy/locked. Please try again after the database lock is cleared.', 'error')
        else:
            flash(f'Unable to load report: {exc}', 'error')
        context = empty_report_page_context()
    return render_template(
        'index.html',
        **context
    )

def build_dashboard_context():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_report_rows(cursor)
    cursor.execute('SELECT firm_name, short_name FROM firm_master ORDER BY firm_name')
    firm_options = cursor.fetchall()
    conn.close()

    total_amount = sum(float(row.get('amount') or 0) for row in rows)
    total_bills = len(rows)
    avg_bill = total_amount / total_bills if total_bills else 0
    today = datetime.now()

    def add_amount(summary, key, amount):
        label = (key or 'Unassigned').strip() or 'Unassigned'
        item = summary.setdefault(label, {'label': label, 'count': 0, 'total': 0})
        item['count'] += 1
        item['total'] += amount

    firm_summary = {}
    followup_summary = {}
    ep_summary = {}
    category_summary = {}
    fy_summary = {}
    ageing_summary = {
        '0-30 Days': {'label': '0-30 Days', 'count': 0, 'total': 0},
        '31-90 Days': {'label': '31-90 Days', 'count': 0, 'total': 0},
        '91-180 Days': {'label': '91-180 Days', 'count': 0, 'total': 0},
        '180+ Days': {'label': '180+ Days', 'count': 0, 'total': 0},
    }

    overdue_total = 0
    overdue_count = 0
    for row in rows:
        amount = float(row.get('amount') or 0)
        add_amount(firm_summary, row.get('short_name') or row.get('firm_name'), amount)
        add_amount(followup_summary, row.get('followup_partner'), amount)
        add_amount(ep_summary, row.get('final_ep'), amount)
        add_amount(category_summary, row.get('client_category'), amount)
        add_amount(fy_summary, row.get('financial_year') or 'Unknown', amount)

        bucket = due_ageing_bucket(row)
        ageing_summary[bucket]['count'] += 1
        ageing_summary[bucket]['total'] += amount

        try:
            overdue_days = int(float(row.get('overdue_days') or 0))
        except (TypeError, ValueError):
            overdue_days = 0
        if overdue_days > 30:
            overdue_total += amount
            overdue_count += 1

    def sorted_items(summary, limit=None):
        items = sorted(summary.values(), key=lambda item: item['total'], reverse=True)
        return items[:limit] if limit else items

    high_risk_rows = sorted(rows, key=lambda row: float(row.get('amount') or 0), reverse=True)[:8]

    detail_base_url = f"{request.script_root}/sub-report/detail"
    ageing_items = list(ageing_summary.values())
    for item in ageing_items:
        item['url'] = f"{detail_base_url}?{urlencode({'ageing': item['label']})}"

    chart_data = {
        'firm': sorted_items(firm_summary, 8),
        'followup': sorted_items(followup_summary, 8),
        'ep': sorted_items(ep_summary, 8),
        'category': sorted_items(category_summary, 8),
        'fy': sorted_items(fy_summary),
        'ageing': ageing_items,
    }

    return {
        'active_page': 'dashboard',
        'firm_options': firm_options,
        'total_amount': total_amount,
        'total_bills': total_bills,
        'avg_bill': avg_bill,
        'overdue_total': overdue_total,
        'overdue_count': overdue_count,
        'chart_data': chart_data,
        'top_followups': sorted_items(followup_summary, 6),
        'top_categories': sorted_items(category_summary, 6),
        'high_risk_rows': high_risk_rows,
        'page_title': 'Debtor Dashboard',
    }

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html', **build_dashboard_context())

def overdue_report_rows():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_report_rows(cursor)
    conn.close()

    overdue_rows = []
    for row in rows:
        try:
            overdue_days = int(float(row.get('overdue_days') or 0))
        except (TypeError, ValueError):
            overdue_days = 0
        if overdue_days <= 30:
            continue
        overdue_rows.append({
            'id': row.get('id'),
            'firm_name': row.get('firm_name') or '',
            'short_name': row.get('short_name') or '',
            'bill_date': row.get('bill_date') or '',
            'bill_date_display': row.get('bill_date_display') or format_display_date(row.get('bill_date')),
            'ref_no': row.get('ref_no') or '',
            'party_name': row.get('party_name') or '',
            'amount': float(row.get('amount') or 0),
            'amount_display': format_indian_currency(row.get('amount') or 0, decimals=False),
            'due_date': row.get('due_date') or '',
            'due_date_display': row.get('due_date_display') or format_display_date(row.get('due_date')),
            'overdue_days': overdue_days,
            'client_group': row.get('client_group') or '',
            'followup_partner': row.get('followup_partner') or '',
            'final_ep': row.get('final_ep') or '',
            'crp_of_group': row.get('crp_of_group') or '',
            'client_category': row.get('client_category') or '',
            'financial_year': row.get('financial_year') or '',
            'import_batch_id': row.get('import_batch_id') or '',
        })

    overdue_rows.sort(key=lambda item: (-item['overdue_days'], item['party_name'].lower(), item['ref_no']))
    return overdue_rows

@app.route('/overdue-report')
def overdue_report():
    rows = overdue_report_rows()
    total_amount = sum(row['amount'] for row in rows)
    return render_template(
        'overdue_report.html',
        rows=rows,
        total_amount=total_amount,
        active_page='overdue_report',
        page_title='Overdue Amount Report'
    )

@app.route('/download/overdue-report-excel')
def overdue_report_excel():
    rows = overdue_report_rows()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Overdue Report'
    columns = [
        ('Bill Date', 'bill_date'),
        ('Firm', 'short_name'),
        ('Bill No', 'ref_no'),
        ('Entity', 'party_name'),
        ('Amount', 'amount'),
        ('Due Date', 'due_date'),
        ('Over Due', 'overdue_days'),
        ('Group', 'client_group'),
        ('Followup Partner', 'followup_partner'),
        ('EP', 'final_ep'),
        ('CRP', 'crp_of_group'),
        ('Category', 'client_category'),
        ('F.Y.', 'financial_year'),
    ]

    header_fill = PatternFill('solid', fgColor='34495E')
    for col_index, (label, _) in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=label)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_index, row in enumerate(rows, start=2):
        for col_index, (_, key) in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            if key in {'bill_date', 'due_date'}:
                date_value = parse_excel_date(row.get(key))
                if date_value:
                    cell.value = date_value
                    cell.number_format = 'dd-mm-yyyy'
                else:
                    cell.value = ''
            elif key == 'amount':
                cell.value = float(row.get(key) or 0)
                cell.number_format = '#,##,##0'
            elif key == 'overdue_days':
                cell.value = int(row.get(key) or 0)
            else:
                cell.value = row.get(key) or ''

    widths = [13, 12, 18, 34, 14, 13, 10, 18, 18, 16, 16, 16, 16]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"overdue_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/download/current-overdue-report-excel', methods=['POST'])
def current_overdue_report_excel():
    try:
        row_ids = json.loads(request.form.get('row_ids') or '[]')
        column_indexes = json.loads(request.form.get('columns') or '[]')
    except json.JSONDecodeError:
        row_ids = []
        column_indexes = []

    cleaned_ids = []
    for row_id in row_ids:
        try:
            cleaned_ids.append(int(row_id))
        except (TypeError, ValueError):
            continue

    if not cleaned_ids:
        flash('No visible overdue rows found for export.')
        return redirect(debtor_url('/overdue-report'))

    columns = [
        ('Bill Date', 'bill_date', 13),
        ('Firm', 'short_name', 12),
        ('Bill No', 'ref_no', 18),
        ('Entity', 'party_name', 34),
        ('Amount', 'amount', 14),
        ('Due Date', 'due_date', 13),
        ('Over Due', 'overdue_days', 10),
        ('Group', 'client_group', 18),
        ('Followup Partner', 'followup_partner', 18),
        ('EP', 'final_ep', 16),
        ('CRP', 'crp_of_group', 16),
        ('Category', 'client_category', 16),
        ('F.Y.', 'financial_year', 16),
    ]
    valid_indexes = []
    for index in column_indexes:
        try:
            index = int(index)
        except (TypeError, ValueError):
            continue
        if 0 <= index < len(columns):
            valid_indexes.append(index)
    if not valid_indexes:
        valid_indexes = list(range(len(columns)))

    rows_by_id = {
        int(row.get('id')): row
        for row in overdue_report_rows()
        if row.get('id') is not None
    }
    selected_rows = [rows_by_id[row_id] for row_id in cleaned_ids if row_id in rows_by_id]
    if not selected_rows:
        flash('No matching overdue rows found for export.')
        return redirect(debtor_url('/overdue-report'))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Overdue Report'
    header_fill = PatternFill('solid', fgColor='34495E')
    selected_columns = [columns[index] for index in valid_indexes]

    for col_index, (label, _, _) in enumerate(selected_columns, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=label)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_index, row in enumerate(selected_rows, start=2):
        for col_index, (_, key, _) in enumerate(selected_columns, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            if key in {'bill_date', 'due_date'}:
                date_value = parse_excel_date(row.get(key))
                if date_value:
                    cell.value = date_value
                    cell.number_format = 'dd-mm-yyyy'
                else:
                    cell.value = ''
            elif key == 'amount':
                cell.value = float(row.get(key) or 0)
                cell.number_format = '#,##,##0'
            elif key == 'overdue_days':
                cell.value = int(row.get(key) or 0)
            else:
                cell.value = row.get(key) or ''

    for index, (_, _, width) in enumerate(selected_columns, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='current_filtered_overdue_report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/receipts')
def receipts():
    return render_template(
        'index.html',
        **build_report_page_context(active_page='receipts', page_title='Receipts')
    )

@app.route('/receipt-register')
def receipt_register():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT *
        FROM receipt_register
        ORDER BY receipt_date DESC, id DESC
    ''')
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    for row in rows:
        row['bill_date_display'] = format_display_date(row.get('bill_date'))
        row['due_date_display'] = format_display_date(row.get('due_date'))
        row['receipt_date_display'] = format_display_date(row.get('receipt_date'))

    return render_template('receipt_register.html', rows=rows, active_page='receipt_register')

@app.route('/receipt-register/cheque-bounce', methods=['POST'])
def mark_receipts_cheque_bounce():
    receipt_ids = request.form.getlist('receipt_ids')
    reason = (request.form.get('reason') or '').strip()
    bounce_date = (request.form.get('bounce_date') or datetime.now().strftime('%Y-%m-%d')).strip()
    if not receipt_ids:
        flash('Please select at least one receipt for cheque bounce.')
        return redirect(debtor_url('/receipt-register'))

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    result = readd_receipts_as_cheque_bounce(cursor, receipt_ids, reason, bounce_date)
    conn.commit()
    conn.close()

    message = f"{result['added_count']} cheque bounce record(s) re-added to debtor report."
    if result['skipped_count']:
        message += f" {result['skipped_count']} record(s) skipped because they were already bounced or invalid."
    message += ' Receipt Register entries are unchanged.'
    flash(message)
    return redirect(debtor_url('/cheque-bounce'))

@app.route('/cheque-bounce')
def cheque_bounce_report():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    ensure_cheque_bounce_register_table(cursor)
    cursor.execute('''
        SELECT *
        FROM cheque_bounce_register
        ORDER BY bounced_at DESC, id DESC
    ''')
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    for row in rows:
        row['bounced_at_display'] = format_display_datetime(row.get('bounced_at'))
        row['bounce_date_display'] = format_display_date(row.get('bounce_date'))
        row['receipt_date_display'] = format_display_date(row.get('receipt_date'))
        row['bill_date_display'] = format_display_date(row.get('bill_date'))
        row['due_date_display'] = format_display_date(row.get('due_date'))

    return render_template('cheque_bounce.html', rows=rows, active_page='cheque_bounce')

@app.route('/adjustment-register')
def adjustment_register():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT *
        FROM receipt_adjustment_register
        ORDER BY adjustment_date DESC, id DESC
    ''')
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    for row in rows:
        row['bill_date_display'] = format_display_date(row.get('bill_date'))
        row['due_date_display'] = format_display_date(row.get('due_date'))
        row['adjustment_date_display'] = format_display_date(row.get('adjustment_date'))

    return render_template('adjustment_register.html', rows=rows, active_page='adjustment_register')

@app.route('/receipt-register/weekly-report')
def receipt_register_weekly_report():
    start_date_obj = parse_input_date(request.args.get('start_date', '').strip())
    end_date_obj = parse_input_date(request.args.get('end_date', '').strip())

    if not start_date_obj or not end_date_obj:
        flash('Please select both From Date and To Date for weekly report.')
        return redirect(url_for('receipt_register'))

    start_date = start_date_obj.strftime('%Y-%m-%d')
    end_date = end_date_obj.strftime('%Y-%m-%d')
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT *
        FROM receipt_register
        WHERE receipt_date BETWEEN ? AND ?
        ORDER BY followup_partner, receipt_date, id
    ''', (start_date, end_date))
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    for row in rows:
        row['receipt_date_display'] = format_header_date(row.get('receipt_date'))
        row['bill_date_display'] = format_header_date(row.get('bill_date'))
        row['due_date_display'] = format_header_date(row.get('due_date'))

    detail_rows = []
    summary = {}
    for row in rows:
        followup_partner = row.get('followup_partner') or 'Unassigned'
        received_amount = float(row.get('received_amount') or 0)
        balance_amount = float(row.get('balance_amount') or 0)
        item = summary.setdefault(followup_partner, {
            'Followup Partner': followup_partner,
            'Receipt Count': 0,
            'Received Amount': 0,
            'Balance Amount': 0,
        })
        item['Receipt Count'] += 1
        item['Received Amount'] += received_amount
        item['Balance Amount'] += balance_amount

        detail_rows.append({
            'Receipt Date': row.get('receipt_date_display') or '',
            'Mode': row.get('receipt_mode') or '',
            'Source': row.get('import_source') or 'Manual',
            'Tally Vch': row.get('tally_voucher_no') or '',
            'Firm': row.get('short_name') or '',
            'Bill Date': row.get('bill_date_display') or '',
            'Bill No': row.get('ref_no') or '',
            'Entity': row.get('party_name') or '',
            'Bill Amount': row.get('bill_amount') or 0,
            'Received Amount': received_amount,
            'Balance': balance_amount,
            'Due Date': row.get('due_date_display') or '',
            'Over Due': row.get('overdue_days') or 0,
            'Group': row.get('client_group') or '',
            'Followup Partner': followup_partner,
            'EP': row.get('final_ep') or '',
            'CRP': row.get('crp_of_group') or '',
            'Category': row.get('client_category') or '',
            'F.Y.': row.get('financial_year') or '',
            'Posted At': row.get('posted_at') or '',
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(sorted(summary.values(), key=lambda item: item['Followup Partner'])).to_excel(
            writer,
            index=False,
            sheet_name='Followup Summary',
        )
        pd.DataFrame(detail_rows).to_excel(writer, index=False, sheet_name='Receipt Details')
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = 'A2'
            for column_cells in sheet.columns:
                max_length = 0
                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 42)

    output.seek(0)
    filename = f"weekly_receipt_report_{start_date}_to_{end_date}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def format_message_date(value):
    date_obj = parse_input_date(value)
    if date_obj:
        return date_obj.strftime('%d/%m/%Y')
    return value or ''

def build_weekly_recovery_context(start_date_obj, end_date_obj, cumulative_start_obj=None, partner=''):
    start_date = start_date_obj.strftime('%Y-%m-%d')
    end_date = end_date_obj.strftime('%Y-%m-%d')
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    cumulative_start = (cumulative_start_obj or start_date_obj).strftime('%Y-%m-%d')
    if cumulative_start > end_date:
        cumulative_start = start_date
    selected_partner = (partner or '').strip()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    row_params = [start_date, end_date]
    partner_filter = ''
    if selected_partner:
        partner_filter = " AND COALESCE(NULLIF(followup_partner, ''), 'Unassigned') = ?"
        row_params.append(selected_partner)
    cursor.execute(f'''
        SELECT *
        FROM receipt_register
        WHERE receipt_date BETWEEN ? AND ?
        {partner_filter}
        ORDER BY followup_partner, receipt_date, id
    ''', row_params)
    rows = [dict(row) for row in cursor.fetchall()]

    cumulative_params = [cumulative_start, end_date]
    if selected_partner:
        cumulative_params.append(selected_partner)
    cursor.execute(f'''
        SELECT COALESCE(SUM(received_amount), 0) AS cumulative_total
        FROM receipt_register
        WHERE receipt_date BETWEEN ? AND ?
        {partner_filter}
    ''', cumulative_params)
    cumulative_total = float(cursor.fetchone()['cumulative_total'] or 0)
    conn.close()

    partner_summary = {}
    weekly_total = 0
    for row in rows:
        partner = row.get('followup_partner') or 'Unassigned'
        received_amount = float(row.get('received_amount') or 0)
        weekly_total += received_amount
        row['receipt_date_display'] = format_message_date(row.get('receipt_date'))
        row['bill_amount_display'] = format_indian_currency(row.get('bill_amount') or 0, decimals=False)
        row['received_amount_display'] = format_indian_currency(received_amount, decimals=False)
        row['balance_amount_display'] = format_indian_currency(row.get('balance_amount') or 0, decimals=False)
        item = partner_summary.setdefault(partner, {
            'partner': partner,
            'receipt_count': 0,
            'received_amount': 0,
        })
        item['receipt_count'] += 1
        item['received_amount'] += received_amount

    summary_rows = sorted(
        partner_summary.values(),
        key=lambda item: item['received_amount'],
        reverse=True
    )
    top_partners = summary_rows[:3]

    if top_partners:
        if len(top_partners) == 1:
            top_line = (
                f"{top_partners[0]['partner']} has done a good job by recovering "
                f"Rs {format_indian_currency(top_partners[0]['received_amount'], decimals=False)}."
            )
        else:
            partner_parts = [
                f"{item['partner']} (Rs {format_indian_currency(item['received_amount'], decimals=False)})"
                for item in top_partners
            ]
            top_line = f"{', '.join(partner_parts[:-1])} and {partner_parts[-1]} have done a good job."
    else:
        top_line = "No recovery has been posted for the selected period."

    message = (
        "A Very Good Afternoon to Everyone,\n\n"
        f"I am hereby sharing the client recovery for the Week from {format_message_date(start_date)} to {format_message_date(end_date)}.\n\n"
        f"With our combined efforts, we have been able to recover a sum of Rs {format_indian_currency(weekly_total, decimals=False)} /- during this period.\n\n"
        f"{top_line}\n\n"
        "I am sure others have also worked hard and tried to recover as much as possible.\n\n"
        f"During this period from {format_message_date(cumulative_start)} to {format_message_date(end_date)} "
        f"we have recovered total Rs {format_indian_currency(cumulative_total, decimals=False)} /-"
    )

    for row in summary_rows:
        row['received_amount_display'] = format_indian_currency(row['received_amount'], decimals=False)

    return {
        'active_page': 'weekly_recovery_report',
        'start_date': start_date,
        'end_date': end_date,
        'cumulative_start': cumulative_start,
        'selected_partner': selected_partner,
        'start_date_display': format_message_date(start_date),
        'end_date_display': format_message_date(end_date),
        'weekly_total': weekly_total,
        'weekly_receipt_count': len(rows),
        'weekly_total_display': format_indian_currency(weekly_total, decimals=False),
        'cumulative_start_display': format_message_date(cumulative_start),
        'cumulative_total_display': format_indian_currency(cumulative_total, decimals=False),
        'summary_rows': summary_rows,
        'receipt_rows': rows,
        'message': message,
        'page_title': 'Weekly Recovery Report',
    }

@app.route('/weekly-recovery-report')
def weekly_recovery_report():
    start_date_obj = parse_input_date(request.args.get('start_date', '').strip())
    end_date_obj = parse_input_date(request.args.get('end_date', '').strip())
    cumulative_start_obj = parse_input_date(request.args.get('cumulative_start_date', '').strip())
    partner = request.args.get('partner', '').strip()

    if not start_date_obj or not end_date_obj:
        flash('Please select both From Date and To Date for Weekly Recovery Report.')
        return redirect(url_for('receipt_register'))

    return render_template(
        'weekly_recovery_report.html',
        **build_weekly_recovery_context(start_date_obj, end_date_obj, cumulative_start_obj, partner)
    )

def build_tally_receipt_import_summary(start_date='', end_date=''):
    start_date_obj = parse_input_date(start_date)
    end_date_obj = parse_input_date(end_date)
    if not start_date_obj or not end_date_obj:
        return {
            'summary_rows': [],
            'total_count': 0,
            'total_amount': 0,
            'start_date': start_date,
            'end_date': end_date,
            'message': '',
        }

    start_value = start_date_obj.strftime('%Y-%m-%d')
    end_value = end_date_obj.strftime('%Y-%m-%d')
    if start_value > end_value:
        start_value, end_value = end_value, start_value

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    ensure_receipt_register_tally_columns(cursor)
    cursor.execute('''
        SELECT
            date(posted_at) AS import_date,
            COUNT(*) AS entry_count,
            COALESCE(SUM(received_amount), 0) AS total_amount,
            MIN(receipt_date) AS first_receipt_date,
            MAX(receipt_date) AS last_receipt_date
        FROM receipt_register
        WHERE import_source = 'Tally'
          AND posted_at IS NOT NULL
          AND trim(posted_at) != ''
          AND date(posted_at) BETWEEN ? AND ?
        GROUP BY date(posted_at)
        ORDER BY date(posted_at) DESC
    ''', (start_value, end_value))
    rows = [dict(row) for row in cursor.fetchall()]
    cursor.execute('''
        SELECT *
        FROM receipt_register
        WHERE import_source = 'Tally'
          AND posted_at IS NOT NULL
          AND trim(posted_at) != ''
          AND date(posted_at) BETWEEN ? AND ?
        ORDER BY posted_at DESC, receipt_date DESC, id DESC
    ''', (start_value, end_value))
    detail_rows = [dict(row) for row in cursor.fetchall()]
    conn.commit()
    conn.close()

    for row in rows:
        row['import_date_display'] = format_header_date(row.get('import_date'))
        row['receipt_period_display'] = (
            format_header_date(row.get('first_receipt_date'))
            if row.get('first_receipt_date') == row.get('last_receipt_date')
            else f"{format_header_date(row.get('first_receipt_date'))} to {format_header_date(row.get('last_receipt_date'))}"
        )
    for row in detail_rows:
        row['import_date_display'] = format_header_date(str(row.get('posted_at') or '')[:10])
        row['posted_at_display'] = row.get('posted_at') or ''
        row['receipt_date_display'] = format_header_date(row.get('receipt_date'))
        row['bill_date_display'] = format_header_date(row.get('bill_date'))

    total_count = sum(int(row.get('entry_count') or 0) for row in rows)
    total_amount = sum(float(row.get('total_amount') or 0) for row in rows)
    return {
        'summary_rows': rows,
        'detail_rows': detail_rows,
        'total_count': total_count,
        'total_amount': total_amount,
        'start_date': start_value,
        'end_date': end_value,
        'message': f'Tally imports by actual import/posting date from {format_header_date(start_value)} to {format_header_date(end_value)}.',
    }

@app.route('/import-tally-receipts')
def import_tally_receipts():
    return render_template(
        'tally_receipts_import.html',
        active_page='import_tally_receipts',
        import_summary=build_tally_receipt_import_summary()
    )

@app.route('/tally-receipts/import-summary')
def tally_receipt_import_summary():
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    if not start_date or not end_date:
        flash('Please select both From Date and To Date for import summary.', 'error')
        return redirect(url_for('import_tally_receipts'))

    return render_template(
        'tally_receipts_import.html',
        active_page='import_tally_receipts',
        import_summary=build_tally_receipt_import_summary(start_date, end_date)
    )

def get_missing_report_clients(cursor):
    cursor.execute('''
        SELECT
            billing_report.party_name,
            GROUP_CONCAT(DISTINCT COALESCE(NULLIF(billing_report.short_name, ''), billing_report.firm_name)) AS firm_names,
            COUNT(*) AS bill_count,
            COALESCE(SUM(billing_report.amount), 0) AS total_amount,
            GROUP_CONCAT(billing_report.ref_no) AS sample_ref_nos
        FROM billing_report
        LEFT JOIN client_master
            ON lower(trim(billing_report.party_name)) = lower(trim(client_master.client_name))
        WHERE billing_report.deleted_at IS NULL
          AND COALESCE(billing_report.receipt_status, 'open') != 'full_paid'
          AND billing_report.party_name IS NOT NULL
          AND trim(billing_report.party_name) != ''
          AND client_master.id IS NULL
        GROUP BY lower(trim(billing_report.party_name)), billing_report.party_name
        ORDER BY lower(billing_report.party_name)
    ''')
    rows = []
    for row in cursor.fetchall():
        row_dict = dict(row)
        refs = [
            ref.strip()
            for ref in str(row_dict.get('sample_ref_nos') or '').split(',')
            if ref and ref.strip()
        ]
        row_dict['sample_ref_nos'] = ', '.join(refs[:5])
        rows.append(row_dict)
    return rows

def get_client_master_dropdown_options(cursor):
    group_master_columns = {
        'Client Group': 'group_name',
        'CRP of Group': 'crp_name',
        'Referred By': 'reffered_by',
    }
    client_master_columns = {
        'Client Category': 'client_category',
        'Whatapp Group': 'whatapp_group',
    }
    options = {}
    for label, column_name in group_master_columns.items():
        cursor.execute(f'''
            SELECT DISTINCT {column_name} AS option_value
            FROM client_group_master
            WHERE {column_name} IS NOT NULL AND trim({column_name}) != ''
            ORDER BY lower({column_name})
        ''')
        options[label] = [row['option_value'] for row in cursor.fetchall()]
    for label, column_name in client_master_columns.items():
        cursor.execute(f'''
            SELECT DISTINCT {column_name} AS option_value
            FROM client_master
            WHERE {column_name} IS NOT NULL AND trim({column_name}) != ''
            ORDER BY lower({column_name})
        ''')
        options[label] = [row['option_value'] for row in cursor.fetchall()]
    return options

def add_client_master_dropdowns(writer, sheet_name, dropdown_options, max_rows=1000):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    options_sheet = workbook.create_sheet('_Master Dropdowns')
    options_sheet.sheet_state = 'hidden'

    header_by_column = {
        str(cell.value or '').strip(): cell.column
        for cell in worksheet[1]
        if cell.value
    }

    for option_index, (header, values) in enumerate(dropdown_options.items(), start=1):
        options_sheet.cell(row=1, column=option_index, value=header)
        for row_index, value in enumerate(values, start=2):
            options_sheet.cell(row=row_index, column=option_index, value=value)

        target_column = header_by_column.get(header)
        if not target_column or not values:
            continue

        option_column_letter = options_sheet.cell(row=1, column=option_index).column_letter
        target_column_letter = worksheet.cell(row=1, column=target_column).column_letter
        formula = f"'_Master Dropdowns'!${option_column_letter}$2:${option_column_letter}${len(values) + 1}"
        validation = DataValidation(type='list', formula1=formula, allow_blank=True)
        validation.error = 'Please select a value from master dropdown or type a valid new value.'
        validation.errorTitle = 'Invalid master value'
        validation.prompt = f'Select {header} from Client Master options.'
        validation.promptTitle = header
        worksheet.add_data_validation(validation)
        validation.add(f'{target_column_letter}2:{target_column_letter}{max_rows}')

@app.route('/import-errors')
def debtor_import_errors():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_missing_report_clients(cursor)
    total_bills = sum(int(row['bill_count'] or 0) for row in rows)
    total_amount = sum(float(row['total_amount'] or 0) for row in rows)

    cursor.execute('SELECT firm_name, short_name FROM firm_master ORDER BY firm_name')
    firm_options = cursor.fetchall()
    conn.close()

    return render_template(
        'import_errors.html',
        rows=rows,
        total_bills=total_bills,
        total_amount=total_amount,
        firm_options=firm_options,
        active_page='import_errors'
    )

@app.route('/import-tally-receipts/upload', methods=['POST'])
def import_tally_receipts_upload():
    if 'file' not in request.files:
        flash('No receipt file selected.', 'error')
        return redirect(url_for('import_tally_receipts'))

    file = request.files['file']
    if file.filename == '':
        flash('No receipt file selected.', 'error')
        return redirect(url_for('import_tally_receipts'))

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in {'.xlsx', '.xlsm'}:
        flash('Invalid receipt file. Please upload fcRece.xlsx or another .xlsx receipt export.', 'error')
        return redirect(url_for('import_tally_receipts'))

    try:
        file_bytes = file.read()
        preview = parse_tally_receipt_trial('', file_path=io.BytesIO(file_bytes), file_name=file.filename)
        if not preview.get('data_start_row') or not preview.get('rows'):
            flash('Receipt file was not imported because Date header/receipt rows were not found. Old fcRece.xlsx is unchanged.', 'error')
            return redirect(url_for('import_tally_receipts'))

        with open(TALLY_RECEIPT_TRIAL_PATH, 'wb') as target_file:
            target_file.write(file_bytes)

        try:
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            archive_path = os.path.join(UPLOAD_DIR, f'tally_receipts_{timestamp}{ext}')
            with open(archive_path, 'wb') as archive_file:
                archive_file.write(file_bytes)
        except OSError:
            pass
    except OSError as exc:
        flash(f'Unable to save receipt file: {exc}', 'error')
        return redirect(url_for('import_tally_receipts'))
    except Exception as exc:
        flash(f'Receipt file was not imported: {exc}. Old fcRece.xlsx is unchanged.', 'error')
        return redirect(url_for('import_tally_receipts'))

    flash(f'Receipt file imported as {os.path.basename(TALLY_RECEIPT_TRIAL_PATH)}.', 'success')
    return redirect(url_for('import_tally_receipts', uploaded='1'))

@app.route('/api/import-tally-receipts/preview')
def import_tally_receipts_preview():
    search_date = request.args.get('date', '').strip()
    try:
        preview = parse_tally_receipt_trial(search_date)
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        ensure_receipt_register_tally_columns(cursor)
        preview['rows'] = enrich_tally_receipt_rows(preview.get('rows') or [], cursor)
        conn.commit()
        conn.close()
        return jsonify({'success': True, **preview})
    except Exception as exc:
        return jsonify({'success': False, 'message': f'Unable to preview Tally receipts: {exc}'}), 500

@app.route('/api/import-tally-receipts/post', methods=['POST'])
def import_tally_receipts_post():
    payload = request.get_json(silent=True) or {}
    search_date = (payload.get('date') or '').strip()

    try:
        preview = parse_tally_receipt_trial(search_date)
    except Exception as exc:
        return jsonify({'success': False, 'message': f'Unable to read Tally receipts: {exc}'}), 500

    receipt_date_obj = parse_input_date(search_date or '')
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        ensure_receipt_register_tally_columns(cursor)
        enriched_rows = enrich_tally_receipt_rows(preview.get('rows') or [], cursor)
        postable_rows = [row for row in enriched_rows if row.get('is_postable')]

        if not postable_rows:
            raise ValueError('No ready Tally receipts found to post.')

        receipt_rows_by_date = {}
        import_meta_by_date = {}
        for row in postable_rows:
            bill_id = int(row.get('matched_bill_id'))
            row_receipt_date = parse_input_date(row.get('receipt_date') or '')
            if receipt_date_obj and row_receipt_date and row_receipt_date.strftime('%Y-%m-%d') != receipt_date_obj.strftime('%Y-%m-%d'):
                continue
            receipt_date = (row_receipt_date or receipt_date_obj or datetime.now()).strftime('%Y-%m-%d')

            receipt_rows_by_date.setdefault(receipt_date, []).append({
                'row_id': bill_id,
                'received_amount': row.get('post_amount'),
            })
            import_meta_by_date.setdefault(receipt_date, {})[bill_id] = {
                'import_source': 'Tally',
                'tally_voucher_no': row.get('voucher_no') or '',
                'tally_reference_no': row.get('reference_no') or '',
                'tally_bank_name': row.get('bank_name') or '',
                'tally_excel_rows': ', '.join(str(value) for value in (row.get('raw_rows') or [])),
            }

        if not receipt_rows_by_date:
            raise ValueError('No receipts matched the selected date.')

        posted_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        posted_count = 0
        posted_total = 0
        for receipt_date, receipt_rows in receipt_rows_by_date.items():
            group_count, group_total = post_receipt_rows(
                cursor,
                receipt_rows,
                'Bank',
                receipt_date,
                posted_at,
                import_meta_by_bill=import_meta_by_date.get(receipt_date, {}),
            )
            posted_count += group_count
            posted_total += group_total
        set_app_meta(cursor, 'last_receipt_activity_at', max(receipt_rows_by_date.keys()))
        set_app_meta(cursor, 'last_receipt_activity_type', 'Import')
        conn.commit()
    except ValueError as exc:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception as exc:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': f'Unable to post Tally receipts: {exc}'}), 500

    conn.close()
    return jsonify({
        'success': True,
        'message': f'{posted_count} Tally receipt(s) posted. Total: Rs {format_indian_currency(posted_total)}',
        'posted_count': posted_count,
        'posted_total': posted_total,
        'redirect_url': url_for('receipt_register')
    })

@app.route('/receipts/post', methods=['POST'])
def post_receipts():
    payload = request.get_json(silent=True) or {}
    receipt_mode = (payload.get('receipt_mode') or '').strip()
    receipt_date_obj = parse_input_date(payload.get('receipt_date') or '')
    receipt_rows = payload.get('rows') or []
    is_adjustment = receipt_mode in {'Bad Debt', 'Discount'}

    if receipt_mode not in {'Cash', 'Bank', 'Online', 'UPI', 'Bad Debt', 'Discount'}:
        return jsonify({'success': False, 'message': 'Please select a valid receipt mode.'}), 400
    if not receipt_date_obj:
        return jsonify({'success': False, 'message': 'Please select a valid date.'}), 400
    if not isinstance(receipt_rows, list) or not receipt_rows:
        return jsonify({'success': False, 'message': 'Please select at least one bill.'}), 400

    receipt_date = receipt_date_obj.strftime('%Y-%m-%d')
    posted_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        if is_adjustment:
            post_adjustment_rows(cursor, receipt_rows, receipt_mode, receipt_date, posted_at)
            redirect_endpoint = 'adjustment_register'
        else:
            ensure_receipt_register_tally_columns(cursor)
            post_receipt_rows(cursor, receipt_rows, receipt_mode, receipt_date, posted_at)
            set_app_meta(cursor, 'last_receipt_activity_at', receipt_date)
            set_app_meta(cursor, 'last_receipt_activity_type', 'Manual')
            redirect_endpoint = 'receipt_register'

        conn.commit()
    except ValueError as exc:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception as exc:
        traceback.print_exc()
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': f'Unable to post receipt: {exc}'}), 500

    conn.close()
    return jsonify({'success': True, 'redirect_url': url_for(redirect_endpoint)})

@app.route('/sub-report')
def sub_report():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_report_rows(cursor)
    conn.close()

    followup_summary = {}
    ep_summary = {}
    category_summary = {}
    period_ep_summary = {}

    for row in rows:
        f_partner = row.get('followup_partner') or 'Unassigned'
        ep = row.get('final_ep') or 'Unassigned'
        cat = row.get('client_category') or 'Unassigned'
        fy = row.get('financial_year') or 'Unknown'
        amt = int(round(row.get('amount') or 0))

        # Aggregate Followup Summary
        followup_summary.setdefault(f_partner, {'count': 0, 'total': 0})
        followup_summary[f_partner]['count'] += 1
        followup_summary[f_partner]['total'] += amt

        # Aggregate EP Summary
        ep_summary.setdefault(ep, {'count': 0, 'total': 0})
        ep_summary[ep]['count'] += 1
        ep_summary[ep]['total'] += amt

        # Aggregate Category Summary
        category_summary.setdefault(cat, {'count': 0, 'total': 0})
        category_summary[cat]['count'] += 1
        category_summary[cat]['total'] += amt

        # Aggregate EP + Period (FY) Summary
        ep_periods = period_ep_summary.setdefault(ep, {})
        ep_periods.setdefault(fy, {'count': 0, 'total': 0})
        ep_periods[fy]['count'] += 1
        ep_periods[fy]['total'] += amt

    # Sort summaries by total amount descending
    followup_summary = dict(sorted(followup_summary.items(), key=lambda x: x[1]['total'], reverse=True))
    ep_summary = dict(sorted(ep_summary.items(), key=lambda x: x[1]['total'], reverse=True))
    category_summary = dict(sorted(category_summary.items(), key=lambda x: x[1]['total'], reverse=True))

    all_fys = sorted(list(set(row.get('financial_year') or 'Unknown' for row in rows)), reverse=True)
    all_eps = list(ep_summary.keys())

    return render_template('sub_report.html', 
                           followup_summary=followup_summary, 
                           ep_summary=ep_summary, 
                           category_summary=category_summary,
                           period_ep_summary=period_ep_summary, 
                           rows=rows,
                           active_page='sub_report',
                           all_fys=all_fys,
                           all_eps=all_eps)

@app.route('/sub-report/followup-detail')
def sub_report_followup_detail():
    partner = (request.args.get('partner') or 'Unassigned').strip() or 'Unassigned'
    return render_sub_report_detail(partner=partner)

@app.route('/sub-report/detail')
def sub_report_detail():
    return render_sub_report_detail(
        ep=(request.args.get('ep') or '').strip(),
        category=(request.args.get('category') or '').strip(),
        fy=(request.args.get('fy') or '').strip(),
        ageing=(request.args.get('ageing') or '').strip(),
    )

def due_ageing_bucket(row):
    try:
        days = int(float(row.get('overdue_days') or 0))
    except (TypeError, ValueError):
        days = 0
    days = max(0, days)
    if days <= 30:
        return '0-30 Days'
    if days <= 90:
        return '31-90 Days'
    if days <= 180:
        return '91-180 Days'
    return '180+ Days'

def filter_sub_report_rows(rows, partner='', ep='', category='', fy='', ageing=''):
    partner = (partner or '').strip()
    ep = (ep or '').strip()
    category = (category or '').strip()
    fy = (fy or '').strip()
    ageing = (ageing or '').strip()

    def normalized(value, fallback='Unassigned'):
        text = (value or fallback).strip()
        return text or fallback

    filtered_rows = []
    for row in rows:
        if partner and normalized(row.get('followup_partner')) != partner:
            continue
        if ep and normalized(row.get('final_ep')) != ep:
            continue
        if category and normalized(row.get('client_category')) != category:
            continue
        if fy and normalized(row.get('financial_year'), 'Unknown') != fy:
            continue
        if ageing and due_ageing_bucket(row) != ageing:
            continue
        filtered_rows.append(row)

    return filtered_rows

def build_sub_report_detail_title(partner='', ep='', category='', fy='', ageing=''):
    parts = []
    if partner:
        parts.append(f'Followup Partner: {partner}')
    if ep:
        parts.append(f'EP: {ep}')
    if category:
        parts.append(f'Category: {category}')
    if fy:
        parts.append(f'Financial Year: {fy}')
    if ageing:
        parts.append(f'Overdue Ageing: {ageing}')
    return ' | '.join(parts) if parts else 'Detailed Report'

def get_sub_report_detail_rows(partner='', ep='', category='', fy='', ageing=''):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_report_rows(cursor)
    conn.close()
    return filter_sub_report_rows(rows, partner=partner, ep=ep, category=category, fy=fy, ageing=ageing)

def render_sub_report_detail(partner='', ep='', category='', fy='', ageing=''):
    partner = (partner or '').strip()
    ep = (ep or '').strip()
    category = (category or '').strip()
    fy = (fy or '').strip()
    ageing = (ageing or '').strip()

    filtered_rows = get_sub_report_detail_rows(partner=partner, ep=ep, category=category, fy=fy, ageing=ageing)
    total_amount = sum(int(round(row.get('amount') or 0)) for row in filtered_rows)
    title = build_sub_report_detail_title(partner=partner, ep=ep, category=category, fy=fy, ageing=ageing)

    return render_template(
        'sub_report_followup_detail.html',
        title=title,
        partner=partner,
        ep=ep,
        category=category,
        fy=fy,
        ageing=ageing,
        rows=filtered_rows,
        total_amount=total_amount,
        active_page='sub_report'
    )

@app.route('/sub-report/detail-excel')
def sub_report_detail_excel():
    partner = (request.args.get('partner') or '').strip()
    ep = (request.args.get('ep') or '').strip()
    category = (request.args.get('category') or '').strip()
    fy = (request.args.get('fy') or '').strip()
    ageing = (request.args.get('ageing') or '').strip()
    rows = get_sub_report_detail_rows(partner=partner, ep=ep, category=category, fy=fy, ageing=ageing)
    title = build_sub_report_detail_title(partner=partner, ep=ep, category=category, fy=fy, ageing=ageing)

    export_rows = []
    for row in rows:
        export_rows.append({
            'Date': row.get('bill_date_display') or '',
            'Ref No': row.get('ref_no') or '',
            'Party Name': row.get('party_name') or '',
            'Group': row.get('client_group') or '',
            'Category': row.get('client_category') or '',
            'Amount': row.get('amount') or 0,
            'EP': row.get('final_ep') or '',
            'Followup': row.get('followup_partner') or '',
            'FY': row.get('financial_year') or '',
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(export_rows).to_excel(writer, index=False, sheet_name='Detailed Report')
        sheet = writer.book['Detailed Report']
        sheet.insert_rows(1, 3)
        sheet['A1'] = title
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A2'] = f'Records: {len(rows)}'
        sheet['B2'] = 'Total'
        sheet['C2'] = sum(row.get('amount') or 0 for row in rows)
        sheet['B2'].font = Font(bold=True)
        sheet['C2'].font = Font(bold=True)
        sheet.freeze_panes = 'A5'

    output.seek(0)
    filename = f"sub_report_detail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/periodic-report')
def periodic_report():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_report_rows(cursor)
    conn.close()

    current_fy_str = get_current_fy_string()
    group_analysis = {} # group_name -> {'total': 0, 'fys': set()}
    ep_analysis = {}
    all_categories = set()

    for row in rows:
        fy = row.get('financial_year') or 'Unknown'
        amt = int(round(row.get('amount') or 0))
        g_name = (row.get('client_group') or row.get('party_name') or 'Unassigned').strip()
        bill_date = row.get('bill_date')
        ep = row.get('final_ep') or 'Unassigned'
        followup_partner = row.get('followup_partner') or 'Unassigned'
        cat = row.get('client_category') or 'Other'
        all_categories.add(cat)

        if g_name not in group_analysis:
            group_analysis[g_name] = {'total': 0, 'fys': set(), 'first_date': bill_date, 'followup_partners': set()}
        group_analysis[g_name]['total'] += amt
        group_analysis[g_name]['fys'].add(fy)
        if bill_date and (not group_analysis[g_name]['first_date'] or bill_date < group_analysis[g_name]['first_date']):
            group_analysis[g_name]['first_date'] = bill_date
        if followup_partner != 'Unassigned': # Only add if actually assigned
            group_analysis[g_name]['followup_partners'].add(followup_partner)
        ep_analysis[ep] = ep_analysis.get(ep, 0) + amt

    current_only_summary = {}
    prev_accum_summary = {}

    for name, data in group_analysis.items():
        has_previous = any(fy != current_fy_str for fy in data['fys'])
        
        # Determine Followup Partner display
        if len(data['followup_partners']) == 1:
            followup_partner_display = next(iter(data['followup_partners']))
        elif len(data['followup_partners']) > 1:
            followup_partner_display = 'Multiple'
        else:
            followup_partner_display = 'N/A'

        summary_item = {
            'total': data['total'],
            'first_date': format_display_date(data['first_date']),
            'followup_partner': followup_partner_display
        }
        if has_previous:
            prev_accum_summary[name] = summary_item
        elif current_fy_str in data['fys']:
            current_only_summary[name] = summary_item

    # Sort by total amount descending
    current_only_summary = dict(sorted(current_only_summary.items(), key=lambda x: x[1]['total'], reverse=True))
    prev_accum_summary = dict(sorted(prev_accum_summary.items(), key=lambda x: x[1]['total'], reverse=True))
    ep_summary = dict(sorted(ep_analysis.items(), key=lambda x: x[1], reverse=True))
    sorted_categories = sorted(list(all_categories))

    # Calculate Followup vs Category / Period Pivot Data
    current_fp_cat = {} # {fp: {cat: amt}}
    prev_fp_cat = {}
    prev_fp_period = {}

    for row in rows:
        amt = int(round(row.get('amount') or 0))
        g_name = (row.get('client_group') or row.get('party_name') or 'Unassigned').strip()
        cat = row.get('client_category') or 'Other'
        fp = row.get('followup_partner') or 'Unassigned'
        fy = row.get('financial_year') or 'Unknown'

        if g_name in current_only_summary:
            d = current_fp_cat.setdefault(fp, {})
            d[cat] = d.get(cat, 0) + amt
        elif g_name in prev_accum_summary:
            d = prev_fp_cat.setdefault(fp, {})
            d[cat] = d.get(cat, 0) + amt
            period_data = prev_fp_period.setdefault(fp, {})
            period_data[fy] = period_data.get(fy, 0) + amt

    # Sort FPs by total in their respective pivot
    current_fp_cat = dict(sorted(current_fp_cat.items(), key=lambda x: sum(x[1].values()), reverse=True))
    prev_fp_cat = dict(sorted(prev_fp_cat.items(), key=lambda x: sum(x[1].values()), reverse=True))
    prev_fp_period = dict(sorted(prev_fp_period.items(), key=lambda x: sum(x[1].values()), reverse=True))
    sorted_periods = sorted(
        {period for periods in prev_fp_period.values() for period in periods},
        reverse=True
    )
    current_fp_cat_totals = {
        cat: sum(cats.get(cat, 0) for cats in current_fp_cat.values())
        for cat in sorted_categories
    }
    prev_fp_cat_totals = {
        cat: sum(cats.get(cat, 0) for cats in prev_fp_cat.values())
        for cat in sorted_categories
    }
    prev_fp_period_totals = {
        period: sum(periods.get(period, 0) for periods in prev_fp_period.values())
        for period in sorted_periods
    }

    return render_template('periodic_report.html',
                           current_only_summary=current_only_summary,
                           prev_accum_summary=prev_accum_summary,
                           ep_summary=ep_summary,
                           current_fp_cat=current_fp_cat,
                           prev_fp_cat=prev_fp_cat,
                           prev_fp_period=prev_fp_period,
                           current_fp_cat_totals=current_fp_cat_totals,
                           prev_fp_cat_totals=prev_fp_cat_totals,
                           prev_fp_period_totals=prev_fp_period_totals,
                           all_categories=sorted_categories,
                           all_periods=sorted_periods,
                           active_page='periodic_report')

def periodic_group_detail_rows(group_name, period_scope=''):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_report_rows(cursor)
    conn.close()

    detail_rows = []
    current_fy = get_current_fy_string()
    for row in rows:
        row_group = (row.get('client_group') or row.get('party_name') or 'Unassigned').strip()
        if row_group != group_name:
            continue
        row_fy = row.get('financial_year') or ''
        if period_scope == 'current' and row_fy != current_fy:
            continue
        if period_scope == 'previous' and row_fy == current_fy:
            continue

        amount = float(row.get('amount') or 0)
        detail_rows.append({
            'bill_date': row.get('bill_date') or '',
            'bill_date_display': format_display_date(row.get('bill_date')),
            'firm_name': row.get('firm_name') or '',
            'short_name': row.get('short_name') or '',
            'ref_no': row.get('ref_no') or '',
            'party_name': row.get('party_name') or '',
            'amount': amount,
            'amount_display': format_indian_currency(amount, decimals=False),
            'due_date': row.get('due_date') or '',
            'due_date_display': format_display_date(row.get('due_date')),
            'overdue_days': row.get('overdue_days') or 0,
            'financial_year': row.get('financial_year') or '',
            'followup_partner': row.get('followup_partner') or '',
            'final_ep': row.get('final_ep') or '',
            'client_category': row.get('client_category') or '',
        })

    detail_rows.sort(key=lambda item: (
        item['bill_date'] or '',
        item['ref_no'] or '',
        item['party_name'] or '',
    ))
    return detail_rows

@app.route('/api/periodic-report/group-detail')
def periodic_group_detail():
    group_name = request.args.get('group', '').strip()
    period_scope = request.args.get('period', '').strip()
    if not group_name:
        return jsonify({'success': False, 'message': 'Group name is required.'}), 400

    detail_rows = periodic_group_detail_rows(group_name, period_scope)
    total_amount = sum(item['amount'] for item in detail_rows)

    return jsonify({
        'success': True,
        'group': group_name,
        'rows': detail_rows,
        'count': len(detail_rows),
        'total_amount_display': format_indian_currency(total_amount, decimals=False),
    })

def safe_export_filename(value):
    cleaned = ''.join(char if char.isalnum() or char in ('-', '_') else '_' for char in str(value or '').strip())
    return cleaned.strip('_')[:80] or 'periodic_detail'

def parse_excel_date(value):
    parsed = parse_input_date(value or '')
    return parsed if parsed else None

@app.route('/periodic-report/group-detail/export/excel')
def periodic_group_detail_export_excel():
    group_name = request.args.get('group', '').strip()
    period_scope = request.args.get('period', '').strip()
    if not group_name:
        flash('Group name is required.')
        return redirect(url_for('periodic_report'))

    rows = periodic_group_detail_rows(group_name, period_scope)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Current Detail'
    columns = [
        ('Bill Date', 'bill_date'),
        ('Firm', 'short_name'),
        ('Ref No.', 'ref_no'),
        ('Party', 'party_name'),
        ('Amount', 'amount'),
        ('Due Date', 'due_date'),
        ('Overdue', 'overdue_days'),
        ('Followup', 'followup_partner'),
        ('EP', 'final_ep'),
        ('Category', 'client_category'),
        ('FY', 'financial_year'),
    ]

    header_fill = PatternFill('solid', fgColor='34495E')
    for col_index, (label, _) in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=label)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_index, row in enumerate(rows, start=2):
        for col_index, (_, key) in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            if key in {'bill_date', 'due_date'}:
                date_value = parse_excel_date(row.get(key))
                if date_value:
                    cell.value = date_value
                    cell.number_format = 'dd-mm-yyyy'
                else:
                    cell.value = ''
            elif key == 'amount':
                cell.value = float(row.get(key) or 0)
                cell.number_format = '#,##,##0'
            elif key == 'overdue_days':
                cell.value = int(row.get(key) or 0)
            else:
                cell.value = row.get(key) or ''

    widths = [13, 12, 18, 34, 14, 13, 10, 18, 16, 16, 14]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"periodic_detail_{safe_export_filename(group_name)}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def periodic_export_visible_columns():
    return [
        ('Bill Date', 'bill_date_display', 92),
        ('Firm', 'short_name', 82),
        ('Ref No.', 'ref_no', 140),
        ('Party', 'party_name', 280),
        ('Amount', 'amount_display', 118),
        ('Due Date', 'due_date_display', 92),
        ('Overdue', 'overdue_days', 76),
    ]

def export_cell_text(row, key):
    value = row.get(key)
    if key == 'amount_display':
        return f"Rs {value or ''}"
    return str(value if value is not None else '')

def truncate_text(value, limit):
    value = str(value or '')
    return value if len(value) <= limit else value[:max(0, limit - 1)] + '~'

def periodic_group_detail_svg(group_name, rows):
    columns = [
        (label, key, width, max(8, int(width / 7)))
        for label, key, width in periodic_export_visible_columns()
    ]
    row_height = 26
    header_height = 30
    title_height = 42
    padding = 16
    table_width = sum(col[2] for col in columns)
    width = table_width + padding * 2
    height = title_height + header_height + max(len(rows), 1) * row_height + padding
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#ffffff"/>',
        f'<text x="{padding}" y="24" fill="#1f2d3a" font-family="Arial, sans-serif" font-size="14" font-weight="700">{html.escape("Periodic Detail - " + group_name)}</text>',
    ]
    x = padding
    y = title_height
    for label, _, col_width, _ in columns:
        parts.append(f'<rect x="{x}" y="{y}" width="{col_width}" height="{header_height}" fill="#34495e" stroke="#b7c4cc"/>')
        parts.append(f'<text x="{x + 5}" y="{y + 19}" fill="#ffffff" font-family="Arial, sans-serif" font-size="11" font-weight="700">{html.escape(label)}</text>')
        x += col_width

    if not rows:
        y_row = y + header_height
        parts.append(f'<rect x="{padding}" y="{y_row}" width="{table_width}" height="{row_height}" fill="#ffffff" stroke="#cfd8dc"/>')
        parts.append(f'<text x="{padding + 8}" y="{y_row + 18}" fill="#34495e" font-family="Arial, sans-serif" font-size="11">No records found.</text>')
    for row_number, row in enumerate(rows or []):
        x = padding
        y_row = y + header_height + row_number * row_height
        fill = '#ffffff' if row_number % 2 == 0 else '#f8fafb'
        for _, key, col_width, char_limit in columns:
            value = truncate_text(export_cell_text(row, key), char_limit)
            parts.append(f'<rect x="{x}" y="{y_row}" width="{col_width}" height="{row_height}" fill="{fill}" stroke="#cfd8dc"/>')
            parts.append(f'<text x="{x + 5}" y="{y_row + 17}" fill="#24384a" font-family="Arial, sans-serif" font-size="10">{html.escape(value)}</text>')
            x += col_width

    parts.append('</svg>')
    return '\n'.join(parts).encode('utf-8')

def pdf_escape(value):
    value = str(value or '')
    value = value.encode('latin-1', errors='replace').decode('latin-1')
    return value.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')

def periodic_group_detail_pdf(group_name, rows):
    columns = [
        (label, key, width, max(8, int(width / 7)))
        for label, key, width in periodic_export_visible_columns()
    ]
    row_height = 18
    width = 842
    height = max(595, 70 + (len(rows) + 2) * row_height)
    left = 24
    y = height - 34
    commands = [
        'q',
        '1 1 1 rg 0 0 842 {0} re f'.format(height),
        '0 0 0 rg /F1 14 Tf',
        f'BT {left} {y} Td (Periodic Detail - {pdf_escape(group_name)}) Tj ET',
    ]
    y -= 28
    x = left
    commands.append('0.20 0.29 0.37 rg')
    commands.append(f'{left} {y - 5} {sum(col[2] for col in columns)} 20 re f')
    commands.append('1 1 1 rg /F1 8 Tf')
    for label, _, col_width, _ in columns:
        commands.append(f'BT {x + 3} {y} Td ({pdf_escape(label)}) Tj ET')
        x += col_width
    y -= row_height
    commands.append('0 0 0 rg /F1 8 Tf')
    if not rows:
        commands.append(f'BT {left + 3} {y} Td (No records found.) Tj ET')
    for row in rows:
        x = left
        for _, key, col_width, char_limit in columns:
            value = truncate_text(export_cell_text(row, key), char_limit)
            commands.append(f'BT {x + 3} {y} Td ({pdf_escape(value)}) Tj ET')
            x += col_width
        y -= row_height
    commands.append('Q')
    stream = '\n'.join(commands).encode('latin-1', errors='replace')
    objects = [
        b'<< /Type /Catalog /Pages 2 0 R >>',
        b'<< /Type /Pages /Kids [3 0 R] /Count 1 >>',
        f'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>'.encode('ascii'),
        b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
        b'<< /Length ' + str(len(stream)).encode('ascii') + b' >>\nstream\n' + stream + b'\nendstream',
    ]
    output = io.BytesIO()
    output.write(b'%PDF-1.4\n')
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f'{index} 0 obj\n'.encode('ascii'))
        output.write(obj)
        output.write(b'\nendobj\n')
    xref_offset = output.tell()
    output.write(f'xref\n0 {len(objects) + 1}\n'.encode('ascii'))
    output.write(b'0000000000 65535 f \n')
    for offset in offsets[1:]:
        output.write(f'{offset:010d} 00000 n \n'.encode('ascii'))
    output.write(f'trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF'.encode('ascii'))
    output.seek(0)
    return output

@app.route('/periodic-report/group-detail/export/image')
def periodic_group_detail_export_image():
    group_name = request.args.get('group', '').strip()
    period_scope = request.args.get('period', '').strip()
    if not group_name:
        flash('Group name is required.')
        return redirect(url_for('periodic_report'))

    rows = periodic_group_detail_rows(group_name, period_scope)
    output = io.BytesIO(periodic_group_detail_svg(group_name, rows))
    return send_file(
        output,
        as_attachment=True,
        download_name=f"periodic_detail_{safe_export_filename(group_name)}.svg",
        mimetype='image/svg+xml'
    )

@app.route('/periodic-report/group-detail/export/pdf')
def periodic_group_detail_export_pdf():
    group_name = request.args.get('group', '').strip()
    period_scope = request.args.get('period', '').strip()
    if not group_name:
        flash('Group name is required.')
        return redirect(url_for('periodic_report'))

    rows = periodic_group_detail_rows(group_name, period_scope)
    output = periodic_group_detail_pdf(group_name, rows)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"periodic_detail_{safe_export_filename(group_name)}.pdf",
        mimetype='application/pdf'
    )

def build_report_export_data():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_report_rows(cursor)
    conn.close()

    followup_summary = {}
    ep_summary = {}
    category_summary = {}
    period_ep_summary = {}

    for row in rows:
        f_partner = row.get('followup_partner') or 'Unassigned'
        ep = row.get('final_ep') or 'Unassigned'
        cat = row.get('client_category') or 'Unassigned'
        fy = row.get('financial_year') or 'Unknown'
        amt = int(round(row.get('amount') or 0))

        followup_summary.setdefault(f_partner, {'count': 0, 'total': 0})
        followup_summary[f_partner]['count'] += 1
        followup_summary[f_partner]['total'] += amt

        ep_summary.setdefault(ep, {'count': 0, 'total': 0})
        ep_summary[ep]['count'] += 1
        ep_summary[ep]['total'] += amt

        category_summary.setdefault(cat, {'count': 0, 'total': 0})
        category_summary[cat]['count'] += 1
        category_summary[cat]['total'] += amt

        ep_periods = period_ep_summary.setdefault(ep, {})
        ep_periods.setdefault(fy, {'count': 0, 'total': 0})
        ep_periods[fy]['count'] += 1
        ep_periods[fy]['total'] += amt

    followup_summary = dict(sorted(followup_summary.items(), key=lambda x: x[1]['total'], reverse=True))
    ep_summary = dict(sorted(ep_summary.items(), key=lambda x: x[1]['total'], reverse=True))
    category_summary = dict(sorted(category_summary.items(), key=lambda x: x[1]['total'], reverse=True))
    all_fys = sorted(list(set(row.get('financial_year') or 'Unknown' for row in rows)), reverse=True)
    all_eps = list(ep_summary.keys())

    current_fy_str = get_current_fy_string()
    group_analysis = {}
    ep_analysis = {}
    all_categories = set()

    for row in rows:
        fy = row.get('financial_year') or 'Unknown'
        amt = int(round(row.get('amount') or 0))
        g_name = (row.get('client_group') or row.get('party_name') or 'Unassigned').strip()
        bill_date = row.get('bill_date')
        ep = row.get('final_ep') or 'Unassigned'
        followup_partner = row.get('followup_partner') or 'Unassigned'
        cat = row.get('client_category') or 'Other'
        all_categories.add(cat)

        if g_name not in group_analysis:
            group_analysis[g_name] = {'total': 0, 'fys': set(), 'first_date': bill_date, 'followup_partners': set()}
        group_analysis[g_name]['total'] += amt
        group_analysis[g_name]['fys'].add(fy)
        if bill_date and (not group_analysis[g_name]['first_date'] or bill_date < group_analysis[g_name]['first_date']):
            group_analysis[g_name]['first_date'] = bill_date
        if followup_partner != 'Unassigned':
            group_analysis[g_name]['followup_partners'].add(followup_partner)
        ep_analysis[ep] = ep_analysis.get(ep, 0) + amt

    current_only_summary = {}
    prev_accum_summary = {}

    for name, data in group_analysis.items():
        has_previous = any(fy != current_fy_str for fy in data['fys'])
        if len(data['followup_partners']) == 1:
            followup_partner_display = next(iter(data['followup_partners']))
        elif len(data['followup_partners']) > 1:
            followup_partner_display = 'Multiple'
        else:
            followup_partner_display = 'N/A'

        summary_item = {
            'total': data['total'],
            'first_date': format_display_date(data['first_date']),
            'followup_partner': followup_partner_display,
        }
        if has_previous:
            prev_accum_summary[name] = summary_item
        elif current_fy_str in data['fys']:
            current_only_summary[name] = summary_item

    current_only_summary = dict(sorted(current_only_summary.items(), key=lambda x: x[1]['total'], reverse=True))
    prev_accum_summary = dict(sorted(prev_accum_summary.items(), key=lambda x: x[1]['total'], reverse=True))
    periodic_ep_summary = dict(sorted(ep_analysis.items(), key=lambda x: x[1], reverse=True))
    sorted_categories = sorted(list(all_categories))

    current_fp_cat = {}
    prev_fp_cat = {}

    for row in rows:
        amt = int(round(row.get('amount') or 0))
        g_name = (row.get('client_group') or row.get('party_name') or 'Unassigned').strip()
        cat = row.get('client_category') or 'Other'
        fp = row.get('followup_partner') or 'Unassigned'

        if g_name in current_only_summary:
            d = current_fp_cat.setdefault(fp, {})
            d[cat] = d.get(cat, 0) + amt
        elif g_name in prev_accum_summary:
            d = prev_fp_cat.setdefault(fp, {})
            d[cat] = d.get(cat, 0) + amt

    current_fp_cat = dict(sorted(current_fp_cat.items(), key=lambda x: sum(x[1].values()), reverse=True))
    prev_fp_cat = dict(sorted(prev_fp_cat.items(), key=lambda x: sum(x[1].values()), reverse=True))

    return {
        'rows': rows,
        'followup_summary': followup_summary,
        'ep_summary': ep_summary,
        'category_summary': category_summary,
        'period_ep_summary': period_ep_summary,
        'all_fys': all_fys,
        'all_eps': all_eps,
        'current_only_summary': current_only_summary,
        'prev_accum_summary': prev_accum_summary,
        'periodic_ep_summary': periodic_ep_summary,
        'current_fp_cat': current_fp_cat,
        'prev_fp_cat': prev_fp_cat,
        'all_categories': sorted_categories,
    }

def summary_export_rows(summary, name_label):
    grand_total = sum(data['total'] for data in summary.values())
    export_rows = []
    for name, data in summary.items():
        export_rows.append({
            name_label: name,
            'Total': data.get('total', 0),
            'Percent': round((data.get('total', 0) / grand_total * 100), 2) if grand_total else 0,
        })
    export_rows.append({name_label: 'Grand Total', 'Total': grand_total, 'Percent': 100 if grand_total else 0})
    return export_rows

def periodic_group_export_rows(summary, period_label=''):
    total = sum(data['total'] for data in summary.values())
    export_rows = []
    for name, data in summary.items():
        export_rows.append({
            'Period': period_label,
            'Group / Party Name': name,
            'Total': data['total'],
            'Percent': round((data['total'] / total * 100), 2) if total else 0,
            'Followup Partner': data['followup_partner'],
            'First Bill Date': data['first_date'],
        })
    export_rows.append({'Period': period_label, 'Group / Party Name': 'Grand Total', 'Total': total, 'Percent': 100 if total else 0, 'Followup Partner': '', 'First Bill Date': ''})
    return export_rows

def pivot_export_rows(pivot_data, categories, total_amount):
    export_rows = []
    for partner, cats in pivot_data.items():
        partner_total = sum(cats.values())
        row = {'Followup Partner': partner}
        for cat in categories:
            row[cat] = cats.get(cat, 0)
        row['Grand Total'] = partner_total
        row['Percent'] = round((partner_total / total_amount * 100), 2) if total_amount else 0
        export_rows.append(row)
    return export_rows

def apply_workbook_formatting(writer):
    for sheet in writer.book.worksheets:
        if not sheet.freeze_panes:
            sheet.freeze_panes = 'A2'
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column_cells in sheet.columns:
            max_length = 10
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 42)

def apply_excel_date_columns(writer, headers=('Bill Date', 'Due Date', 'First Bill Date', 'Date')):
    header_names = set(headers)
    for sheet in writer.book.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10)):
            date_columns = [
                cell.column
                for cell in row
                if str(cell.value or '').strip() in header_names
            ]
            if not date_columns:
                continue

            for column_index in date_columns:
                for data_row in range(row[0].row + 1, sheet.max_row + 1):
                    cell = sheet.cell(row=data_row, column=column_index)
                    if cell.value in (None, ''):
                        continue
                    parsed_date = parse_excel_date(cell.value)
                    if not parsed_date:
                        continue
                    cell.value = parsed_date
                    cell.number_format = 'dd-mm-yyyy'
            break

def main_report_export_row(row):
    return {
        'Firm': row.get('short_name') or '',
        'Bill Date': parse_excel_date(row.get('bill_date')),
        'Ref. No.': row.get('ref_no') or '',
        "Party's Name": row.get('party_name') or '',
        'Amount': row.get('amount') or 0,
        'Due Date': parse_excel_date(row.get('due_date')),
        'Over Due': row.get('overdue_days') or 0,
        'Group': row.get('client_group') or '',
        'Followup Partner': row.get('followup_partner') or '',
        'EP': row.get('final_ep') or '',
        'CRP': row.get('crp_of_group') or '',
        'Category': row.get('client_category') or '',
        'F.Y.': row.get('financial_year') or '',
    }

MAIN_REPORT_EXPORT_COLUMNS = [
    'Firm',
    'Bill Date',
    'Ref. No.',
    "Party's Name",
    'Amount',
    'Due Date',
    'Over Due',
    'Group',
    'Followup Partner',
    'EP',
    'CRP',
    'Category',
    'F.Y.',
]

def selected_main_report_columns(column_indexes):
    selected_columns = []
    for index in column_indexes:
        try:
            column_index = int(index)
        except (TypeError, ValueError):
            continue
        if 0 <= column_index < len(MAIN_REPORT_EXPORT_COLUMNS):
            selected_columns.append(MAIN_REPORT_EXPORT_COLUMNS[column_index])
    return selected_columns or MAIN_REPORT_EXPORT_COLUMNS

def write_titled_sheet(writer, sheet_name, title, rows):
    pd.DataFrame(rows).to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
    sheet = writer.book[sheet_name]
    sheet['A1'] = title
    sheet['A1'].font = Font(bold=True, size=14)
    sheet.freeze_panes = 'A4'

def write_combined_summary_sheet(writer, data):
    sheet_name = 'Summaries'
    sheet = writer.book.create_sheet(sheet_name)
    current_row = 1
    sections = [
        ('Followup Summary', summary_export_rows(data['followup_summary'], 'Followup Partner')),
        ('EP Summary', summary_export_rows(data['ep_summary'], 'EP Name')),
        ('Category Summary', summary_export_rows(data['category_summary'], 'Category')),
    ]
    for title, rows in sections:
        sheet.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=13)
        current_row += 1
        for row in dataframe_to_rows(pd.DataFrame(rows), index=False, header=True):
            for col_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=current_row, column=col_idx, value=value)
                if current_row == 2 or sheet.cell(row=current_row - 1, column=1).value == title:
                    cell.font = Font(bold=True)
            current_row += 1
        current_row += 2

@app.route('/report/update', methods=['POST'])
def update_report_row():
    row_id = request.form.get('row_id')
    ref_no = request.form.get('ref_no', '').strip()
    party_name = request.form.get('party_name', '').strip()
    client_group = request.form.get('client_group', '').strip()
    add_to_master = request.form.get('add_to_master') == 'yes'
    bill_date_obj = parse_input_date(request.form.get('bill_date', ''))
    amount_value = request.form.get('amount', '').strip()
    ep_override = request.form.get('final_ep', '').strip()
    firm_name = request.form.get('firm_name', '').strip()

    if not row_id or not bill_date_obj or not party_name:
        flash("Unable to update row. Please check the bill date and Party's Name.")
        return redirect(url_for('report'))

    try:
        amount = float(amount_value)
    except ValueError:
        flash('Unable to update row. Please check the amount.')
        return redirect(url_for('report'))

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    short_name = get_firm_short_name(cursor, firm_name) if firm_name else None

    due_date_obj = bill_date_obj + timedelta(days=30)
    overdue_days = (datetime.now() - bill_date_obj).days

    cursor.execute('''
        SELECT client_group
        FROM client_master
        WHERE lower(trim(client_name)) = lower(trim(?))
        LIMIT 1
    ''', (party_name,))
    matching_client = cursor.fetchone()
    matching_master_group = (matching_client[0] if matching_client else '') or ''
    should_add_to_master = add_to_master and not matching_client
    group_override = client_group
    if should_add_to_master or (matching_client and matching_master_group.strip() == client_group):
        group_override = ''

    cursor.execute('''
        UPDATE billing_report
        SET ref_no = ?, bill_date = ?, party_name = ?, group_override = ?, amount = ?, due_date = ?, overdue_days = ?, ep_override = ?,
            firm_name = COALESCE(NULLIF(?, ''), firm_name),
            short_name = COALESCE(NULLIF(?, ''), short_name)
        WHERE id = ?
    ''', (
        ref_no, bill_date_obj.strftime('%Y-%m-%d'), party_name, group_override, amount, 
        due_date_obj.strftime('%Y-%m-%d'), overdue_days, ep_override,
        firm_name, short_name,
        row_id,
    ))

    if should_add_to_master:
        cursor.execute('''
            SELECT crp_of_group
            FROM client_master
            WHERE lower(trim(client_group)) = lower(trim(?))
              AND crp_of_group IS NOT NULL AND trim(crp_of_group) != ''
            GROUP BY crp_of_group
            ORDER BY COUNT(*) DESC, crp_of_group
            LIMIT 1
        ''', (client_group,))
        group_crp = cursor.fetchone()
        cursor.execute('''
            INSERT INTO client_master (client_name, client_group, crp_of_group)
            VALUES (?, ?, ?)
        ''', (party_name, client_group, group_crp[0] if group_crp else ''))

    followup_partner = resolve_or_copy_followup_partner_for_row(cursor, row_id)
    cursor.execute(
        'UPDATE billing_report SET followup_partner = ? WHERE id = ?',
        (followup_partner, row_id)
    )

    conn.commit()
    conn.close()

    flash('Report row updated successfully.')
    return redirect(url_for('report'))

@app.route('/report/delete', methods=['POST'])
def soft_delete_report_row():
    row_id = request.form.get('row_id')
    if not row_id:
        flash('Unable to delete row. Row id is missing.')
        return redirect(url_for('report'))

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    deleted_count = soft_delete_billing_rows(cursor, [row_id], request.form.get('delete_reason', '').strip())
    conn.commit()
    conn.close()

    flash('Report row moved to deleted records.' if deleted_count else 'Report row not found.')
    return redirect(url_for('report'))

@app.route('/control-panel/bulk-delete-report')
def control_panel_bulk_delete_report():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = get_report_rows(cursor)
    conn.close()

    summary_lookup = {}
    for row in rows:
        short_name = row.get('short_name') or row.get('firm_name') or ''
        item = summary_lookup.setdefault(short_name, {'short_name': short_name, 'bill_count': 0, 'total_amount': 0})
        item['bill_count'] += 1
        item['total_amount'] += float(row.get('amount') or 0)

    return render_template(
        'report_bulk_delete.html',
        rows=rows,
        summary=list(summary_lookup.values()),
        active_page='control_panel'
    )

@app.route('/control-panel/bulk-delete-report', methods=['POST'])
def control_panel_bulk_delete_report_post():
    row_ids = request.form.getlist('row_ids')
    cleaned_ids = []
    for row_id in row_ids:
        try:
            cleaned_ids.append(int(row_id))
        except (TypeError, ValueError):
            continue
    cleaned_ids = list(dict.fromkeys(cleaned_ids))

    if not cleaned_ids:
        flash('Please select at least one row to delete.')
        return redirect(url_for('control_panel_bulk_delete_report'))

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    deleted_count = soft_delete_billing_rows(
        cursor,
        cleaned_ids,
        request.form.get('delete_reason', '').strip() or 'Bulk delete from main report'
    )
    conn.commit()
    conn.close()

    flash(f'{deleted_count} report row(s) moved to deleted records.')
    return redirect(url_for('control_panel_bulk_delete_report'))

@app.route('/report/add', methods=['POST'])
def add_report_row():
    ref_no = request.form.get('ref_no', '').strip()
    party_name = request.form.get('party_name', '').strip()
    client_group = request.form.get('client_group', '').strip()
    add_to_master = request.form.get('add_to_master') == 'yes'
    bill_date_obj = parse_input_date(request.form.get('bill_date', ''))
    amount_value = request.form.get('amount', '').strip()
    ep_override = request.form.get('final_ep', '').strip()
    firm_name = request.form.get('firm_name', '').strip()

    if not firm_name or not bill_date_obj or not ref_no or not party_name:
        flash("Unable to add row. Please check firm, bill date, ref no., and Party's Name.")
        return redirect(url_for('report'))

    amount = to_float(amount_value)
    if amount is None:
        flash('Unable to add row. Please check the amount.')
        return redirect(url_for('report'))

    conn = connect_debtor_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            SELECT client_group
            FROM client_master
            WHERE lower(trim(client_name)) = lower(trim(?))
            LIMIT 1
        ''', (party_name,))
        matching_client = cursor.fetchone()
        matching_master_group = (matching_client[0] if matching_client else '') or ''
        should_add_to_master = add_to_master and not matching_client
        group_override = client_group
        if should_add_to_master or (matching_client and matching_master_group.strip() == client_group):
            group_override = ''

        inserted = insert_billing_row(cursor, firm_name, bill_date_obj, ref_no, party_name, amount)
        if not inserted:
            flash("Unable to add row. Please check firm, bill date, ref no., Party's Name, and amount.")
            return redirect(url_for('report'))

        row_id = cursor.lastrowid
        cursor.execute('''
            UPDATE billing_report
            SET group_override = ?, ep_override = ?
            WHERE id = ?
        ''', (group_override, ep_override, row_id))

        if should_add_to_master:
            cursor.execute('''
                SELECT crp_of_group
                FROM client_master
                WHERE lower(trim(client_group)) = lower(trim(?))
                  AND crp_of_group IS NOT NULL AND trim(crp_of_group) != ''
                GROUP BY crp_of_group
                ORDER BY COUNT(*) DESC, crp_of_group
                LIMIT 1
            ''', (client_group,))
            group_crp = cursor.fetchone()
            cursor.execute('''
                INSERT INTO client_master (client_name, client_group, crp_of_group)
                VALUES (?, ?, ?)
            ''', (party_name, client_group, group_crp[0] if group_crp else ''))

        followup_partner = resolve_or_copy_followup_partner_for_row(cursor, row_id)
        cursor.execute(
            'UPDATE billing_report SET followup_partner = ? WHERE id = ?',
            (followup_partner, row_id)
        )

        conn.commit()
        flash('Report row added successfully.')
    except sqlite3.OperationalError as exc:
        conn.rollback()
        if 'database is locked' in str(exc).lower():
            flash("Report row could not be added because the database is busy. Please try again in a few seconds.", 'error')
        else:
            flash(f"Unable to add report row: {exc}", 'error')
    except sqlite3.IntegrityError as exc:
        conn.rollback()
        flash(f"Unable to add report row: {exc}", 'error')
    finally:
        conn.close()
    return redirect(url_for('report'))

@app.route('/report/run-followup-logic', methods=['POST'])
def run_report_followup_logic():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        updated_count = run_followup_partner_logic(cursor)
        conn.commit()
        flash(f'Followup partner logic run successfully for {updated_count} open row(s).')
    except Exception as exc:
        conn.rollback()
        flash(f'Unable to run followup partner logic: {exc}')
    finally:
        conn.close()
    return redirect(url_for('report'))

@app.route('/import-followup-choice', methods=['POST'])
def import_followup_choice():
    batch_id = (request.form.get('batch_id') or '').strip()
    choice = (request.form.get('choice') or 'no').strip().lower()
    return_to = request.form.get('return_to') or url_for('report')

    if not batch_id:
        flash('Followup partner choice skipped: import batch not found.')
        return redirect(return_to)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        if choice == 'yes':
            updated_count = run_followup_partner_logic(cursor)
            flash(f'Followup partner rule run successfully for {updated_count} open row(s).')
        else:
            updated_count = copy_existing_followup_for_import_batch(cursor, batch_id)
            flash(f'Existing group followup partner copied for {updated_count} imported row(s).')
        conn.commit()
    except Exception as exc:
        conn.rollback()
        flash(f'Unable to update followup partners after import: {exc}')
    finally:
        conn.close()

    return redirect(return_to)

@app.route('/download/report-excel')
def download_report_excel():
    data = build_report_export_data()
    rows = data['rows']

    export_rows = []
    for row in rows:
        export_rows.append(main_report_export_row(row))

    output = io.BytesIO()
    as_on_label = datetime.now().strftime('%d.%m.%Y')
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(export_rows).to_excel(writer, index=False, sheet_name='Main Report')
        write_combined_summary_sheet(writer, data)

        ep_period_rows = []
        for fy in data['all_fys']:
            row_data = {'Financial Year': fy}
            row_total = 0
            for ep in data['all_eps']:
                amount = data['period_ep_summary'].get(ep, {}).get(fy, {'total': 0})['total']
                row_data[ep] = amount
                row_total += amount
            row_data['Total'] = row_total
            ep_period_rows.append(row_data)
        total_row = {'Financial Year': 'Grand Total'}
        grand_total = 0
        for ep in data['all_eps']:
            col_total = sum(data['period_ep_summary'].get(ep, {}).get(fy, {'total': 0})['total'] for fy in data['all_fys'])
            total_row[ep] = col_total
            grand_total += col_total
        total_row['Total'] = grand_total
        ep_period_rows.append(total_row)
        write_titled_sheet(
            writer,
            'EP Period Summary',
            f'Summary as on {as_on_label}',
            ep_period_rows
        )

        pd.DataFrame(periodic_group_export_rows(data['current_only_summary'], 'Current Period')).to_excel(writer, index=False, sheet_name='Current Period')
        writer.book['Current Period'].insert_rows(1, 2)
        writer.book['Current Period']['A1'] = 'Current Period Report'
        writer.book['Current Period']['A1'].font = Font(bold=True, size=14)
        writer.book['Current Period']['A2'] = f'Period: Current Period Only | Export as on {as_on_label}'
        writer.book['Current Period'].freeze_panes = 'A4'

        pd.DataFrame(periodic_group_export_rows(data['prev_accum_summary'], 'Previous Years')).to_excel(writer, index=False, sheet_name='Previous Years')
        writer.book['Previous Years'].insert_rows(1, 2)
        writer.book['Previous Years']['A1'] = 'Previous Years Accumulated Report'
        writer.book['Previous Years']['A1'].font = Font(bold=True, size=14)
        writer.book['Previous Years']['A2'] = f'Period: Previous Years | Export as on {as_on_label}'
        writer.book['Previous Years'].freeze_panes = 'A4'

        total_current = sum(item['total'] for item in data['current_only_summary'].values())
        total_previous = sum(item['total'] for item in data['prev_accum_summary'].values())
        pd.DataFrame(pivot_export_rows(data['current_fp_cat'], data['all_categories'], total_current)).to_excel(writer, index=False, sheet_name='Current FP Category')
        pd.DataFrame(pivot_export_rows(data['prev_fp_cat'], data['all_categories'], total_previous)).to_excel(writer, index=False, sheet_name='Previous FP Category')

        periodic_ep_rows = summary_export_rows(
            {name: {'count': 0, 'total': total} for name, total in data['periodic_ep_summary'].items()},
            'EP Name'
        )
        pd.DataFrame(periodic_ep_rows).to_excel(writer, index=False, sheet_name='Periodic EP Summary')

        apply_excel_date_columns(writer)
        apply_workbook_formatting(writer)

    output.seek(0)
    return send_file(
        output,
        download_name='debtor_reports_all_sheets.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/download/current-report-excel', methods=['POST'])
def download_current_report_excel():
    try:
        row_ids = json.loads(request.form.get('row_ids') or '[]')
        column_indexes = json.loads(request.form.get('columns') or '[]')
    except json.JSONDecodeError:
        row_ids = []
        column_indexes = []

    cleaned_ids = []
    for row_id in row_ids:
        try:
            cleaned_ids.append(int(row_id))
        except (TypeError, ValueError):
            continue

    if not cleaned_ids:
        flash('No visible report rows found for export.')
        return redirect(url_for('report'))

    data = build_report_export_data()
    rows_by_id = {int(row.get('id')): row for row in data['rows'] if row.get('id') is not None}
    selected_columns = selected_main_report_columns(column_indexes)
    export_rows = [
        {
            column: main_report_export_row(rows_by_id[row_id]).get(column, '')
            for column in selected_columns
        }
        for row_id in cleaned_ids
        if row_id in rows_by_id
    ]

    if not export_rows:
        flash('No matching report rows found for export.')
        return redirect(url_for('report'))

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(export_rows, columns=selected_columns).to_excel(writer, index=False, sheet_name='Current Report')
        sheet = writer.book['Current Report']
        sheet['A1'].font = Font(bold=True)
        sheet.freeze_panes = 'A2'
        apply_excel_date_columns(writer)
        apply_workbook_formatting(writer)

    output.seek(0)
    return send_file(
        output,
        download_name='current_filtered_report.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/client-master')
def client_master_view():
    conn = connect_debtor_db()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    ensure_client_group_master_seed(cursor)
    sync_client_master_from_group_master(cursor)
    conn.commit()
    cursor.execute('SELECT * FROM client_master')
    clients = cursor.fetchall()
    cursor.execute('''
        SELECT group_name AS client_group
        FROM client_group_master
        WHERE group_name IS NOT NULL AND trim(group_name) != ''
        ORDER BY lower(group_name)
    ''')
    group_options = [row['client_group'] for row in cursor.fetchall()]
    cursor.execute('''
        SELECT group_name AS client_group, crp_name AS crp_of_group
        FROM client_group_master
        WHERE group_name IS NOT NULL AND trim(group_name) != ''
          AND crp_name IS NOT NULL AND trim(crp_name) != ''
        ORDER BY lower(group_name), lower(crp_name)
    ''')
    group_crp_map = {}
    group_master_map = {}
    for row in cursor.fetchall():
        group_crp_map.setdefault(row['client_group'], []).append(row['crp_of_group'])
    cursor.execute('''
        SELECT group_name, COALESCE(crp_name, '') AS crp_name, COALESCE(reffered_by, '') AS reffered_by
        FROM client_group_master
        WHERE group_name IS NOT NULL AND trim(group_name) != ''
        ORDER BY lower(group_name)
    ''')
    for row in cursor.fetchall():
        group_master_map[row['group_name']] = {
            'crp': row['crp_name'],
            'referred_by': row['reffered_by'],
        }
    crp_options = sorted({crp for crp_list in group_crp_map.values() for crp in crp_list})
    conn.close()
    return render_template(
        'client_master.html',
        clients=clients,
        group_options=group_options,
        group_crp_map=group_crp_map,
        group_master_map=group_master_map,
        crp_options=crp_options,
    )

@app.route('/client-group-master')
def client_group_master_view():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    ensure_client_group_master_seed(cursor)
    ensure_crp_master_seed(cursor)
    conn.commit()
    cursor.execute('''
        SELECT
            client_group_master.*,
            (
                SELECT COUNT(*)
                FROM client_master
                WHERE lower(trim(client_master.client_group)) = lower(trim(client_group_master.group_name))
            ) AS client_count
        FROM client_group_master
        ORDER BY lower(group_name)
    ''')
    groups = cursor.fetchall()
    crp_options = get_crp_options(cursor)
    conn.close()
    return render_template('client_group_master.html', groups=groups, crp_options=crp_options)

@app.route('/client-group-master/<int:group_id>/clients')
def client_group_clients_report(group_id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM client_group_master WHERE id = ?', (group_id,))
    group = cursor.fetchone()
    if not group:
        conn.close()
        flash('Client group not found.')
        return redirect(url_for('client_group_master_view'))

    cursor.execute('''
        SELECT *
        FROM client_master
        WHERE lower(trim(client_group)) = lower(trim(?))
        ORDER BY lower(client_name)
    ''', (group['group_name'],))
    clients = cursor.fetchall()
    conn.close()
    return render_template(
        'client_group_clients.html',
        group=group,
        clients=clients,
    )

@app.route('/client-group-master/<int:group_id>/clients-data')
def client_group_clients_data(group_id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM client_group_master WHERE id = ?', (group_id,))
    group = cursor.fetchone()
    if not group:
        conn.close()
        return jsonify({'ok': False, 'message': 'Client group not found.'}), 404

    cursor.execute('''
        SELECT id, client_name, client_group, client_category, crp_of_group,
               reffered_by, whatapp_group, phone, email, gstin
        FROM client_master
        WHERE lower(trim(client_group)) = lower(trim(?))
        ORDER BY lower(client_name)
    ''', (group['group_name'],))
    clients = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({
        'ok': True,
        'group_name': group['group_name'],
        'export_url': url_for('client_group_clients_excel', group_id=group_id),
        'clients': clients,
    })

@app.route('/client-group-master/<int:group_id>/clients-excel')
def client_group_clients_excel(group_id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM client_group_master WHERE id = ?', (group_id,))
    group = cursor.fetchone()
    if not group:
        conn.close()
        flash('Client group not found.')
        return redirect(url_for('client_group_master_view'))

    cursor.execute('''
        SELECT id, client_name, client_group, client_category, crp_of_group,
               reffered_by, whatapp_group, phone, email, gstin
        FROM client_master
        WHERE lower(trim(client_group)) = lower(trim(?))
        ORDER BY lower(client_name)
    ''', (group['group_name'],))
    clients = cursor.fetchall()
    conn.close()

    export_rows = []
    for client in clients:
        export_rows.append({
            'ID': client['id'],
            'Client Name': client['client_name'],
            'Group': client['client_group'],
            'Category': client['client_category'],
            'CRP': client['crp_of_group'],
            'Referred By': client['reffered_by'],
            'WhatsApp Group': client['whatapp_group'],
            'Phone': client['phone'],
            'Email': client['email'],
            'GSTIN': client['gstin'],
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(export_rows).to_excel(writer, index=False, sheet_name='Group Clients')
        apply_workbook_formatting(writer)
    output.seek(0)
    safe_group = re.sub(r'[^A-Za-z0-9_-]+', '_', group['group_name']).strip('_') or 'client_group'
    return send_file(
        output,
        download_name=f'{safe_group}_clients.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/download/client-group-master-excel')
def download_client_group_master_excel():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    ensure_client_group_master_seed(cursor)
    ensure_crp_master_seed(cursor)
    conn.commit()
    cursor.execute('''
        SELECT
            client_group_master.id,
            client_group_master.group_name,
            COALESCE(client_group_master.crp_name, '') AS crp_name,
            COALESCE(client_group_master.reffered_by, '') AS reffered_by,
            (
                SELECT COUNT(*)
                FROM client_master
                WHERE lower(trim(client_master.client_group)) = lower(trim(client_group_master.group_name))
            ) AS client_count
        FROM client_group_master
        ORDER BY lower(group_name)
    ''')
    groups = cursor.fetchall()
    conn.close()

    export_rows = []
    for group in groups:
        export_rows.append({
            'ID': group['id'],
            'Group Name': group['group_name'],
            'CRP Name': group['crp_name'],
            'Referred By': group['reffered_by'],
            'Clients': group['client_count'],
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(export_rows).to_excel(writer, index=False, sheet_name='Client Groups')
        apply_workbook_formatting(writer)
    output.seek(0)
    return send_file(
        output,
        download_name='client_group_master.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/client-group-master/add', methods=['POST'])
def add_client_group():
    group_name = request.form.get('group_name', '').strip()
    crp_name = request.form.get('crp_name', '').strip()
    reffered_by = request.form.get('reffered_by', '').strip()

    if not group_name:
        flash('Group name is required.')
        return redirect(url_for('client_group_master_view'))

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO client_group_master (group_name, crp_name, reffered_by, updated_at)
            VALUES (?, ?, ?, ?)
        ''', (group_name, crp_name, reffered_by, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()
        flash('Client group added successfully.')
    except sqlite3.IntegrityError:
        flash('Client group already exists.')
    finally:
        conn.close()
    return redirect(url_for('client_group_master_view'))

@app.route('/client-group-master/update', methods=['POST'])
def update_client_group():
    group_id = request.form.get('group_id')
    group_name = request.form.get('group_name', '').strip()
    crp_name = request.form.get('crp_name', '').strip()
    reffered_by = request.form.get('reffered_by', '').strip()

    if not group_id or not group_name:
        flash('Group name is required.')
        return redirect(url_for('client_group_master_view'))

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT group_name FROM client_group_master WHERE id = ?', (group_id,))
    old_row = cursor.fetchone()
    if not old_row:
        conn.close()
        flash('Client group not found.')
        return redirect(url_for('client_group_master_view'))

    old_group_name = old_row[0]
    try:
        cursor.execute('''
            UPDATE client_group_master
            SET group_name = ?, crp_name = ?, reffered_by = ?, updated_at = ?
            WHERE id = ?
        ''', (
            group_name,
            crp_name,
            reffered_by,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            group_id,
        ))
        cursor.execute('''
            UPDATE client_master
            SET client_group = ?, crp_of_group = ?, reffered_by = ?
            WHERE lower(trim(client_group)) = lower(trim(?))
        ''', (group_name, crp_name, reffered_by, old_group_name))
        conn.commit()
        flash('Client group updated successfully.')
    except sqlite3.IntegrityError:
        conn.rollback()
        flash('Client group already exists.')
    finally:
        conn.close()
    return redirect(url_for('client_group_master_view'))

@app.route('/client-group-master/delete', methods=['POST'])
def delete_client_group():
    group_id = request.form.get('group_id')
    if group_id:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM client_group_master WHERE id = ?', (group_id,))
        group = cursor.fetchone()
        if group:
            group_dict = sqlite_record_to_dict(group)
            log_deleted_record(
                cursor,
                'client_group_master',
                group_id,
                'Client Group Master',
                group_dict.get('group_name') or f"Group #{group_id}",
                group_dict,
                ' | '.join(part for part in [
                    f"CRP: {group_dict.get('crp_name')}" if group_dict.get('crp_name') else '',
                    f"Referred By: {group_dict.get('reffered_by')}" if group_dict.get('reffered_by') else '',
                ] if part),
            )
        cursor.execute('DELETE FROM client_group_master WHERE id = ?', (group_id,))
        conn.commit()
        conn.close()
        flash('Client group deleted from master. Client records were not deleted.')
    return redirect(url_for('client_group_master_view'))

@app.route('/crp-master')
def crp_master_view():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    ensure_client_group_master_seed(cursor)
    ensure_crp_master_seed(cursor)
    conn.commit()
    cursor.execute('''
        SELECT
            crp_master.*,
            (
                SELECT COUNT(*)
                FROM client_group_master
                WHERE lower(trim(client_group_master.crp_name)) = lower(trim(crp_master.crp_name))
            ) AS group_count
        FROM crp_master
        ORDER BY lower(crp_name)
    ''')
    crps = cursor.fetchall()
    conn.close()
    return render_template('crp_master.html', crps=crps)

@app.route('/crp-master/<int:crp_id>/groups-data')
def crp_master_groups_data(crp_id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM crp_master WHERE id = ?', (crp_id,))
    crp = cursor.fetchone()
    if not crp:
        conn.close()
        return jsonify({'ok': False, 'message': 'CRP not found.'}), 404

    cursor.execute('''
        SELECT
            client_group_master.id,
            client_group_master.group_name,
            COALESCE(client_group_master.crp_name, '') AS crp_name,
            COALESCE(client_group_master.reffered_by, '') AS reffered_by,
            (
                SELECT COUNT(*)
                FROM client_master
                WHERE lower(trim(client_master.client_group)) = lower(trim(client_group_master.group_name))
            ) AS client_count
        FROM client_group_master
        WHERE lower(trim(client_group_master.crp_name)) = lower(trim(?))
        ORDER BY lower(client_group_master.group_name)
    ''', (crp['crp_name'],))
    groups = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({
        'ok': True,
        'crp_name': crp['crp_name'],
        'export_url': url_for('crp_master_groups_excel', crp_id=crp_id),
        'groups': groups,
    })

@app.route('/crp-master/<int:crp_id>/groups-excel')
def crp_master_groups_excel(crp_id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM crp_master WHERE id = ?', (crp_id,))
    crp = cursor.fetchone()
    if not crp:
        conn.close()
        flash('CRP not found.')
        return redirect(url_for('crp_master_view'))

    cursor.execute('''
        SELECT
            client_group_master.id,
            client_group_master.group_name,
            COALESCE(client_group_master.crp_name, '') AS crp_name,
            COALESCE(client_group_master.reffered_by, '') AS reffered_by,
            (
                SELECT COUNT(*)
                FROM client_master
                WHERE lower(trim(client_master.client_group)) = lower(trim(client_group_master.group_name))
            ) AS client_count
        FROM client_group_master
        WHERE lower(trim(client_group_master.crp_name)) = lower(trim(?))
        ORDER BY lower(client_group_master.group_name)
    ''', (crp['crp_name'],))
    groups = cursor.fetchall()
    conn.close()

    export_rows = []
    for group in groups:
        export_rows.append({
            'ID': group['id'],
            'Group Name': group['group_name'],
            'CRP Name': group['crp_name'],
            'Referred By': group['reffered_by'],
            'Clients': group['client_count'],
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(export_rows).to_excel(writer, index=False, sheet_name='CRP Groups')
        apply_workbook_formatting(writer)
    output.seek(0)
    safe_crp = re.sub(r'[^A-Za-z0-9_-]+', '_', crp['crp_name']).strip('_') or 'crp'
    return send_file(
        output,
        download_name=f'{safe_crp}_groups.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/download/crp-master-excel')
def download_crp_master_excel():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    ensure_client_group_master_seed(cursor)
    ensure_crp_master_seed(cursor)
    conn.commit()
    cursor.execute('''
        SELECT
            crp_master.id,
            crp_master.crp_name,
            (
                SELECT COUNT(*)
                FROM client_group_master
                WHERE lower(trim(client_group_master.crp_name)) = lower(trim(crp_master.crp_name))
            ) AS group_count
        FROM crp_master
        ORDER BY lower(crp_name)
    ''')
    crps = cursor.fetchall()
    conn.close()

    export_rows = []
    for crp in crps:
        export_rows.append({
            'ID': crp['id'],
            'CRP Name': crp['crp_name'],
            'Groups': crp['group_count'],
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(export_rows).to_excel(writer, index=False, sheet_name='CRP Master')
        apply_workbook_formatting(writer)
    output.seek(0)
    return send_file(
        output,
        download_name='crp_master.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/crp-master/add', methods=['POST'])
def add_crp():
    crp_name = request.form.get('crp_name', '').strip()
    if not crp_name:
        flash('CRP name is required.')
        return redirect(url_for('crp_master_view'))

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute(
            'INSERT INTO crp_master (crp_name, updated_at) VALUES (?, ?)',
            (crp_name, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        )
        conn.commit()
        flash('CRP added successfully.')
    except sqlite3.IntegrityError:
        flash('CRP already exists.')
    finally:
        conn.close()
    return redirect(url_for('crp_master_view'))

@app.route('/crp-master/update', methods=['POST'])
def update_crp():
    crp_id = request.form.get('crp_id')
    crp_name = request.form.get('crp_name', '').strip()
    if not crp_id or not crp_name:
        flash('CRP name is required.')
        return redirect(url_for('crp_master_view'))

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT crp_name FROM crp_master WHERE id = ?', (crp_id,))
    old_row = cursor.fetchone()
    if not old_row:
        conn.close()
        flash('CRP not found.')
        return redirect(url_for('crp_master_view'))

    old_crp_name = old_row[0]
    try:
        cursor.execute(
            'UPDATE crp_master SET crp_name = ?, updated_at = ? WHERE id = ?',
            (crp_name, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), crp_id)
        )
        cursor.execute(
            'UPDATE client_group_master SET crp_name = ?, updated_at = ? WHERE lower(trim(crp_name)) = lower(trim(?))',
            (crp_name, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), old_crp_name)
        )
        cursor.execute(
            'UPDATE client_master SET crp_of_group = ? WHERE lower(trim(crp_of_group)) = lower(trim(?))',
            (crp_name, old_crp_name)
        )
        conn.commit()
        flash('CRP updated successfully.')
    except sqlite3.IntegrityError:
        conn.rollback()
        flash('CRP already exists.')
    finally:
        conn.close()
    return redirect(url_for('crp_master_view'))

@app.route('/crp-master/delete', methods=['POST'])
def delete_crp():
    crp_id = request.form.get('crp_id')
    if crp_id:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM crp_master WHERE id = ?', (crp_id,))
        crp = cursor.fetchone()
        if crp:
            crp_dict = sqlite_record_to_dict(crp)
            cursor.execute('''
                SELECT COUNT(*)
                FROM client_group_master
                WHERE lower(trim(crp_name)) = lower(trim(?))
            ''', (crp_dict.get('crp_name') or '',))
            group_count = cursor.fetchone()[0]
            log_deleted_record(
                cursor,
                'crp_master',
                crp_id,
                'CRP Master',
                crp_dict.get('crp_name') or f"CRP #{crp_id}",
                crp_dict,
                f"Groups linked: {group_count}",
            )
        cursor.execute('DELETE FROM crp_master WHERE id = ?', (crp_id,))
        conn.commit()
        conn.close()
        flash('CRP deleted from master. Existing group/client records were not deleted.')
    return redirect(url_for('crp_master_view'))

@app.route('/executive-partner-master')
def executive_partner_master_view():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM executive_partner_master ORDER BY partner_name')
    partners = cursor.fetchall()
    conn.close()
    return render_template('executive_partner_master.html', partners=partners)

@app.route('/firm-master')
def firm_master_view():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM firm_master ORDER BY firm_name')
    firms = cursor.fetchall()
    conn.close()
    return render_template('firm_master.html', firms=firms)

@app.route('/firm-master/add', methods=['POST'])
def add_firm():
    firm_name = request.form.get('firm_name', '').strip()
    short_name = request.form.get('short_name', '').strip() or get_short_name(firm_name)

    if not firm_name:
        flash("Firm name is required.")
        return redirect(url_for('firm_master_view'))

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Check for case-insensitive duplicates manually
    cursor.execute(
        'SELECT id FROM firm_master WHERE lower(trim(firm_name)) = lower(trim(?)) LIMIT 1',
        (firm_name,)
    )
    if cursor.fetchone():
        conn.close()
        flash("Firm with this name already exists.")
        return redirect(url_for('firm_master_view'))

    try:
        cursor.execute(
            'INSERT INTO firm_master (firm_name, short_name) VALUES (?, ?)',
            (firm_name, short_name)
        )
        conn.commit()
        flash("Firm added successfully!")
    except sqlite3.IntegrityError:
        flash("Firm already exists.")
    finally:
        conn.close()
    return redirect(url_for('firm_master_view'))

@app.route('/firm-master/update', methods=['POST'])
def update_firm():
    firm_id = request.form.get('firm_id')
    firm_name = request.form.get('firm_name', '').strip()
    short_name = request.form.get('short_name', '').strip() or get_short_name(firm_name)

    if not firm_id or not firm_name:
        flash("Firm name is required.")
        return redirect(url_for('firm_master_view'))

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM firm_master WHERE id = ?', (firm_id,))
    old_row = cursor.fetchone()
    if not old_row:
        conn.close()
        flash("Firm not found.")
        return redirect(url_for('firm_master_view'))

    # Check if we are merging into an existing firm
    cursor.execute(
        'SELECT id, short_name FROM firm_master WHERE lower(trim(firm_name)) = lower(trim(?)) AND id != ? LIMIT 1',
        (firm_name, firm_id)
    )
    target_firm = cursor.fetchone()

    old_firm_name = old_row['firm_name']
    try:
        if target_firm:
            # MERGE LOGIC: Move bills to target firm and delete the old master entry
            target_short = target_firm[1]
            old_firm_dict = sqlite_record_to_dict(old_row)
            log_deleted_record(
                cursor,
                'firm_master',
                firm_id,
                'Firm Master',
                old_firm_dict.get('firm_name') or f"Firm #{firm_id}",
                old_firm_dict,
                f"Merged into: {firm_name}",
            )
            cursor.execute(
                'UPDATE billing_report SET firm_name = ?, short_name = ? WHERE firm_name = ?',
                (firm_name, target_short, old_firm_name)
            )
            cursor.execute('DELETE FROM firm_master WHERE id = ?', (firm_id,))
            flash(f"Firm '{old_firm_name}' has been merged into '{firm_name}'. All bills updated.")
        else:
            # NORMAL RENAME
            cursor.execute(
                'UPDATE firm_master SET firm_name = ?, short_name = ? WHERE id = ?',
                (firm_name, short_name, firm_id)
            )
            cursor.execute(
                'UPDATE billing_report SET firm_name = ?, short_name = ? WHERE firm_name = ?',
                (firm_name, short_name, old_firm_name)
            )
            flash("Firm updated successfully!")
            
        conn.commit()
    except sqlite3.IntegrityError:
        flash("Firm already exists.")
    finally:
        conn.close()
    return redirect(url_for('firm_master_view'))

@app.route('/firm-master/delete', methods=['POST'])
def delete_firm():
    firm_id = request.form.get('firm_id')
    if firm_id:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM firm_master WHERE id = ?', (firm_id,))
        firm = cursor.fetchone()
        if firm:
            firm_dict = sqlite_record_to_dict(firm)
            log_deleted_record(
                cursor,
                'firm_master',
                firm_id,
                'Firm Master',
                firm_dict.get('firm_name') or f"Firm #{firm_id}",
                firm_dict,
                f"Short Name: {firm_dict.get('short_name')}" if firm_dict.get('short_name') else '',
            )
        cursor.execute('DELETE FROM firm_master WHERE id = ?', (firm_id,))
        conn.commit()
        conn.close()
        flash("Firm deleted from master. Report data was not deleted.")
    return redirect(url_for('firm_master_view'))

@app.route('/control-panel')
def control_panel():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    backups = read_backup_log()
    report_update_dates = get_manual_report_update_dates()
    return render_template(
        'control_panel.html',
        backups=backups,
        report_update_dates=report_update_dates,
        active_page='control_panel'
    )

@app.route('/control-panel/report-update', methods=['POST'])
def control_panel_report_update():
    report_as_on_date = request.form.get('report_as_on_date', '').strip()
    last_bill_update_date = request.form.get('last_bill_update_date', '').strip()
    last_receipt_update_date = request.form.get('last_receipt_update_date', '').strip()

    if (
        not parse_input_date(report_as_on_date)
        or not parse_input_date(last_bill_update_date)
        or not parse_input_date(last_receipt_update_date)
    ):
        flash('Please enter Report as on, Last Bill update, and Last Receipt update dates.')
        return redirect(url_for('control_panel'))

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    set_app_meta(cursor, 'manual_report_as_on_date', report_as_on_date)
    set_app_meta(cursor, 'manual_last_bill_update_date', last_bill_update_date)
    set_app_meta(cursor, 'manual_last_receipt_update_date', last_receipt_update_date)
    conn.commit()
    conn.close()

    flash('Report update dates saved successfully.')
    return redirect(url_for('control_panel'))

@app.route('/control-panel/navbar-access')
def debtor_navbar_access():
    users = get_debtor_users()
    saved_access = get_all_debtor_nav_access()
    for user in users:
        email = str(user.get('email') or '').lower()
        user['is_arif'] = email == 'arif.siddiqui@asija.in'
        user['access_keys'] = set(DEBTOR_NAV_ACCESS_KEYS) if user['is_arif'] else saved_access.get(email, set())
    return render_template(
        'navbar_access.html',
        users=users,
        access_items=DEBTOR_NAV_ACCESS_ITEMS,
        active_page='control_panel'
    )

@app.route('/control-panel/navbar-access/update', methods=['POST'])
def update_debtor_navbar_access():
    users = get_debtor_users()
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    ensure_debtor_nav_access_table(cursor)
    cursor.execute('DELETE FROM debtor_nav_access')

    for user in users:
        email = str(user.get('email') or '').strip().lower()
        if not email or email == 'arif.siddiqui@asija.in':
            continue
        selected_keys = set()
        for item in DEBTOR_NAV_ACCESS_ITEMS:
            mode = request.form.get(f'access_{user["id"]}_{item["id"]}', 'hide')
            if mode == 'view' and item.get('view_key'):
                selected_keys.add(item['view_key'])
            elif mode == 'full' and item.get('full_key'):
                selected_keys.add(item['full_key'])
                if item.get('view_key'):
                    selected_keys.add(item['view_key'])
        for access_key in selected_keys:
            if access_key in DEBTOR_NAV_ACCESS_KEYS:
                cursor.execute(
                    'INSERT OR IGNORE INTO debtor_nav_access (user_email, access_key) VALUES (?, ?)',
                    (email, access_key)
                )

    conn.commit()
    conn.close()
    flash('Debtor navbar access updated successfully.')
    return redirect(url_for('debtor_navbar_access'))

@app.route('/deleted-records')
@app.route('/control-panel/deleted-records')
def debtor_deleted_records():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    rows = build_deleted_records_rows(cursor)
    conn.close()
    return render_template(
        'deleted_records.html',
        deleted_rows=rows,
        active_page='control_panel'
    )

@app.route('/deleted-records/restore', methods=['POST'])
@app.route('/control-panel/deleted-records/restore', methods=['POST'])
def restore_deleted_report_record():
    delete_log_id = request.form.get('delete_log_id')
    legacy_billing_id = request.form.get('legacy_billing_id')
    if delete_log_id or legacy_billing_id:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            if delete_log_id:
                recalled, message = restore_deleted_log_record(cursor, delete_log_id)
                flash(message)
            else:
                cursor.execute(
                    '''
                    UPDATE billing_report
                    SET deleted_at = NULL, deleted_by = NULL, delete_reason = NULL
                    WHERE id = ?
                    ''',
                    (legacy_billing_id,)
                )
                recalled = cursor.rowcount > 0
                flash('Deleted record recalled successfully.' if recalled else 'Deleted record not found.')
            conn.commit()
        except (sqlite3.IntegrityError, ValueError) as exc:
            conn.rollback()
            flash(f'Unable to recall this record: {exc}')
        finally:
            conn.close()
    return redirect(url_for('debtor_deleted_records'))

@app.route('/control-panel/backup', methods=['POST'])
def control_panel_backup():
    note = request.form.get('note', '').strip() or 'Manual backup from Control Panel'
    backup_id, skipped = create_project_backup(note=note, backup_type='manual')
    if skipped:
        flash(f"Backup {backup_id} created. {len(skipped)} locked/skipped file(s).")
    else:
        flash(f"Backup {backup_id} created successfully.")
    return redirect(url_for('control_panel'))

@app.route('/control-panel/restore', methods=['POST'])
def control_panel_restore():
    backup_id = request.form.get('backup_id', '').strip()
    success, message = restore_project_backup(backup_id)
    flash(message)
    return redirect(url_for('control_panel'))

@app.route('/executive-partner-master/add', methods=['POST'])
def add_executive_partner():
    partner_name = request.form.get('partner_name', '').strip()
    final_ep = request.form.get('final_ep', '').strip()

    if partner_name:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute(
                'INSERT INTO executive_partner_master (partner_name, final_ep) VALUES (?, ?)',
                (partner_name, final_ep)
            )
            conn.commit()
            flash("Executive partner added successfully!")
        except sqlite3.IntegrityError:
            flash("Executive partner already exists.")
        finally:
            conn.close()
    return redirect(url_for('executive_partner_master_view'))

@app.route('/client-master/add', methods=['POST'])
def add_client():
    client_name = request.form.get('client_name', '').strip()
    phone = request.form.get('phone', '').strip()
    email = request.form.get('email', '').strip()
    gstin = request.form.get('gstin', '').strip()
    client_group = request.form.get('client_group', '').strip()
    crp_of_group = request.form.get('crp_of_group', '').strip()
    reffered_by = request.form.get('reffered_by', '').strip()
    whatapp_group = request.form.get('whatapp_group', '').strip()
    client_category = request.form.get('client_category', '').strip()
    
    if client_name:
        conn = connect_debtor_db()
        cursor = conn.cursor()
        try:
            client_group, crp_of_group, reffered_by = resolve_client_group_parent(
                cursor,
                client_group,
                crp_of_group,
                reffered_by
            )
            cursor.execute('INSERT INTO client_master (client_name, phone, email, gstin, client_group, crp_of_group, reffered_by, whatapp_group, client_category) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                           (client_name, phone, email, gstin, client_group, crp_of_group, reffered_by, whatapp_group, client_category))
            conn.commit()
            flash("Client added successfully!")
        except sqlite3.OperationalError as exc:
            conn.rollback()
            if 'database is locked' in str(exc).lower():
                flash("Client could not be added because the database is busy. Please try again in a few seconds.", 'error')
            else:
                flash(f"Unable to add client: {exc}", 'error')
        finally:
            conn.close()
    return redirect(url_for('client_master_view'))

@app.route('/client-master/update', methods=['POST'])
def update_client():
    client_id = request.form.get('client_id')
    client_name = request.form.get('client_name', '').strip()
    phone = request.form.get('phone', '').strip()
    email = request.form.get('email', '').strip()
    gstin = request.form.get('gstin', '').strip()
    client_group = request.form.get('client_group', '').strip()
    crp_of_group = request.form.get('crp_of_group', '').strip()
    reffered_by = request.form.get('reffered_by', '').strip()
    whatapp_group = request.form.get('whatapp_group', '').strip()
    client_category = request.form.get('client_category', '').strip()

    if not client_id or not client_name:
        flash("Client name is required.")
        return redirect(url_for('client_master_view'))

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    client_group, crp_of_group, reffered_by = resolve_client_group_parent(
        cursor,
        client_group,
        crp_of_group,
        reffered_by
    )
    cursor.execute('''
        UPDATE client_master
        SET client_name = ?, phone = ?, email = ?, gstin = ?, client_group = ?,
            crp_of_group = ?, reffered_by = ?, whatapp_group = ?, client_category = ?
        WHERE id = ?
    ''', (
        client_name,
        phone,
        email,
        gstin,
        client_group,
        crp_of_group,
        reffered_by,
        whatapp_group,
        client_category,
        client_id,
    ))
    conn.commit()
    conn.close()
    flash("Client updated successfully!")
    return redirect(url_for('client_master_view'))

@app.route('/download/client-list')
def download_client_list():
    """Exports the entire client master to Excel for editing."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    ensure_client_group_master_seed(cursor)
    sync_client_master_from_group_master(cursor)
    conn.commit()
    cursor.execute('SELECT * FROM client_master')
    clients = cursor.fetchall()
    conn.close()

    export_rows = []
    for client in clients:
        export_rows.append({
            'ID': client['id'],
            'Client Name': client['client_name'],
            'Client Group': client['client_group'],
            'Client Category': client['client_category'],
            'CRP of Group': client['crp_of_group'],
            'Referred By': client['reffered_by'],
            'Whatapp Group': client['whatapp_group'],
            'Phone': client['phone'],
            'Email': client['email'],
            'GSTIN': client['gstin']
        })

    output = io.BytesIO()
    df = pd.DataFrame(export_rows)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Clients')

    output.seek(0)
    return send_file(
        output,
        download_name='client_list_export.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/download/missing-clients-report')
def download_missing_clients_report():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    missing_clients = get_missing_report_clients(cursor)
    dropdown_options = get_client_master_dropdown_options(cursor)
    conn.close()

    export_rows = []
    for row in missing_clients:
        export_rows.append({
            'ID': '',
            'Client Name': row.get('party_name') or '',
            'Client Group': '',
            'Client Category': '',
            'CRP of Group': '',
            'Referred By': '',
            'Whatapp Group': '',
            'Phone': '',
            'Email': '',
            'GSTIN': '',
            'Report Firm(s)': row.get('firm_names') or '',
            'Report Bill Count': row.get('bill_count') or 0,
            'Report Total Amount': row.get('total_amount') or 0,
            'Sample Ref. No.': row.get('sample_ref_nos') or '',
        })

    output = io.BytesIO()
    df = pd.DataFrame(export_rows)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Missing Clients')
        add_client_master_dropdowns(
            writer,
            'Missing Clients',
            dropdown_options,
            max_rows=max(len(export_rows) + 200, 1000)
        )

    output.seek(0)
    return send_file(
        output,
        download_name='missing_clients_for_client_master.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handles report upload and appends rows to the database."""
    return_to = request.form.get('return_to') or request.referrer or url_for('report')
    if 'file' not in request.files:
        flash("No file part in the request.", 'error')
        return redirect(return_to)
    file = request.files['file']
    if file.filename == '':
        flash("No selected file.", 'error')
        return redirect(return_to)
    manual_firm_name = request.form.get('firm_name', '').strip()
    
    if file and file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        ext = os.path.splitext(file.filename)[1].lower()
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        target_path = os.path.join(UPLOAD_DIR, f'import_{timestamp}{ext}')
        file.save(target_path)
        import_result = process_data_file(target_path, manual_firm_name)
        imported_count = import_result.get('imported_count', 0)
        missing_client_count = import_result.get('missing_client_count', 0)
        batch_id = import_result.get('batch_id', '')
        if missing_client_count:
            flash(f"Import successful: {imported_count} rows imported. {missing_client_count} imported party name(s) are not in Client Master. Check Missing Clients Report for current Main Report status.")
        else:
            flash(f"Import successful: {imported_count} rows imported. No new client name missing in Client Master.")
        if batch_id:
            separator = '&' if '?' in return_to else '?'
            followup_prompt = '&followup_prompt=1' if imported_count > 0 else ''
            return_to = f'{return_to}{separator}import_batch={batch_id}{followup_prompt}'
    else:
        flash("Invalid file type. Please upload .csv, .xlsx, or .xls.")
    return redirect(return_to)

@app.route('/client-master/upload', methods=['POST'])
def client_master_upload():
    """Handles Excel file upload for client master, ignoring duplicates."""
    if 'file' not in request.files:
        flash("No file part in the request.", 'error')
        return redirect(url_for('client_master_view'))
    file = request.files['file']
    if file.filename == '':
        flash("No selected file.", 'error')
        return redirect(url_for('client_master_view'))
    
    if file and file.filename.endswith('.xlsx'):
        temp_path = os.path.join(BASE_DIR, 'temp_client_master.xlsx')
        file.save(temp_path)
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            df = pd.read_excel(temp_path)
            # Normalize column names for flexible matching
            df.columns = [str(c).strip() for c in df.columns]
            def normalize_header(header):
                return ''.join(ch for ch in str(header).lower() if ch.isalnum())

            col_map = {normalize_header(c): c for c in df.columns}

            def get_col_val(row_data, aliases):
                for alias in aliases:
                    actual_col = col_map.get(normalize_header(alias))
                    if actual_col:
                        val = row_data[actual_col]
                        return str(val).strip() if pd.notna(val) else ""
                return ""

            imported_count = 0
            for index, row in df.iterrows():
                client_id = get_col_val(row, ['ID', 'id'])
                client_name = get_col_val(row, ['Client Name', 'Name', 'Party Name', 'Party', 'Client'])
                if not client_name or client_name.lower() == 'nan': 
                    continue
                
                phone = get_col_val(row, ['Phone', 'Mobile', 'Contact', 'Phone Number'])
                email = get_col_val(row, ['Email', 'Email ID', 'Mail'])
                gstin = get_col_val(row, ['GSTIN', 'GST No', 'GST', 'GST Number'])
                client_group = get_col_val(row, ['Client Group', 'Group Name', 'Group'])
                crp_of_group = get_col_val(row, ['CRP of Group', 'CRP'])
                reffered_by = get_col_val(row, ['Reffered By', 'Referred By', 'RefferedBy', 'ReferredBy'])
                whatapp_group = get_col_val(row, ['Whatapp Group', 'Whatapp', 'Whatsapp Group', 'WhatsApp Group'])
                client_category = get_col_val(row, ['Client Category', 'Category'])
                client_group, crp_of_group, reffered_by = resolve_client_group_parent(
                    cursor,
                    client_group,
                    crp_of_group,
                    reffered_by
                )
                
                if client_id:
                    # Update existing client if ID is present
                    cursor.execute('''
                        UPDATE client_master 
                        SET client_name=?, phone=?, email=?, gstin=?, client_group=?, 
                            crp_of_group=?, reffered_by=?, whatapp_group=?, client_category=?
                        WHERE id=?
                    ''', (client_name, phone, email, gstin, client_group, crp_of_group, reffered_by, whatapp_group, client_category, client_id))
                    imported_count += 1
                else:
                    # Insert as new if client name doesn't exist
                    cursor.execute('SELECT COUNT(*) FROM client_master WHERE client_name = ?', (client_name,))
                    if cursor.fetchone()[0] == 0:
                        cursor.execute('INSERT INTO client_master (client_name, phone, email, gstin, client_group, crp_of_group, reffered_by, whatapp_group, client_category) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                                       (client_name, phone, email, gstin, client_group, crp_of_group, reffered_by, whatapp_group, client_category))
                        imported_count += 1
            conn.commit()
            flash(f"{imported_count} clients processed (updated/added) successfully!", 'success')
        except Exception as e:
            flash(f"Error importing client master: {e}", 'error')
        finally:
            conn.close()
            if os.path.exists(temp_path): os.remove(temp_path) # Clean up temp file
    else:
        flash("Invalid file type. Please upload an .xlsx file.", 'error')
    return redirect(url_for('client_master_view'))

@app.route('/clear', methods=['POST'])
def clear_report():
    """Clears all data from the database and removes source data files."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM billing_report')
    for row in cursor.fetchall():
        row_dict = sqlite_record_to_dict(row)
        log_deleted_record(
            cursor,
            'billing_report',
            row_dict.get('id'),
            'Billing Report',
            row_dict.get('party_name') or row_dict.get('ref_no') or f"Bill #{row_dict.get('id')}",
            row_dict,
            ' | '.join(part for part in [
                row_dict.get('short_name') or row_dict.get('firm_name') or '',
                row_dict.get('bill_date') or '',
                row_dict.get('ref_no') or '',
                'Deleted by Clear Report',
            ] if part),
        )
    cursor.execute('DELETE FROM billing_report')
    conn.commit()
    conn.close()
    # Clean up source files to prevent re-import on server restart
    for name in ['data', 'input']:
        for ext in ['.csv', '.xlsx']:
            p = os.path.join(BASE_DIR, f'{name}{ext}')
            if os.path.exists(p):
                try:
                    os.remove(p)
                except OSError:
                    # File might be open in Excel; we skip deletion but the DB is already cleared
                    continue
    flash("Data deleted successfully!")
    return redirect(url_for('index'))

@app.route('/download/billing-template')
def download_billing_template():
    """Generates a template for the Billing Report."""
    output = io.BytesIO()
    # Create a DataFrame where the first row is the Firm Name placeholder
    # and the second row contains example headers or data
    data = [
        ["[Enter Your Firm Name Here]", "", "", ""],
        ["2024-01-01", "REF001", "Example Party", 5000.00]
    ]
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False)
    
    output.seek(0)
    return send_file(output, 
                     download_name="billing_template.xlsx", 
                     as_attachment=True, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download/client-template')
def download_client_template():
    """Generates a template for the Client Master."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    dropdown_options = get_client_master_dropdown_options(cursor)
    conn.close()

    output = io.BytesIO()
    # Standard headers expected by the import logic
    df = pd.DataFrame(columns=['Client Name', 'Client Group', 'CRP of Group', 'Referred By', 'Whatapp Group', 'Phone', 'Email', 'GSTIN'])
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Clients')
        add_client_master_dropdowns(writer, 'Clients', dropdown_options)
    
    output.seek(0)
    return send_file(output, 
                     download_name="client_master_template.xlsx", 
                     as_attachment=True, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

try:
    init_db()
except Exception as exc:
    print(f"Debtor report database initialization skipped: {exc}")

if __name__ == '__main__':
    app.run(debug=True, port=5005)
