"""
Tally Sales Register - Standalone GUI App
==========================================

Connects DIRECTLY to Tally Prime over its XML/HTTP interface (no Claude,
no MCP, no internet needed) and generates a combined, columnar Sales
Register across one or more open Tally companies, into a single Excel
sheet with a "Company" column.

REQUIREMENTS (one-time setup)
------------------------------
1. Tally Prime must be running with the companies you want loaded/open.
2. Enable Tally's XML server:
     Tally > F1 (Help) > Settings > Connectivity
     - "Tally Prime is Action as"  -> Server
     - Port -> 9000  (default)
3. Python 3.8+ with these packages:
     pip install requests openpyxl

HOW TO RUN
----------
    python tally_sales_register_app.py

A window opens: click "Refresh companies" to list open Tally companies,
tick the ones you want, set a From/To date, click "Generate Excel".
A Save-As dialog lets you choose where to save the .xlsx file.

NOTES ON HOW IT TALKS TO TALLY
-------------------------------
Tally exposes a simple HTTP endpoint at http://localhost:9000 that
accepts an XML request body and returns an XML response. This script
builds small ad-hoc TDL ("Tally Definition Language") collection
requests to:
  1. List the companies currently open in Tally.
  2. Pull all vouchers of type "Sales" for a company within a date
     range, including each voucher's ledger entries (so we can see
     which income/sales ledgers were credited and for how much).

The "Sales Accounts" ledgers (whatever they're called in each
company's chart of accounts) become the pivoted columns; each
voucher becomes one row.
"""

import datetime as dt
import queue
import threading
import tkinter as tk
import xml.etree.ElementTree as ET
from tkinter import ttk, messagebox, filedialog

import requests

TALLY_URL = "http://localhost:9000"
TIMEOUT = 60


# --------------------------------------------------------------------------
# Tally XML/HTTP client
# --------------------------------------------------------------------------

class TallyError(Exception):
    pass


def _post_xml(xml_request: str) -> str:
    try:
        resp = requests.post(TALLY_URL, data=xml_request.encode("utf-8"), timeout=TIMEOUT)
        resp.raise_for_status()
        return resp.text
    except requests.exceptions.ConnectionError as e:
        raise TallyError(
            "Could not connect to Tally at {}. Make sure Tally Prime is running "
            "and the XML server is enabled (F1 > Settings > Connectivity > "
            "'Tally Prime is Action as' = Server, Port 9000).".format(TALLY_URL)
        ) from e
    except requests.exceptions.RequestException as e:
        raise TallyError("Error talking to Tally: {}".format(e)) from e


def _clean_xml(raw: str) -> str:
    """Tally's XML output is sometimes not fully well-formed (stray control
    chars, unescaped & etc). Do a light cleanup before parsing."""
    import re
    raw = raw.replace("\x04", "").replace("\x00", "")
    raw = raw.replace("&#4;", "")
    # Tally often returns UDF:* elements without an xmlns:UDF declaration.
    # ElementTree then rejects an otherwise valid export as an "unbound
    # prefix".  These UDF fields are not used by this report, so make their
    # tag names XML-safe before parsing.
    raw = re.sub(r"(?<=<)(/?)UDF:", r"\1UDF_", raw)
    # Escape bare ampersands that aren't part of a valid entity
    raw = re.sub(r"&(?!(amp|lt|gt|quot|apos|#\d+|#x[0-9a-fA-F]+);)", "&amp;", raw)
    return raw


def get_open_companies() -> list:
    """Returns the list of company names currently open in Tally."""
    request_xml = """
<ENVELOPE>
  <HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>Export</TALLYREQUEST>
    <TYPE>Collection</TYPE>
    <ID>List of Companies</ID>
  </HEADER>
  <BODY>
    <DESC>
      <STATICVARIABLES>
        <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
      </STATICVARIABLES>
      <TDL>
        <TDLMESSAGE>
          <COLLECTION NAME="List of Companies" ISMODIFY="No">
            <TYPE>Company</TYPE>
            <FETCH>NAME</FETCH>
          </COLLECTION>
        </TDLMESSAGE>
      </TDL>
    </DESC>
  </BODY>
</ENVELOPE>
""".strip()

    raw = _post_xml(request_xml)
    raw = _clean_xml(raw)
    try:
        root = ET.fromstring(raw)
    except ET.ParseError as e:
        raise TallyError("Could not parse Tally's company list response: {}".format(e))

    names = []
    for name_el in root.iter("NAME"):
        if name_el.text:
            names.append(name_el.text.strip())
    # de-dup, keep order
    seen = set()
    result = []
    for n in names:
        if n not in seen:
            seen.add(n)
            result.append(n)
    return result


def _tally_date(d: dt.date) -> str:
    return d.strftime("%Y%m%d")


def get_sales_vouchers(company: str, from_date: dt.date, to_date: dt.date) -> list:
    """
    Returns a list of dicts, one per (voucher, income-ledger) line:
      {date, voucher_type, voucher_number, party_name, ledger_name, amount}

    'party_name' is the voucher's party ledger (the debtor / client).
    'ledger_name' is each non-party ledger entry on the voucher (the
    income/sales ledger that was credited) and 'amount' its value.
    """
    request_xml = """
<ENVELOPE>
  <HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>Export</TALLYREQUEST>
    <TYPE>Collection</TYPE>
    <ID>SalesVouchersCollection</ID>
  </HEADER>
  <BODY>
    <DESC>
      <STATICVARIABLES>
        <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
        <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
        <SVFROMDATE>{from_date}</SVFROMDATE>
        <SVTODATE>{to_date}</SVTODATE>
      </STATICVARIABLES>
      <TDL>
        <TDLMESSAGE>
          <COLLECTION NAME="SalesVouchersCollection" ISMODIFY="No">
            <TYPE>Voucher</TYPE>
            <FILTER>IsSalesVoucher</FILTER>
            <FETCH>DATE, VOUCHERNUMBER, VOUCHERTYPENAME, PARTYLEDGERNAME, LEDGERENTRIES.LIST</FETCH>
          </COLLECTION>
          <SYSTEM TYPE="Formulae" NAME="IsSalesVoucher">$VoucherTypeName Contains "Sale"</SYSTEM>
        </TDLMESSAGE>
      </TDL>
    </DESC>
  </BODY>
</ENVELOPE>
""".strip().format(
        company=_xml_escape(company),
        from_date=_tally_date(from_date),
        to_date=_tally_date(to_date),
    )

    raw = _post_xml(request_xml)
    raw = _clean_xml(raw)
    try:
        root = ET.fromstring(raw)
    except ET.ParseError as e:
        raise TallyError(
            "Could not parse Tally's voucher response for '{}': {}".format(company, e)
        )

    rows = []
    for voucher in root.iter("VOUCHER"):
        date_raw = _text(voucher, "DATE")
        voucher_date = _parse_tally_date(date_raw)
        # Some Tally Prime installations do not honour SVFROMDATE/SVTODATE
        # for a custom Voucher collection.  Apply the requested period here
        # as the authoritative filter so the report never includes vouchers
        # after the selected end date.
        if voucher_date is None or not (from_date <= voucher_date <= to_date):
            continue
        voucher_type = _text(voucher, "VOUCHERTYPENAME")
        voucher_number = _text(voucher, "VOUCHERNUMBER")
        party_name = _text(voucher, "PARTYLEDGERNAME")

        entries = voucher.findall("./LEDGERENTRIES.LIST")
        for entry in entries:
            ledger_name = _text(entry, "LEDGERNAME")
            amount_raw = _text(entry, "AMOUNT")
            if ledger_name is None or amount_raw is None:
                continue
            try:
                amount = float(amount_raw)
            except ValueError:
                continue

            # Skip the party's own ledger entry (that's the debit side /
            # debtor) - we only want the income/sales ledger credit lines.
            if party_name and ledger_name.strip() == party_name.strip():
                continue

            rows.append({
                "date": voucher_date,
                "voucher_type": voucher_type or "",
                "voucher_number": voucher_number or "",
                "party_name": party_name or "",
                "ledger_name": ledger_name.strip(),
                # Income ledger credits come through as negative in Tally's
                # raw export convention; normalize to a positive sales amount.
                "amount": abs(amount),
            })

    return rows


def _xml_escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _text(el, tag):
    child = el.find(tag)
    if child is None:
        return None
    return child.text


def _parse_tally_date(raw):
    if not raw:
        return None
    raw = raw.strip()
    try:
        return dt.datetime.strptime(raw, "%Y%m%d").date()
    except ValueError:
        return raw


def _parse_tally_amount(raw) -> float:
    """Parse normal and foreign-currency amounts returned by Tally.

    Tally can return an amount such as ``-$250.00 @ Rs. 88.13/$ =
    -Rs. 22032.50``.  For reports in INR, the value after ``=`` is the
    authoritative base-currency amount.
    """
    import re

    text = (raw or "").strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        pass

    amount_text = text.rsplit("=", 1)[-1].strip()
    numbers = re.findall(r"\d+(?:\.\d+)?", amount_text)
    if not numbers:
        raise ValueError("Invalid Tally amount: {!r}".format(raw))
    value = float(numbers[-1])
    return -value if amount_text.startswith("-") else value


def _get_parent_map(company: str, object_type: str, collection_name: str) -> dict:
    """Return {name: parent} for Tally Ledger or Group masters."""
    request_xml = """
<ENVELOPE>
  <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Collection</TYPE><ID>{collection_name}</ID></HEADER>
  <BODY><DESC>
    <STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT><SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY></STATICVARIABLES>
    <TDL><TDLMESSAGE>
      <COLLECTION NAME="{collection_name}" ISMODIFY="No"><TYPE>{object_type}</TYPE><FETCH>NAME,PARENT</FETCH></COLLECTION>
    </TDLMESSAGE></TDL>
  </DESC></BODY>
</ENVELOPE>
""".strip().format(
        company=_xml_escape(company), object_type=object_type, collection_name=collection_name
    )
    try:
        root = ET.fromstring(_clean_xml(_post_xml(request_xml)))
    except ET.ParseError as e:
        raise TallyError("Could not parse {} masters for '{}': {}".format(object_type, company, e))

    result = {}
    for element in root.findall(".//COLLECTION/{}".format(object_type.upper())):
        name = element.get("NAME")
        if name:
            result[name.strip()] = (_text(element, "PARENT") or "").strip()
    return result


def _get_bills_receivable(company: str, as_of_date: dt.date) -> list:
    """Read Tally's built-in Bills Receivable report as of one date."""
    request_xml = """
<ENVELOPE>
  <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Data</TYPE><ID>Bills Receivable</ID></HEADER>
  <BODY><DESC><STATICVARIABLES>
    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
    <SVTODATE>{as_of_date}</SVTODATE>
  </STATICVARIABLES></DESC></BODY>
</ENVELOPE>
""".strip().format(company=_xml_escape(company), as_of_date=_tally_date(as_of_date))
    try:
        root = ET.fromstring(_clean_xml(_post_xml(request_xml)))
    except ET.ParseError as e:
        raise TallyError("Could not parse Bills Receivable for '{}': {}".format(company, e))

    rows = []
    current = None
    for element in root:
        if element.tag == "BILLFIXED":
            current = (
                (_text(element, "BILLDATE") or "").strip(),
                (_text(element, "BILLREF") or "").replace("\r", "").replace("\n", "").strip(),
                (_text(element, "BILLPARTY") or "").strip(),
            )
        elif element.tag == "BILLCL" and current:
            try:
                amount = _parse_tally_amount(element.text)
            except ValueError:
                amount = 0.0
            rows.append((*current, amount))
            current = None
    return rows


def _get_bill_opening_amounts(company: str) -> dict:
    """Return Bill-master balances keyed by reference number.

    A reference can occur more than once (for example, a receivable bill and
    a positive adjustment). Keep every record so its sign can be matched to
    the Bills Receivable entry instead of overwriting one with another.
    """
    request_xml = """
<ENVELOPE>
  <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Collection</TYPE><ID>BillOpeningBalances</ID></HEADER>
  <BODY><DESC>
    <STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT><SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY></STATICVARIABLES>
    <TDL><TDLMESSAGE>
      <COLLECTION NAME="BillOpeningBalances" ISMODIFY="No"><TYPE>Bill</TYPE><FETCH>NAME,OPENINGBALANCE,CLOSINGBALANCE</FETCH></COLLECTION>
    </TDLMESSAGE></TDL>
  </DESC></BODY>
</ENVELOPE>
""".strip().format(company=_xml_escape(company))
    try:
        root = ET.fromstring(_clean_xml(_post_xml(request_xml)))
    except ET.ParseError as e:
        raise TallyError("Could not parse Bill masters for '{}': {}".format(company, e))

    amounts = {}
    for bill in root.findall(".//COLLECTION/BILL"):
        name = (bill.get("NAME") or "").strip()
        if not name:
            continue
        try:
            opening = _parse_tally_amount(_text(bill, "OPENINGBALANCE"))
            closing = _parse_tally_amount(_text(bill, "CLOSINGBALANCE"))
            amounts.setdefault(name, []).append((opening, closing))
        except ValueError:
            continue
    return amounts


def _matching_bill_opening(candidates, receivable_closing):
    """Choose the Bill record belonging to the receivable, not an adjustment."""
    if not candidates:
        return 0.0
    same_sign = [item for item in candidates if item[1] * receivable_closing > 0]
    matches = same_sign or candidates
    # If more than one record has the same sign, the one whose current Bill
    # balance is closest to the receivable report's balance is the best match.
    opening, _closing = min(matches, key=lambda item: abs(abs(item[1]) - abs(receivable_closing)))
    # Tally stores debtor balances as negative (debit) and credit-side bill
    # adjustments as positive.  Reverse that sign for the report total: a
    # normal debtor bill contributes positively, while a credit adjustment
    # reduces the Opening Amount total.
    return -opening


def _parse_bill_date(raw):
    try:
        return dt.datetime.strptime(raw, "%d-%b-%y").date()
    except ValueError:
        return raw


def get_billwise_debtor_rows(company: str, from_date: dt.date, to_date: dt.date) -> list:
    """Outstanding bills for ledger groups under Sundry Debtors.

    Opening Amount is the original bill amount from Tally's Bill master;
    Pending Amount is its balance at to_date. Only bills outstanding at
    to_date are included, which is the normal debtor-report view.
    """
    ledgers = _get_parent_map(company, "Ledger", "DebtorLedgerMasters")
    groups = _get_parent_map(company, "Group", "DebtorGroupMasters")

    def is_sundry_debtor(party):
        name = ledgers.get(party, "")
        seen = set()
        while name and name not in seen:
            if name == "Sundry Debtors":
                return True
            seen.add(name)
            name = groups.get(name, "")
        return False

    opening_amounts = _get_bill_opening_amounts(company)
    closing_rows = _get_bills_receivable(company, to_date)

    result = []
    for raw_date, ref, party, closing_amount in closing_rows:
        if not is_sundry_debtor(party):
            continue
        result.append({
            "company": company,
            "date": _parse_bill_date(raw_date),
            "reference_number": ref,
            "party_name": party,
            "group_name": ledgers.get(party, ""),
            "opening_amount": _matching_bill_opening(opening_amounts.get(ref), closing_amount),
            "pending_amount": abs(closing_amount),
        })
    return result


# --------------------------------------------------------------------------
# Cost Centre Breakup
#
# Tally's native "Cost Centre Breakup" (Display > Statement of Accounts >
# Cost Centre > Cost Centre Breakup) report is exported directly (same
# TYPE=Data technique as Bills Receivable) instead of being reconstructed
# from raw vouchers. This is important because:
#   1. Tally computes the Closing Balance itself, correctly including any
#      pre-period opening balance a cost centre carries on a Balance-Sheet
#      type ledger (something we cannot derive from period vouchers alone).
#   2. The report's raw XML export is a flat, unlabelled sequence: a
#      <DSPACCNAME><DSPDISPNAME>Group Name</DSPDISPNAME></DSPACCNAME>
#      marks the start of a group/particulars row, immediately followed by
#      one <DSPACCINFO> block per Cost Centre column (Debit/Credit/Closing),
#      in the SAME left-to-right order the columns appear on screen - with
#      no cost-centre name attached to each block. So we separately fetch
#      the Cost Centre master list (which Tally returns in the same
#      creation order used for the report's columns) and zip the two
#      together by position.
# --------------------------------------------------------------------------

def get_cost_centre_names(company: str, debug_dump_path: str = None) -> list:
    """Cost centre master names, in Tally's native (creation) order - this
    is the order the Cost Centre Breakup report's columns follow.

    If debug_dump_path is given, the raw XML response is also saved there
    (useful for troubleshooting if this ever returns an empty list even
    though Cost Centres exist in Tally)."""
    request_xml = """
<ENVELOPE>
  <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Collection</TYPE><ID>CCNames</ID></HEADER>
  <BODY><DESC>
    <STATICVARIABLES>
      <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
      <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
    </STATICVARIABLES>
    <TDL><TDLMESSAGE>
      <COLLECTION NAME="CCNames" ISMODIFY="No">
        <TYPE>CostCentre</TYPE>
        <FETCH>NAME</FETCH>
      </COLLECTION>
    </TDLMESSAGE></TDL>
  </DESC></BODY>
</ENVELOPE>
""".strip().format(company=_xml_escape(company))

    raw = _post_xml(request_xml)
    raw = _clean_xml(raw)

    if debug_dump_path:
        with open(debug_dump_path, "w", encoding="utf-8") as f:
            f.write(raw)

    try:
        root = ET.fromstring(raw)
    except ET.ParseError as e:
        raise TallyError("Could not parse cost centre list for '{}': {}".format(company, e))

    names = []
    # Search anywhere in the tree (not a fixed path) - Tally's exact
    # nesting for a plain Collection export has varied slightly across
    # versions, but the object tag itself is reliably "COSTCENTRE".
    for cc in root.iter("COSTCENTRE"):
        name_el = cc.find("NAME")
        if name_el is not None and name_el.text and name_el.text.strip():
            names.append(name_el.text.strip())
        elif cc.text and cc.text.strip() and cc.text.strip() not in names:
            # Some Tally versions return NAME as an attribute-less leaf
            # directly instead of a NAME child - fall back to that.
            names.append(cc.text.strip())
    return names


def dump_native_cost_centre_xml(company: str, from_date: dt.date, to_date: dt.date, out_path: str) -> str:
    """Save the raw 'Cost Centre Breakup' native report XML to out_path, for
    troubleshooting if the ID guess below ever stops matching a Tally
    version/menu wording. Returns the cleaned XML text too."""
    request_xml = """
<ENVELOPE>
  <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Data</TYPE><ID>Cost Centre Breakup</ID></HEADER>
  <BODY><DESC><STATICVARIABLES>
    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
    <SVFROMDATE>{from_date}</SVFROMDATE>
    <SVTODATE>{to_date}</SVTODATE>
  </STATICVARIABLES></DESC></BODY>
</ENVELOPE>
""".strip().format(company=_xml_escape(company), from_date=_tally_date(from_date), to_date=_tally_date(to_date))

    raw = _post_xml(request_xml)
    cleaned = _clean_xml(raw)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(cleaned)
    return cleaned


def get_cost_centre_breakup(company: str, from_date: dt.date, to_date: dt.date) -> list:
    """
    Returns rows: {company, group, cost_centre, debit, credit, closing}
    read directly from Tally's own native Cost Centre Breakup report -
    values (including Closing Balance) are exactly what Tally itself shows.
    """
    cost_centre_names = get_cost_centre_names(company)
    if not cost_centre_names:
        raise TallyError(
            "'{}' has no Cost Centres defined (or the master list could not be read).".format(company)
        )
    n = len(cost_centre_names)

    request_xml = """
<ENVELOPE>
  <HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST><TYPE>Data</TYPE><ID>Cost Centre Breakup</ID></HEADER>
  <BODY><DESC><STATICVARIABLES>
    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
    <SVFROMDATE>{from_date}</SVFROMDATE>
    <SVTODATE>{to_date}</SVTODATE>
  </STATICVARIABLES></DESC></BODY>
</ENVELOPE>
""".strip().format(company=_xml_escape(company), from_date=_tally_date(from_date), to_date=_tally_date(to_date))

    raw = _post_xml(request_xml)
    raw = _clean_xml(raw)
    try:
        root = ET.fromstring(raw)
    except ET.ParseError as e:
        raise TallyError(
            "Could not parse the Cost Centre Breakup report for '{}': {}. "
            "Use the debug export button and send the file so this can be fixed.".format(company, e)
        )

    result = []
    current_group = None
    idx_in_group = 0
    for element in root.iter():
        if element.tag == "DSPACCNAME":
            name_el = element.find("DSPDISPNAME")
            current_group = (name_el.text or "").strip() if name_el is not None and name_el.text else None
            idx_in_group = 0
        elif element.tag == "DSPACCINFO" and current_group is not None:
            if idx_in_group >= n:
                # More blocks than known cost centres - master list is out
                # of sync with the report; stop attributing names for the
                # overflow rather than mislabel them.
                idx_in_group += 1
                continue
            cc_name = cost_centre_names[idx_in_group]
            idx_in_group += 1

            dr_el = element.find("./DSPDRAMT/DSPDRAMTA")
            cr_el = element.find("./DSPCRAMT/DSPCRAMTA")
            cl_el = element.find("./DSPCLAMT/DSPCLAMTA")
            dr_text = (dr_el.text or "").strip() if dr_el is not None else ""
            cr_text = (cr_el.text or "").strip() if cr_el is not None else ""
            cl_text = (cl_el.text or "").strip() if cl_el is not None else ""

            try:
                debit = abs(_parse_tally_amount(dr_text)) if dr_text else 0.0
            except ValueError:
                debit = 0.0
            try:
                credit = abs(_parse_tally_amount(cr_text)) if cr_text else 0.0
            except ValueError:
                credit = 0.0
            try:
                closing = abs(_parse_tally_amount(cl_text)) if cl_text else 0.0
            except ValueError:
                closing = 0.0

            if debit or credit or closing:
                result.append({
                    "company": company,
                    "group": current_group,
                    "cost_centre": cc_name,
                    "debit": debit,
                    "credit": credit,
                    "closing": closing,
                })

    if not result:
        raise TallyError(
            "Tally returned the Cost Centre Breakup report for '{}' but it had no "
            "non-zero figures for {} to {}. Double check the period, or use the "
            "debug export button and send the file.".format(company, from_date, to_date)
        )

    return result


def build_cost_centre_breakup_workbook(all_rows, from_date, to_date, out_path):
    """
    all_rows: list of dicts with keys company, group, cost_centre, debit,
    credit, closing (see get_cost_centre_breakup).
    Writes one sheet: Company | Particulars(Group) | then 3 columns
    (Debit / Credit / Closing Balance) per distinct Cost Centre across all
    companies, plus a Grand Total row per company.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    if not all_rows:
        raise ValueError(
            "No cost-centre-tagged transactions found for the selected companies/period."
        )

    companies = sorted({r["company"] for r in all_rows})
    cost_centres = sorted({r["cost_centre"] for r in all_rows})

    grid = {}
    groups_by_company = {}
    for r in all_rows:
        key = (r["company"], r["group"])
        grid.setdefault(key, {})[r["cost_centre"]] = (r["debit"], r["credit"], r["closing"])
        groups_by_company.setdefault(r["company"], set()).add(r["group"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Centre Breakup"
    ws.sheet_view.showGridLines = False

    font_name = "Arial"
    header_font = Font(name=font_name, bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(name=font_name, bold=True, size=14)
    subtitle_font = Font(name=font_name, size=10, italic=True)
    normal_font = Font(name=font_name, size=10)
    bold_font = Font(name=font_name, bold=True, size=10)
    total_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    n_cc = len(cost_centres)
    total_cols = 2 + n_cc * 3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value="Cost Centre Breakup - All Companies").font = title_font
    ws.cell(row=1, column=1).alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(
        row=2, column=1,
        value="{} to {}  |  Debit/Credit = period movement; Closing Balance excludes any pre-period "
              "opening balance carried on a cost centre (see README)".format(
                  from_date.strftime("%d-%b-%Y"), to_date.strftime("%d-%b-%Y")
              )
    ).font = subtitle_font
    ws.cell(row=2, column=1).alignment = center

    header_row = 4
    sub_header_row = 5

    ws.merge_cells(start_row=header_row, start_column=1, end_row=sub_header_row, end_column=1)
    ws.cell(row=header_row, column=1, value="Company")
    ws.merge_cells(start_row=header_row, start_column=2, end_row=sub_header_row, end_column=2)
    ws.cell(row=header_row, column=2, value="Particulars")

    for j, cc in enumerate(cost_centres):
        start_col = 3 + j * 3
        ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=start_col + 2)
        ws.cell(row=header_row, column=start_col, value=cc)
        for k, sub in enumerate(["Debit", "Credit", "Closing Balance"]):
            ws.cell(row=sub_header_row, column=start_col + k, value=sub)

    for row in ws[header_row:sub_header_row]:
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    data_start = sub_header_row + 1
    r = data_start
    for company in companies:
        groups = sorted(groups_by_company.get(company, []))
        company_start_row = r
        for group in groups:
            ws.cell(row=r, column=1, value=company)
            ws.cell(row=r, column=2, value=group)
            row_data = grid.get((company, group), {})
            for j, cc in enumerate(cost_centres):
                start_col = 3 + j * 3
                debit, credit, closing = row_data.get(cc, (0.0, 0.0, 0.0))
                ws.cell(row=r, column=start_col, value=debit or None)
                ws.cell(row=r, column=start_col + 1, value=credit or None)
                ws.cell(row=r, column=start_col + 2, value=closing or None)
                for c in range(start_col, start_col + 3):
                    ws.cell(row=r, column=c).number_format = '#,##0.00;(#,##0.00);"-"'
                    ws.cell(row=r, column=c).alignment = right
            for c in range(1, total_cols + 1):
                ws.cell(row=r, column=c).font = normal_font
                ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=2).alignment = left
            r += 1

        ws.cell(row=r, column=2, value="{} - Grand Total".format(company))
        for j in range(n_cc):
            start_col = 3 + j * 3
            for k in range(3):
                col = start_col + k
                col_letter = get_column_letter(col)
                cell = ws.cell(
                    row=r, column=col,
                    value="=SUM({0}{1}:{0}{2})".format(col_letter, company_start_row, r - 1)
                )
                cell.number_format = '#,##0.00;(#,##0.00);"-"'
                cell.alignment = right
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).font = bold_font
            ws.cell(row=r, column=c).fill = total_fill
            ws.cell(row=r, column=c).border = border
        r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 26
    for j in range(n_cc):
        for k in range(3):
            ws.column_dimensions[get_column_letter(3 + j * 3 + k)].width = 15

    ws.freeze_panes = ws.cell(row=data_start, column=3)
    wb.save(out_path)


# --------------------------------------------------------------------------
# Excel builder
# --------------------------------------------------------------------------

def build_combined_workbook(all_rows, from_date, to_date, out_path):
    """
    all_rows: list of dicts with keys:
      company, date, voucher_type, voucher_number, party_name, ledger_name, amount
    Writes one sheet: Company | Date | Voucher Type | Voucher No | Party Name
    | <one column per distinct ledger_name across all companies> | Total
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    if not all_rows:
        raise ValueError("No sales vouchers found for the selected companies/period.")

    # unique vouchers (company + voucher_number + date + party + type)
    voucher_keys = []
    seen = set()
    for r in all_rows:
        key = (r["company"], r["date"], r["voucher_number"])
        if key not in seen:
            seen.add(key)
            voucher_keys.append(key)

    voucher_meta = {}
    for r in all_rows:
        key = (r["company"], r["date"], r["voucher_number"])
        if key not in voucher_meta:
            voucher_meta[key] = {
                "company": r["company"],
                "date": r["date"],
                "voucher_type": r["voucher_type"],
                "voucher_number": r["voucher_number"],
                "party_name": r["party_name"],
            }

    voucher_keys.sort(key=lambda k: (voucher_meta[k]["company"], voucher_meta[k]["date"] or dt.date.min, voucher_meta[k]["voucher_number"]))

    ledgers = sorted({r["ledger_name"] for r in all_rows})

    pivot = {}
    for r in all_rows:
        key = (r["company"], r["date"], r["voucher_number"])
        pivot[(key, r["ledger_name"])] = pivot.get((key, r["ledger_name"]), 0) + r["amount"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Register"

    font_name = "Arial"
    header_font = Font(name=font_name, bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(name=font_name, bold=True, size=14)
    subtitle_font = Font(name=font_name, size=10, italic=True)
    normal_font = Font(name=font_name, size=10)
    bold_font = Font(name=font_name, bold=True, size=10)
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    n_ledger_cols = len(ledgers)
    total_cols = 5 + n_ledger_cols + 1  # Company, Date, VchType, VchNo, Party + ledgers + Total

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value="Combined Sales Register - All Companies").font = title_font
    ws.cell(row=1, column=1).alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1,
            value="Columnar Format | {} to {}".format(from_date.strftime("%d-%b-%Y"), to_date.strftime("%d-%b-%Y"))
            ).font = subtitle_font
    ws.cell(row=2, column=1).alignment = center

    header_row = 4
    headers = ["Company", "Date", "Voucher Type", "Voucher No.", "Party Name"] + ledgers + ["Total"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    data_start = header_row + 1
    for i, key in enumerate(voucher_keys):
        r = data_start + i
        meta = voucher_meta[key]
        ws.cell(row=r, column=1, value=meta["company"])
        date_val = meta["date"]
        if isinstance(date_val, dt.date):
            ws.cell(row=r, column=2, value=date_val).number_format = "DD-MMM-YYYY"
        else:
            ws.cell(row=r, column=2, value=date_val)
        ws.cell(row=r, column=3, value=meta["voucher_type"])
        ws.cell(row=r, column=4, value=meta["voucher_number"])
        ws.cell(row=r, column=5, value=meta["party_name"])

        for j, ledger in enumerate(ledgers):
            col = 6 + j
            val = pivot.get((key, ledger))
            cell = ws.cell(row=r, column=col, value=val if val else None)
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
            cell.alignment = right

        total_col = 6 + n_ledger_cols
        first_col_letter = get_column_letter(6)
        last_col_letter = get_column_letter(5 + n_ledger_cols)
        ws.cell(row=r, column=total_col,
                value="=SUM({0}{2}:{1}{2})".format(first_col_letter, last_col_letter, r)
                ).number_format = '#,##0.00;(#,##0.00);"-"'

        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).font = normal_font
            ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=1).alignment = left
        ws.cell(row=r, column=5).alignment = left

    data_end = data_start + len(voucher_keys) - 1
    total_row = data_end + 1
    ws.cell(row=total_row, column=5, value="Column Total").font = bold_font
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
    for j in range(n_ledger_cols + 1):
        col = 6 + j
        col_letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col,
                        value="=SUM({0}{1}:{0}{2})".format(col_letter, data_start, data_end))
        cell.font = bold_font
        cell.number_format = '#,##0.00;(#,##0.00);"-"'
        cell.border = border
        cell.alignment = right
    for c in range(1, 6):
        ws.cell(row=total_row, column=c).border = border

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 30
    for j in range(n_ledger_cols):
        ws.column_dimensions[get_column_letter(6 + j)].width = 16
    ws.column_dimensions[get_column_letter(6 + n_ledger_cols)].width = 14

    ws.freeze_panes = ws.cell(row=data_start, column=6)
    wb.save(out_path)


def build_billwise_debtor_workbook(rows, from_date, to_date, out_path):
    """Write a combined, bill-wise debtor report for one or more companies."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    if not rows:
        raise ValueError("No Sundry Debtors bills were outstanding on the selected end date.")

    rows.sort(key=lambda r: (
        r["company"],
        r["date"] if isinstance(r["date"], dt.date) else dt.date.max,
        r["party_name"],
        r["reference_number"],
    ))
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill Wise Debtor"
    ws.sheet_view.showGridLines = False

    headers = ["Company", "Date", "Ref No.", "Party Name", "GroupName", "Opening Amount", "Pending Amount"]
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    title_font = Font(name="Arial", bold=True, size=14)
    subtitle_font = Font(name="Arial", size=10, italic=True)
    normal_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    total_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A1:G1")
    ws["A1"] = "Bill Wise Debtor Report - Sundry Debtors Only"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.merge_cells("A2:G2")
    ws["A2"] = "Opening Amount = original bill amount | Pending as on {}".format(
        to_date.strftime("%d-%b-%Y")
    )
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center

    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    first_data_row = 5
    for index, item in enumerate(rows, start=first_data_row):
        values = [
            item["company"], item["date"], item["reference_number"], item["party_name"],
            item["group_name"], item["opening_amount"], item["pending_amount"],
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=index, column=column, value=value)
            cell.font = normal_font
            cell.border = border
        ws.cell(row=index, column=2).number_format = "DD-MMM-YYYY"
        for column in (6, 7):
            # Show both debit and credit bill amounts without a minus sign,
            # as Tally's Bills Receivable display does. The stored sign is
            # retained so Excel totals match Tally's calculation.
            ws.cell(row=index, column=column).number_format = '#,##0.00;#,##0.00;"-"'
            ws.cell(row=index, column=column).alignment = right

    last_data_row = first_data_row + len(rows) - 1
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=5, value="Total")
    for column in (6, 7):
        letter = "F" if column == 6 else "G"
        cell = ws.cell(row=total_row, column=column,
                       value="=SUM({0}{1}:{0}{2})".format(letter, first_data_row, last_data_row))
        cell.number_format = '#,##0.00;#,##0.00;"-"'
        cell.alignment = right
    for column in range(1, 8):
        cell = ws.cell(row=total_row, column=column)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border

    for letter, width in zip("ABCDEFG", (32, 14, 28, 42, 30, 18, 18)):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:G{}".format(last_data_row)
    wb.save(out_path)


# --------------------------------------------------------------------------
# GUI
# --------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tally - Combined Sales Register")
        self.geometry("940x650")
        self.minsize(760, 560)
        self.resizable(True, True)

        self.company_vars = {}
        self.work_queue = queue.Queue()

        self._build_ui()
        self.after(200, self._poll_queue)

    def _build_ui(self):
        style = ttk.Style(self)
        base_font = ("Segoe UI", 10)
        bold_font = ("Segoe UI", 10, "bold")
        style.configure(".", font=base_font)
        style.configure("TButton", font=base_font, padding=4)
        style.configure("TLabel", font=base_font)
        style.configure("TLabelframe.Label", font=bold_font)
        style.configure("TCheckbutton", font=base_font)
        style.configure("TEntry", font=base_font)
        main_area = ttk.Frame(self)
        main_area.pack(fill="both", expand=True, padx=10, pady=10)
        reports = ttk.LabelFrame(main_area, text="Reports")
        reports.pack(side="left", fill="y", padx=(0, 10))
        content = ttk.Frame(main_area)
        content.pack(side="left", fill="both", expand=True)
        pad = {"padx": 2, "pady": 4}

        top = ttk.Frame(content)
        top.pack(fill="x", **pad)
        ttk.Label(top, text="Tally companies", font=("Arial", 11, "bold")).pack(side="left")
        ttk.Button(top, text="Refresh companies", command=self.refresh_companies).pack(side="right")

        list_frame = ttk.Frame(content, relief="groove", borderwidth=1)
        list_frame.pack(fill="both", expand=True, padx=2, pady=4)

        # Keep this area scrollable and compact so controls below it remain
        # visible even on a 1366x768 laptop display.
        canvas = tk.Canvas(list_frame, height=250)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        self.checklist_frame = ttk.Frame(canvas)
        self.checklist_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.checklist_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.canvas = canvas

        sel_frame = ttk.Frame(content)
        sel_frame.pack(fill="x", padx=2)
        ttk.Button(sel_frame, text="Select all", command=self.select_all).pack(side="left")
        ttk.Button(sel_frame, text="Clear", command=self.clear_all).pack(side="left", padx=6)

        period_frame = ttk.LabelFrame(content, text="Period")
        period_frame.pack(fill="x", padx=2, pady=8)

        today = dt.date.today()
        # Sales data is commonly entered earlier in the financial year.  A
        # financial-year default avoids a misleading empty report caused by
        # limiting the first run to only the current calendar month.
        fy_start_year = today.year if today.month >= 4 else today.year - 1
        default_from = dt.date(fy_start_year, 4, 1).strftime("%Y-%m-%d")
        default_to = today.strftime("%Y-%m-%d")

        ttk.Label(period_frame, text="From (YYYY-MM-DD)").grid(row=0, column=0, padx=8, pady=8, sticky="w")
        self.from_entry = ttk.Entry(period_frame, width=14)
        self.from_entry.insert(0, default_from)
        self.from_entry.grid(row=0, column=1, padx=8, pady=8)

        ttk.Label(period_frame, text="To (YYYY-MM-DD)").grid(row=0, column=2, padx=8, pady=8, sticky="w")
        self.to_entry = ttk.Entry(period_frame, width=14)
        self.to_entry.insert(0, default_to)
        self.to_entry.grid(row=0, column=3, padx=8, pady=8)

        self.status_var = tk.StringVar(value="Click 'Refresh companies' to connect to Tally.")
        ttk.Label(content, textvariable=self.status_var, foreground="#555").pack(fill="x", padx=2, pady=3)

        self.progress = ttk.Progressbar(self, mode="indeterminate")
        self.progress.pack(fill="x", padx=2, pady=3)

        reports.columnconfigure(0, weight=1)
        reports.columnconfigure(1, weight=1)
        ttk.Button(reports, text="Sales Register",
                   command=self.generate).grid(row=0, column=0, sticky="ew", padx=6, pady=(8, 3))
        ttk.Button(reports, text="Debtor Report",
                   command=self.generate_debtor_report).grid(row=0, column=1, sticky="ew", padx=6, pady=(8, 3))
        ttk.Button(reports, text="Cost Centre",
                   command=self.generate_cost_centre_report).grid(row=1, column=0, sticky="ew", padx=6, pady=3)
        ttk.Button(reports, text="Debug XML",
                   command=self.export_debug_xml).grid(row=1, column=1, sticky="ew", padx=6, pady=3)

    # -- actions --------------------------------------------------------

    def refresh_companies(self):
        self.status_var.set("Connecting to Tally...")
        self.progress.start(10)

        def work():
            try:
                companies = get_open_companies()
                self.work_queue.put(("companies_ok", companies))
            except TallyError as e:
                self.work_queue.put(("error", str(e)))

        threading.Thread(target=work, daemon=True).start()

    def select_all(self):
        for var in self.company_vars.values():
            var.set(True)

    def clear_all(self):
        for var in self.company_vars.values():
            var.set(False)

    def generate(self):
        selected = [name for name, var in self.company_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("No companies selected", "Tick at least one company first.")
            return

        try:
            from_date = dt.datetime.strptime(self.from_entry.get().strip(), "%Y-%m-%d").date()
            to_date = dt.datetime.strptime(self.to_entry.get().strip(), "%Y-%m-%d").date()
        except ValueError:
            messagebox.showerror("Invalid date", "Use YYYY-MM-DD format for both dates.")
            return

        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile="Combined_Sales_Register.xlsx",
        )
        if not out_path:
            return

        self.status_var.set("Fetching vouchers from Tally for {} companies...".format(len(selected)))
        self.progress.start(10)

        def work():
            all_rows = []
            errors = []
            for company in selected:
                try:
                    rows = get_sales_vouchers(company, from_date, to_date)
                    for r in rows:
                        r["company"] = company
                    all_rows.extend(rows)
                except TallyError as e:
                    errors.append("{}: {}".format(company, e))

            if not all_rows and errors:
                self.work_queue.put(("error", "\n".join(errors)))
                return

            try:
                build_combined_workbook(all_rows, from_date, to_date, out_path)
            except Exception as e:
                self.work_queue.put(("error", "Failed to build Excel file: {}".format(e)))
                return

            msg = "Saved: {}".format(out_path)
            if errors:
                msg += "\n\nSome companies had issues:\n" + "\n".join(errors)
            self.work_queue.put(("done", msg))

        threading.Thread(target=work, daemon=True).start()

    def generate_debtor_report(self):
        selected = [name for name, var in self.company_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("No companies selected", "Tick at least one company first.")
            return

        try:
            from_date = dt.datetime.strptime(self.from_entry.get().strip(), "%Y-%m-%d").date()
            to_date = dt.datetime.strptime(self.to_entry.get().strip(), "%Y-%m-%d").date()
            if from_date > to_date:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid date", "Use valid YYYY-MM-DD dates and keep From on/before To.")
            return

        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile="Bill_Wise_Debtor_Report.xlsx",
        )
        if not out_path:
            return

        self.status_var.set("Fetching bill-wise debtor data from Tally for {} companies...".format(len(selected)))
        self.progress.start(10)

        def work():
            all_rows = []
            errors = []
            for company in selected:
                try:
                    all_rows.extend(get_billwise_debtor_rows(company, from_date, to_date))
                except TallyError as e:
                    errors.append("{}: {}".format(company, e))

            if not all_rows and errors:
                self.work_queue.put(("error", "\n".join(errors)))
                return
            try:
                build_billwise_debtor_workbook(all_rows, from_date, to_date, out_path)
            except Exception as e:
                self.work_queue.put(("error", "Failed to build Excel file: {}".format(e)))
                return

            msg = "Saved: {}".format(out_path)
            if errors:
                msg += "\n\nSome companies had issues:\n" + "\n".join(errors)
            self.work_queue.put(("debtor_done", msg))

        threading.Thread(target=work, daemon=True).start()

    def generate_cost_centre_report(self):
        selected = [name for name, var in self.company_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("No companies selected", "Tick at least one company first.")
            return

        try:
            from_date = dt.datetime.strptime(self.from_entry.get().strip(), "%Y-%m-%d").date()
            to_date = dt.datetime.strptime(self.to_entry.get().strip(), "%Y-%m-%d").date()
            if from_date > to_date:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid date", "Use valid YYYY-MM-DD dates and keep From on/before To.")
            return

        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile="Cost_Centre_Breakup.xlsx",
        )
        if not out_path:
            return

        self.status_var.set("Fetching cost centre data from Tally for {} companies...".format(len(selected)))
        self.progress.start(10)

        def work():
            all_rows = []
            errors = []
            for company in selected:
                try:
                    all_rows.extend(get_cost_centre_breakup(company, from_date, to_date))
                except TallyError as e:
                    errors.append("{}: {}".format(company, e))

            if not all_rows and errors:
                self.work_queue.put(("error", "\n".join(errors)))
                return
            try:
                build_cost_centre_breakup_workbook(all_rows, from_date, to_date, out_path)
            except Exception as e:
                self.work_queue.put(("error", "Failed to build Excel file: {}".format(e)))
                return

            msg = "Saved: {}".format(out_path)
            if errors:
                msg += "\n\nSome companies had issues:\n" + "\n".join(errors)
            self.work_queue.put(("cc_done", msg))

        threading.Thread(target=work, daemon=True).start()

    def export_debug_xml(self):
        selected = [name for name, var in self.company_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("No companies selected", "Tick at least one company first.")
            return
        company = selected[0]

        try:
            from_date = dt.datetime.strptime(self.from_entry.get().strip(), "%Y-%m-%d").date()
            to_date = dt.datetime.strptime(self.to_entry.get().strip(), "%Y-%m-%d").date()
            if from_date > to_date:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid date", "Use valid YYYY-MM-DD dates and keep From on/before To.")
            return

        out_path = filedialog.asksaveasfilename(
            defaultextension=".xml",
            filetypes=[("XML file", "*.xml")],
            initialfile="debug_cost_centre_breakup.xml",
        )
        if not out_path:
            return

        self.status_var.set("Exporting native Cost Centre Breakup XML for '{}'...".format(company))
        self.progress.start(10)

        def work():
            master_out_path = out_path.rsplit(".", 1)[0] + "_master_list.xml"
            try:
                xml_text = dump_native_cost_centre_xml(company, from_date, to_date, out_path)
                names = get_cost_centre_names(company, debug_dump_path=master_out_path)
            except TallyError as e:
                self.work_queue.put(("error", str(e)))
                return
            group_count = xml_text.count("<DSPACCNAME>")
            info_count = xml_text.count("<DSPACCINFO>")
            msg = ("Saved: {}\nSaved: {}\nGroups found: {}  |  DSPACCINFO blocks: {}  |  Cost centres in master list: {}\n\n"
                   "Send BOTH files to Claude so the cost centre parser can be fixed.").format(
                       out_path, master_out_path, group_count, info_count, len(names))
            self.work_queue.put(("debug_done", msg))

        threading.Thread(target=work, daemon=True).start()

    # -- queue polling (keeps Tally calls off the GUI thread) -----------

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.work_queue.get_nowait()
                self.progress.stop()
                if kind == "companies_ok":
                    self._populate_companies(payload)
                    self.status_var.set("Found {} open companies.".format(len(payload)))
                elif kind == "error":
                    self.status_var.set("Error.")
                    messagebox.showerror("Tally error", payload)
                elif kind == "done":
                    self.status_var.set("Done.")
                    messagebox.showinfo("Sales register generated", payload)
                elif kind == "debtor_done":
                    self.status_var.set("Done.")
                    messagebox.showinfo("Debtor report generated", payload)
                elif kind == "cc_done":
                    self.status_var.set("Done.")
                    messagebox.showinfo("Cost centre breakup generated", payload)
                elif kind == "debug_done":
                    self.status_var.set("Done.")
                    messagebox.showinfo("Debug XML exported", payload)
        except queue.Empty:
            pass
        self.after(200, self._poll_queue)

    def _populate_companies(self, companies):
        for widget in self.checklist_frame.winfo_children():
            widget.destroy()
        self.company_vars = {}
        for name in companies:
            var = tk.BooleanVar(value=True)
            chk = ttk.Checkbutton(self.checklist_frame, text=name, variable=var)
            chk.pack(anchor="w", padx=6, pady=2)
            self.company_vars[name] = var


if __name__ == "__main__":
    app = App()
    app.mainloop()
