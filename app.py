from flask import Flask, render_template, request, jsonify, send_file, Response, session, redirect, url_for, send_from_directory
import sqlite3
import os
from datetime import datetime
import json
import calendar
import zipfile
import io
import shutil
import re
import gzip
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import traceback
from jinja2 import ChoiceLoader, FileSystemLoader
from werkzeug.middleware.dispatcher import DispatcherMiddleware
from debtorapp import db_compat as debtor_db
import main_db_compat as main_db

app = Flask(__name__)
app.secret_key = 'super-secret-key-change-this-in-production'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = timedelta(days=7)
DATABASE = 'tasks.db'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEBTOR_REPORT_DATABASE = os.path.join(BASE_DIR, 'debtorapp', 'database.db')
LINKS_DIR = os.path.join(BASE_DIR, 'linksandothers')
LINKS_DATABASE = os.path.join(LINKS_DIR, 'mywork.db')
RESTORE_POINTS_DIR = os.path.join(BASE_DIR, 'restore_points')
DATABASE_BACKUP_DIR = os.path.join(BASE_DIR, 'database_backups')
app.jinja_loader = ChoiceLoader([app.jinja_loader, FileSystemLoader(BASE_DIR)])

ADMIN_EMAILS = ['arif.siddiqui@asija.in', 'admin@example.com']
PROJECT_PERIOD_START = datetime.strptime('2026-04-01', '%Y-%m-%d').date()
PROJECT_PERIOD_END = datetime.strptime('2027-03-31', '%Y-%m-%d').date()

PAGE_PERMISSIONS = [
    {'key': 'today_work', 'label': "Today's Work", 'path': '/today-work'},
    {'key': 'daily_debtor_report', 'label': 'Debtor Report', 'path': '/daily-debtor-report'},
    {'key': 'client_analytic', 'label': 'Client Analytic', 'path': '/client-analytic'},
    {'key': 'report_section', 'label': 'Report Section', 'path': '/report-section'},
    {'key': 'work_master', 'label': 'Task Master', 'path': '/work-master'},
    {'key': 'pending_approvals', 'label': 'Pending Approvals', 'path': '/pending-approvals'},
    {'key': 'request_logs', 'label': 'Request Logs', 'path': '/request-logs'},
    {'key': 'control_panel', 'label': 'Control Panel', 'path': '/control-panel'},
    {'key': 'assign_to_master', 'label': 'Department Name Master', 'path': '/assign-to-master'},
    {'key': 'user_power_management', 'label': 'User Powers', 'path': '/user-role-management'},
    {'key': 'user_management', 'label': 'User Accounts', 'path': '/user-management'},
    {'key': 'deleted_records', 'label': 'Deleted Records', 'path': '/deleted-records'},
    {'key': 'links_and_others', 'label': 'Links & Others', 'path': '/links-and-others'},
    {'key': 'restore_points', 'label': 'Restore Points', 'path': '/restore-points'},
]

SPECIAL_PERMISSIONS = [
    {'key': 'approve_requests', 'label': 'Approve / Reject Requests'},
]

LEGACY_USER_PERMISSIONS = [
    'today_work',
    'daily_debtor_report',
    'report_section',
    'work_master',
    'pending_approvals',
    'request_logs',
    'control_panel',
    'assign_to_master',
]

PROJECT_DATABASES = {
    'main_app': {
        'label': 'Main App / Tasks',
        'path': os.path.join(BASE_DIR, DATABASE),
        'filename': 'tasks.db',
    },
    'debtor_report': {
        'label': 'Daily Debtor Report',
        'path': DEBTOR_REPORT_DATABASE,
        'filename': 'database.db',
    },
    'links_others': {
        'label': 'Links & Others',
        'path': LINKS_DATABASE,
        'filename': 'mywork.db',
    },
}

def get_debtor_report_update_dates():
    dates = {
        'report_as_on_date': '',
        'last_bill_update_date': '',
        'last_receipt_update_date': '',
    }
    try:
        conn = debtor_db.connect(DEBTOR_REPORT_DATABASE)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS app_meta (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')
        cursor.execute(
            "SELECT key, value FROM app_meta WHERE key IN ('manual_report_as_on_date', 'manual_last_bill_update_date', 'manual_last_receipt_update_date')"
        )
        for key, value in cursor.fetchall():
            if key == 'manual_report_as_on_date':
                dates['report_as_on_date'] = value or ''
            elif key == 'manual_last_bill_update_date':
                dates['last_bill_update_date'] = value or ''
            elif key == 'manual_last_receipt_update_date':
                dates['last_receipt_update_date'] = value or ''
        conn.close()
    except debtor_db.Error:
        pass
    return dates

def set_debtor_report_update_dates(report_as_on_date, last_bill_update_date, last_receipt_update_date):
    conn = debtor_db.connect(DEBTOR_REPORT_DATABASE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS app_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    cursor.executemany('''
        INSERT INTO app_meta (key, value)
        VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    ''', (
        ('manual_report_as_on_date', report_as_on_date),
        ('manual_last_bill_update_date', last_bill_update_date),
        ('manual_last_receipt_update_date', last_receipt_update_date),
    ))
    conn.commit()
    conn.close()

@app.after_request
def set_response_headers(response):
    """Cache static assets, but keep pages private after logout."""
    response = gzip_response_if_supported(response)
    if request.endpoint == 'static' or request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=604800, immutable'
        response.headers.pop('Pragma', None)
        response.headers.pop('Expires', None)
    else:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response


def gzip_response_if_supported(response):
    if response.direct_passthrough:
        return response
    if response.status_code < 200 or response.status_code >= 300:
        return response
    if 'gzip' not in request.headers.get('Accept-Encoding', '').lower():
        return response
    if response.headers.get('Content-Encoding'):
        return response

    content_type = response.headers.get('Content-Type', '').lower()
    compressible_types = (
        'text/html',
        'text/css',
        'text/javascript',
        'application/javascript',
        'application/json',
    )
    if not any(content_type.startswith(item) for item in compressible_types):
        return response

    data = response.get_data()
    if len(data) < 1024:
        return response

    compressed = gzip.compress(data, compresslevel=5)
    if len(compressed) >= len(data):
        return response

    response.set_data(compressed)
    response.headers['Content-Encoding'] = 'gzip'
    response.headers['Content-Length'] = str(len(compressed))
    vary = response.headers.get('Vary')
    if vary:
        if 'Accept-Encoding' not in vary:
            response.headers['Vary'] = f'{vary}, Accept-Encoding'
    else:
        response.headers['Vary'] = 'Accept-Encoding'
    return response


def configure_sqlite_connection(conn):
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA busy_timeout = 30000')
    conn.execute('PRAGMA journal_mode = WAL')
    conn.execute('PRAGMA synchronous = NORMAL')
    return conn

def ensure_database_backup_dir():
    os.makedirs(DATABASE_BACKUP_DIR, exist_ok=True)

def sqlite_file_backup(source_path, destination_path):
    os.makedirs(os.path.dirname(destination_path), exist_ok=True)
    source_conn = sqlite3.connect(source_path)
    backup_conn = sqlite3.connect(destination_path)
    try:
        source_conn.backup(backup_conn)
    finally:
        backup_conn.close()
        source_conn.close()

def create_all_database_backup():
    ensure_database_backup_dir()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_folder = os.path.join(DATABASE_BACKUP_DIR, timestamp)
    os.makedirs(backup_folder, exist_ok=False)

    manifest = {
        'backup_id': timestamp,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'created_by': session.get('user_email', ''),
        'projects': {},
    }

    for project_key, info in PROJECT_DATABASES.items():
        source_path = info['path']
        backup_name = f"{project_key}_{info['filename']}"
        backup_path = os.path.join(backup_folder, backup_name)
        if project_key == 'main_app' and main_db.use_postgres(DATABASE):
            backup_name = 'main_app_POSTGRES_DATABASE.txt'
            backup_path = os.path.join(backup_folder, backup_name)
            with open(backup_path, 'w', encoding='utf-8') as f:
                f.write(
                    'Main app is configured to use PostgreSQL via APP_DATABASE_URL. '
                    'Use pg_dump or Z_backup_postgres_asijaapp.bat to back up the live main app database.\n'
                )
            status = 'postgres_external_backup_required'
        elif os.path.exists(source_path):
            sqlite_file_backup(source_path, backup_path)
            status = 'backed_up'
        else:
            status = 'missing'
        manifest['projects'][project_key] = {
            'label': info['label'],
            'source_path': source_path,
            'backup_file': backup_name,
            'status': status,
        }

    manifest_path = os.path.join(backup_folder, 'manifest.json')
    with open(manifest_path, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2)

    return manifest

def list_database_backups():
    ensure_database_backup_dir()
    backups = []
    for name in sorted(os.listdir(DATABASE_BACKUP_DIR), reverse=True):
        folder = os.path.join(DATABASE_BACKUP_DIR, name)
        manifest_path = os.path.join(folder, 'manifest.json')
        if not os.path.isdir(folder) or not os.path.exists(manifest_path):
            continue
        try:
            with open(manifest_path, 'r', encoding='utf-8') as f:
                manifest = json.load(f)
            backups.append(manifest)
        except (OSError, json.JSONDecodeError):
            continue
    return backups

def safe_restore_database(project_key, backup_id):
    if project_key not in PROJECT_DATABASES:
        raise ValueError('Invalid project selected.')
    if project_key == 'main_app' and main_db.use_postgres(DATABASE):
        raise ValueError('Main app is using PostgreSQL. Restore it with psql from a PostgreSQL backup file.')

    backup_folder = os.path.abspath(os.path.join(DATABASE_BACKUP_DIR, backup_id))
    backup_root = os.path.abspath(DATABASE_BACKUP_DIR)
    if not backup_folder.startswith(backup_root + os.sep):
        raise ValueError('Invalid backup selected.')

    manifest_path = os.path.join(backup_folder, 'manifest.json')
    if not os.path.exists(manifest_path):
        raise FileNotFoundError('Backup manifest not found.')

    with open(manifest_path, 'r', encoding='utf-8') as f:
        manifest = json.load(f)

    project_entry = manifest.get('projects', {}).get(project_key)
    if not project_entry or project_entry.get('status') != 'backed_up':
        raise FileNotFoundError('Selected project database is not available in this backup.')

    backup_file = project_entry.get('backup_file')
    backup_path = os.path.abspath(os.path.join(backup_folder, backup_file))
    if not backup_path.startswith(backup_folder + os.sep) or not os.path.exists(backup_path):
        raise FileNotFoundError('Backup database file not found.')

    target_path = PROJECT_DATABASES[project_key]['path']
    safety_id = datetime.now().strftime('%Y%m%d_%H%M%S')
    safety_folder = os.path.join(DATABASE_BACKUP_DIR, f'before_restore_{project_key}_{safety_id}')
    os.makedirs(safety_folder, exist_ok=False)
    safety_path = os.path.join(safety_folder, os.path.basename(target_path))
    if os.path.exists(target_path):
        sqlite_file_backup(target_path, safety_path)

    os.makedirs(os.path.dirname(target_path), exist_ok=True)
    temp_restore_path = f"{target_path}.restore_tmp"
    shutil.copy2(backup_path, temp_restore_path)
    os.replace(temp_restore_path, target_path)

    for suffix in ('-wal', '-shm'):
        sidecar_path = target_path + suffix
        try:
            if os.path.exists(sidecar_path):
                os.remove(sidecar_path)
        except OSError:
            pass

    safety_manifest = {
        'backup_id': os.path.basename(safety_folder),
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'created_by': session.get('user_email', ''),
        'restore_safety_for': project_key,
        'projects': {
            project_key: {
                'label': PROJECT_DATABASES[project_key]['label'],
                'source_path': target_path,
                'backup_file': os.path.basename(safety_path),
                'status': 'backed_up' if os.path.exists(safety_path) else 'missing',
            }
        },
    }
    with open(os.path.join(safety_folder, 'manifest.json'), 'w', encoding='utf-8') as f:
        json.dump(safety_manifest, f, indent=2)

    return {
        'project': PROJECT_DATABASES[project_key]['label'],
        'backup_id': backup_id,
        'safety_backup_id': os.path.basename(safety_folder),
    }

def init_db():
    """Initialize the database with required tables"""
    conn = configure_sqlite_connection(main_db.connect(DATABASE, timeout=30))
    c = conn.cursor()
    
    # 1. Create users table
    c.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL
    )
    ''')
    
    # 2. Create assignee_master FIRST (referenced by others)
    c.execute('''
    CREATE TABLE IF NOT EXISTS assignee_master (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        email TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')

    # Existing databases created before assignee emails were added need this
    # column before task-loading joins can select am.email.
    try:
        c.execute("PRAGMA table_info(assignee_master)")
        assignee_columns = [col[1] for col in c.fetchall()]
        if 'email' not in assignee_columns:
            c.execute('ALTER TABLE assignee_master ADD COLUMN email TEXT')
        conn.commit()
    except Exception as e:
        print(f"Migration note (assignee_master): {e}")
        conn.rollback()
    
    # 3. Create work_master
    c.execute('''
    CREATE TABLE IF NOT EXISTS work_master (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        work_name TEXT NOT NULL,
        day_of_month INTEGER,
        work_date TEXT,
        work_start_date TEXT,
        work_tat TEXT NOT NULL,
        priority TEXT DEFAULT 'Normal',
        assignee_id INTEGER,
        assigned_user_id INTEGER,
        deleted_at TIMESTAMP,
        deleted_by_user_id INTEGER,
        delete_request_id INTEGER,
        deleted_hidden_at TIMESTAMP,
        deleted_hidden_by_user_id INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(assignee_id) REFERENCES assignee_master(id),
        FOREIGN KEY(assigned_user_id) REFERENCES users(id)
    )
    ''')

    # Create default admin user if none exists
    c.execute("SELECT count(*) FROM users")
    if c.fetchone()[0] == 0:
        hashed_pw = generate_password_hash('admin123')
        c.execute("INSERT INTO users (email, password) VALUES (?, ?)", ('admin@example.com', hashed_pw))

    # Hardened Migration logic for work_master
    try:
        c.execute("PRAGMA table_info(work_master)")
        columns = [col[1] for col in c.fetchall()]
        
        # Handle legacy work_due_date rename if still present
        if 'work_due_date' in columns:
            c.execute('DROP TABLE IF EXISTS work_master_old')
            c.execute('ALTER TABLE work_master RENAME TO work_master_old')
            c.execute('''
            CREATE TABLE work_master (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                work_name TEXT NOT NULL,
                day_of_month INTEGER,
                work_date TEXT,
                work_start_date TEXT,
                work_tat TEXT NOT NULL,
                priority TEXT DEFAULT 'Normal', 
                assignee_id INTEGER,
                assigned_user_id INTEGER,
                deleted_at TIMESTAMP,
                deleted_by_user_id INTEGER,
                delete_request_id INTEGER,
                deleted_hidden_at TIMESTAMP,
                deleted_hidden_by_user_id INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(assignee_id) REFERENCES assignee_master(id),
                FOREIGN KEY(assigned_user_id) REFERENCES users(id)
            )
            ''')
            c.execute('''
            INSERT INTO work_master (id, work_name, work_date, work_tat, created_at)
            SELECT id, work_name, work_due_date, work_tat, CURRENT_TIMESTAMP FROM work_master_old
            ''')
            c.execute('DROP TABLE work_master_old')
            # Refresh column list after table recreation
            c.execute("PRAGMA table_info(work_master)")
            columns = [col[1] for col in c.fetchall()]

        # Ensure all modern columns exist (Incremental Updates)
        new_cols = {
            'day_of_month': 'INTEGER',
            'work_date': 'TEXT',
            'work_start_date': 'TEXT',
            'priority': "TEXT DEFAULT 'Normal'",
            'assignee_id': 'INTEGER',
            'assigned_user_id': 'INTEGER',
            'deleted_at': 'TIMESTAMP',
            'deleted_by_user_id': 'INTEGER',
            'delete_request_id': 'INTEGER',
            'deleted_hidden_at': 'TIMESTAMP',
            'deleted_hidden_by_user_id': 'INTEGER',
            'created_at': 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP'
        }
        for col, definition in new_cols.items():
            if col not in columns:
                c.execute(f'ALTER TABLE work_master ADD COLUMN {col} {definition}')
        
        conn.commit()
    except Exception as e:
        print(f"Migration note: {e}")
        conn.rollback()
    
    # 4. Create work_assigned table for today's work
    c.execute('''
    CREATE TABLE IF NOT EXISTS work_assigned (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        work_id INTEGER NOT NULL,
        assigned_date TEXT NOT NULL,
        status TEXT DEFAULT 'pending',
        actual_start TEXT,
        actual_end TEXT,
        remarks TEXT,
        assignee_id INTEGER,
        FOREIGN KEY(work_id) REFERENCES work_master(id),
        FOREIGN KEY(assignee_id) REFERENCES assignee_master(id)
    )
    ''')
    
    # Migrate work_assigned if needed
    try:
        c.execute("PRAGMA table_info(work_assigned)")
        columns_assigned = [col[1] for col in c.fetchall()]
        if 'actual_start' not in columns_assigned:
            c.execute('ALTER TABLE work_assigned ADD COLUMN actual_start TEXT')
        if 'actual_end' not in columns_assigned:
            c.execute('ALTER TABLE work_assigned ADD COLUMN actual_end TEXT')
        if 'remarks' not in columns_assigned:
            c.execute('ALTER TABLE work_assigned ADD COLUMN remarks TEXT')
        if 'assignee_id' not in columns_assigned:
            c.execute('ALTER TABLE work_assigned ADD COLUMN assignee_id INTEGER')
        conn.commit()
    except Exception as e:
        print(f"Migration note (assigned): {e}")
    
    # Create user_roles table for mapping users to roles/departments
    c.execute('''
    CREATE TABLE IF NOT EXISTS user_roles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        assignee_id INTEGER NOT NULL,
        access_level TEXT DEFAULT 'edit',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(user_id) REFERENCES users(id),
        FOREIGN KEY(assignee_id) REFERENCES assignee_master(id),
        UNIQUE(user_id, assignee_id)
    )
    ''')

    try:
        c.execute("PRAGMA table_info(user_roles)")
        role_columns = [col[1] for col in c.fetchall()]
        if 'access_level' not in role_columns:
            c.execute("ALTER TABLE user_roles ADD COLUMN access_level TEXT DEFAULT 'edit'")
            c.execute("UPDATE user_roles SET access_level = 'edit' WHERE access_level IS NULL")
        conn.commit()
    except Exception as e:
        print(f"Migration note (user_roles): {e}")
        conn.rollback()

    c.execute('''
    CREATE TABLE IF NOT EXISTS user_permissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        permission_key TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(user_id) REFERENCES users(id),
        UNIQUE(user_id, permission_key)
    )
    ''')

    c.execute('SELECT id, email FROM users')
    for user_id, email in c.fetchall():
        c.execute('SELECT COUNT(*) FROM user_permissions WHERE user_id = ?', (user_id,))
        if c.fetchone()[0] == 0:
            default_permissions = [p['key'] for p in PAGE_PERMISSIONS] + [p['key'] for p in SPECIAL_PERMISSIONS]
            if str(email).lower() not in ADMIN_EMAILS:
                default_permissions = LEGACY_USER_PERMISSIONS
            for permission_key in default_permissions:
                c.execute(
                    'INSERT OR IGNORE INTO user_permissions (user_id, permission_key) VALUES (?, ?)',
                    (user_id, permission_key)
                )
        c.execute(
            'INSERT OR IGNORE INTO user_permissions (user_id, permission_key) VALUES (?, ?)',
            (user_id, 'daily_debtor_report')
        )
        if str(email).lower() == 'arif.siddiqui@asija.in':
            c.execute(
                'INSERT OR IGNORE INTO user_permissions (user_id, permission_key) VALUES (?, ?)',
                (user_id, 'links_and_others')
            )
        else:
            c.execute(
                'DELETE FROM user_permissions WHERE user_id = ? AND permission_key = ?',
                (user_id, 'links_and_others')
            )

    # Store deleted/skipped occurrences for recurring tasks without deleting the full series
    c.execute('''
    CREATE TABLE IF NOT EXISTS work_skipped_dates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        work_id INTEGER NOT NULL,
        skipped_date TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(work_id) REFERENCES work_master(id),
        UNIQUE(work_id, skipped_date)
    )
    ''')

    c.execute('''
    CREATE TABLE IF NOT EXISTS work_change_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        requester_user_id INTEGER NOT NULL,
        action TEXT NOT NULL,
        work_id INTEGER,
        payload TEXT,
        status TEXT DEFAULT 'pending',
        reviewer_user_id INTEGER,
        review_note TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        reviewed_at TIMESTAMP,
        FOREIGN KEY(requester_user_id) REFERENCES users(id),
        FOREIGN KEY(reviewer_user_id) REFERENCES users(id),
        FOREIGN KEY(work_id) REFERENCES work_master(id)
    )
    ''')
    
    conn.commit()
    conn.close()

def get_db_connection():
    """Get database connection"""
    return configure_sqlite_connection(main_db.connect(DATABASE, timeout=30))

def get_links_db_connection():
    """Get database connection for the links-and-others module."""
    return configure_sqlite_connection(sqlite3.connect(LINKS_DATABASE, timeout=30))

def init_links_db():
    """Ensure the embedded links-and-others module has the tables it needs."""
    os.makedirs(LINKS_DIR, exist_ok=True)
    conn = configure_sqlite_connection(sqlite3.connect(LINKS_DATABASE, timeout=30))
    c = conn.cursor()
    c.execute('''
    CREATE TABLE IF NOT EXISTS dailywork (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        inflowdate TEXT,
        work_name TEXT,
        target_date TEXT,
        status TEXT,
        allocated_to TEXT,
        total_seconds INTEGER
    )
    ''')
    c.execute('''
    CREATE TABLE IF NOT EXISTS important_links (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category TEXT,
        link_name TEXT,
        url TEXT,
        is_special INTEGER
    )
    ''')
    c.execute('''
    CREATE TABLE IF NOT EXISTS daily_diary (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entry_date TEXT,
        particulars TEXT,
        start_t TEXT,
        end_t TEXT,
        spend_time TEXT,
        remark TEXT
    )
    ''')
    conn.commit()
    conn.close()

# ==================== Security Decorator ====================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_email = str(session.get('user_email', '')).lower()
        if not user_email or user_email not in ADMIN_EMAILS:
            # Redirect to index if not the admin
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def arif_only_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_email = str(session.get('user_email', '')).lower()
        if user_email != 'arif.siddiqui@asija.in':
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def is_arif_user():
    return str(session.get('user_email', '')).lower() == 'arif.siddiqui@asija.in'

def is_admin_user():
    return str(session.get('user_email', '')).lower() in ADMIN_EMAILS

def has_user_permission(permission_key):
    if is_admin_user():
        return True

    user_id = session.get('user_id')
    if not user_id:
        return False

    conn = get_db_connection()
    try:
        row = conn.execute(
            'SELECT 1 FROM user_permissions WHERE user_id = ? AND permission_key = ?',
            (user_id, permission_key)
        ).fetchone()
        return row is not None
    finally:
        conn.close()

def permission_required(permission_key):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not has_user_permission(permission_key):
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_current_user_permissions():
    if is_admin_user():
        return {p['key'] for p in PAGE_PERMISSIONS + SPECIAL_PERMISSIONS}

    user_id = session.get('user_id')
    if not user_id:
        return set()

    conn = get_db_connection()
    try:
        rows = conn.execute(
            'SELECT permission_key FROM user_permissions WHERE user_id = ?',
            (user_id,)
        ).fetchall()
        return {r['permission_key'] for r in rows}
    finally:
        conn.close()

def get_navigation_context():
    permissions = get_current_user_permissions()
    return {
        'user_permissions': permissions,
        'can_approve_requests': has_user_permission('approve_requests')
    }

def links_url_for(endpoint, **values):
    if endpoint == 'static':
        return url_for('links_and_others_static', filename=values.get('filename', ''))
    return url_for(endpoint, **values)

def rewrite_links_content(content):
    replacements = {
        '="/static/': '="/links-and-others/static/',
        "='/static/": "='/links-and-others/static/",
        '"/static/': '"/links-and-others/static/',
        "'/static/": "'/links-and-others/static/",
        'action="/add_link"': 'action="/links-and-others/add_link"',
        'action="/edit_link"': 'action="/links-and-others/edit_link"',
        "window.location.href='/backup_db'": "window.location.href='/links-and-others/backup_db'",
        "location.href='/'": "location.href='/links-and-others'",
        "window.open('/diary'": "window.open('/links-and-others/diary'",
        "fetch('/get_diary'": "fetch('/links-and-others/get_diary'",
        "fetch('/save_diary'": "fetch('/links-and-others/save_diary'",
        'http://127.0.0.1:${PORT}/': '/links-and-others/',
        'http://127.0.0.1:${PORT_SIDEBAR}/': '/links-and-others/',
    }
    for old, new in replacements.items():
        content = content.replace(old, new)
    return content

def render_links_template(template_name, **context):
    context['url_for'] = links_url_for
    return rewrite_links_content(render_template(template_name, **context))

def safe_restore_label(label):
    label = (label or 'manual checkpoint').strip()
    label = re.sub(r'[^A-Za-z0-9._ -]+', '', label)
    label = re.sub(r'\s+', '_', label).strip('._-')
    return label[:60] or 'manual_checkpoint'

def should_skip_restore_path(rel_path):
    parts = rel_path.replace('\\', '/').split('/')
    if not parts:
        return True
    skip_names = {
        'venv',
        '__pycache__',
        '.git',
        '.idea',
        '.vscode',
        'restore_points',
        'backups'
    }
    if any(part in skip_names for part in parts):
        return True
    filename = parts[-1].lower()
    return filename.endswith(('.pyc', '.pyo', '.rar', '.zip'))

def create_restore_point(label, reason='manual'):
    os.makedirs(RESTORE_POINTS_DIR, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_label = safe_restore_label(label)
    archive_name = f'{timestamp}_{safe_label}.zip'
    archive_path = os.path.join(RESTORE_POINTS_DIR, archive_name)
    created_by = session.get('user_email', 'system')
    files = []

    for root, dirs, filenames in os.walk(BASE_DIR):
        rel_root = os.path.relpath(root, BASE_DIR)
        if rel_root == '.':
            rel_root = ''
        dirs[:] = [
            d for d in dirs
            if not should_skip_restore_path(os.path.join(rel_root, d))
        ]
        for filename in filenames:
            file_path = os.path.join(root, filename)
            rel_path = os.path.relpath(file_path, BASE_DIR)
            if should_skip_restore_path(rel_path):
                continue
            files.append((file_path, rel_path))

    manifest = {
        'label': label or 'Manual checkpoint',
        'reason': reason,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'created_by': created_by,
        'file_count': len(files)
    }

    with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('restore_manifest.json', json.dumps(manifest, indent=2))
        for file_path, rel_path in files:
            zf.write(file_path, rel_path)

    manifest['id'] = archive_name
    manifest['size'] = os.path.getsize(archive_path)
    return manifest

def read_restore_point_manifest(filename):
    archive_path = os.path.join(RESTORE_POINTS_DIR, filename)
    manifest = {
        'id': filename,
        'label': filename,
        'created_at': '',
        'created_by': '',
        'reason': '',
        'file_count': 0,
        'size': os.path.getsize(archive_path)
    }
    try:
        with zipfile.ZipFile(archive_path, 'r') as zf:
            data = json.loads(zf.read('restore_manifest.json').decode('utf-8'))
            manifest.update(data)
    except Exception:
        pass
    return manifest

def list_restore_points():
    os.makedirs(RESTORE_POINTS_DIR, exist_ok=True)
    points = []
    for filename in os.listdir(RESTORE_POINTS_DIR):
        if filename.lower().endswith('.zip'):
            points.append(read_restore_point_manifest(filename))
    return sorted(points, key=lambda item: item['id'], reverse=True)

def clean_project_for_restore():
    for name in os.listdir(BASE_DIR):
        rel_path = name
        if should_skip_restore_path(rel_path):
            continue
        target = os.path.join(BASE_DIR, name)
        if os.path.isdir(target):
            shutil.rmtree(target)
        else:
            os.remove(target)

def restore_project_from_point(filename):
    archive_path = os.path.abspath(os.path.join(RESTORE_POINTS_DIR, filename))
    restore_dir_abs = os.path.abspath(RESTORE_POINTS_DIR)
    if not archive_path.startswith(restore_dir_abs + os.sep) or not os.path.isfile(archive_path):
        raise ValueError('Restore point not found')

    safety_point = create_restore_point(f'Before restoring {filename}', reason='before_restore')
    clean_project_for_restore()

    base_abs = os.path.abspath(BASE_DIR)
    with zipfile.ZipFile(archive_path, 'r') as zf:
        for member in zf.infolist():
            rel_path = member.filename.replace('\\', '/')
            if rel_path == 'restore_manifest.json' or should_skip_restore_path(rel_path):
                continue
            target_path = os.path.abspath(os.path.join(BASE_DIR, rel_path))
            if not target_path.startswith(base_abs + os.sep):
                raise ValueError('Unsafe path in restore point')
            if member.is_dir():
                os.makedirs(target_path, exist_ok=True)
                continue
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            with zf.open(member, 'r') as src, open(target_path, 'wb') as dst:
                shutil.copyfileobj(src, dst)

    return safety_point

def get_pending_approval_button_context():
    """Return sidebar badge text/count for approval requests."""
    conn = get_db_connection()
    try:
        if has_user_permission('approve_requests'):
            count = conn.execute(
                "SELECT COUNT(*) FROM work_change_requests WHERE status = 'pending'"
            ).fetchone()[0]
            return {
                'approval_button_label': 'Pending Approvals',
                'approval_button_count': count
            }

        count = conn.execute(
            """
            SELECT COUNT(*)
            FROM work_change_requests
            WHERE requester_user_id = ? AND status = 'pending'
            """,
            (session.get('user_id'),)
        ).fetchone()[0]
        return {
            'approval_button_label': 'Raised Requests',
            'approval_button_count': count
        }
    finally:
        conn.close()

def build_active_work_where(extra_condition='', extra_params=None):
    clauses = ['wm.deleted_at IS NULL']
    params = []
    if extra_condition:
        clauses.append(f'({extra_condition})')
        params.extend(extra_params or [])
    return 'WHERE ' + ' AND '.join(clauses), params

def get_current_user_role_scope(conn):
    """Return whether the current user can see all tasks and their allowed assignees."""
    user_id = session.get('user_id')
    user_email = str(session.get('user_email', '')).lower()
    is_admin = bool(user_email and user_email in ADMIN_EMAILS)

    if is_admin:
        return True, []

    role_rows = conn.execute(
        "SELECT assignee_id FROM user_roles WHERE user_id = ? AND access_level IN ('view', 'edit')",
        (user_id,)
    ).fetchall()
    return False, [r['assignee_id'] for r in role_rows]

def can_manage_users_or_departments():
    return (
        is_admin_user()
        or has_user_permission('user_management')
        or has_user_permission('user_power_management')
        or has_user_permission('assign_to_master')
        or has_user_permission('approve_requests')
    )

def get_current_user_editable_assignee_ids(conn):
    if is_admin_user() or has_user_permission('approve_requests'):
        return None

    rows = conn.execute(
        "SELECT assignee_id FROM user_roles WHERE user_id = ? AND access_level = 'edit'",
        (session.get('user_id'),)
    ).fetchall()
    return {row['assignee_id'] for row in rows}

def can_current_user_edit_assignee(conn, assignee_id):
    editable_ids = get_current_user_editable_assignee_ids(conn)
    if editable_ids is None:
        return True
    if assignee_id in (None, ''):
        return False
    return int(assignee_id) in editable_ids

def ensure_current_user_can_edit_work(conn, work_id, payload=None):
    if is_admin_user() or has_user_permission('approve_requests'):
        return

    work = conn.execute(
        'SELECT assignee_id FROM work_master WHERE id = ? AND deleted_at IS NULL',
        (work_id,)
    ).fetchone()
    if not work:
        raise ValueError('Task not found or already deleted')

    if not can_current_user_edit_assignee(conn, work['assignee_id']):
        raise PermissionError('You only have view access for this department.')

    if payload and not can_current_user_edit_assignee(conn, payload.get('assignee_id')):
        raise PermissionError('You cannot move this task to a department where you do not have edit access.')

def ensure_current_user_can_create_work(conn, payload):
    if is_admin_user() or has_user_permission('approve_requests'):
        return
    if not can_current_user_edit_assignee(conn, payload.get('assignee_id')):
        raise PermissionError('You only have view access for this department.')

def save_user_department_access(conn, user_id, department_access):
    conn.execute('DELETE FROM user_roles WHERE user_id = ?', (user_id,))
    for item in department_access or []:
        assignee_id = item.get('assignee_id')
        access_level = item.get('access_level', 'edit')
        if assignee_id and access_level in ('view', 'edit'):
            conn.execute(
                'INSERT OR IGNORE INTO user_roles (user_id, assignee_id, access_level) VALUES (?, ?, ?)',
                (user_id, assignee_id, access_level)
            )

def is_date_scheduled(start_dt, target_dt, frequency):
    """
    Determines if a task starting on start_dt is scheduled to occur on target_dt
    based on its frequency.
    """
    if target_dt < start_dt:
        return False

    # Detect month-end alignment for monthly/quarterly/yearly tasks
    _, last_day_start = calendar.monthrange(start_dt.year, start_dt.month)
    was_month_end = (start_dt.day == last_day_start)
    _, last_day_target = calendar.monthrange(target_dt.year, target_dt.month)
    is_now_month_end = (target_dt.day == last_day_target)

    is_day_match = is_now_month_end if was_month_end else (target_dt.day == start_dt.day)

    if frequency == 'once':
        return target_dt == start_dt
    if frequency in ['daily', 'ongoing']:
        return True
    if frequency == 'weekly':
        return (target_dt - start_dt).days % 7 == 0
    if frequency == 'monthly':
        return is_day_match
    if frequency == 'quarterly':
        months_diff = (target_dt.year - start_dt.year) * 12 + (target_dt.month - start_dt.month)
        return is_day_match and (months_diff % 3 == 0)
    if frequency == 'yearly':
        return is_day_match and (target_dt.month == start_dt.month)
    
    return False

# Initialize database on startup
init_db()
init_links_db()

# ==================== Routes ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['user_email'] = user['email']
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Invalid email or password')
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Clear session and redirect to login"""
    session.clear()
    return redirect(url_for('login'))

@app.route('/user-management')
@login_required
@permission_required('user_management')
def user_management():
    """Web-based User Management page"""
    return render_template('user_management.html')

@app.route('/user-role-management')
@login_required
@permission_required('user_power_management')
def user_role_management():
    return render_template(
        'user_role_management.html',
        page_permissions=PAGE_PERMISSIONS,
        special_permissions=SPECIAL_PERMISSIONS
    )

@app.route('/')
@login_required
def index():
    """Welcome page with two buttons"""
    context = get_navigation_context()
    context.update(get_pending_approval_button_context())
    return render_template('index.html', **context)

@app.route('/work-master')
@login_required
@permission_required('work_master')
def work_master():
    """Work Master page"""
    context = get_navigation_context()
    context.update(get_pending_approval_button_context())
    return render_template(
        'work_master.html',
        project_period_start=PROJECT_PERIOD_START.strftime('%Y-%m-%d'),
        project_period_end=PROJECT_PERIOD_END.strftime('%Y-%m-%d'),
        **context
    )

@app.route('/today-work')
@login_required
@permission_required('today_work')
def today_work():
    """Today's Work page with calendar"""
    context = get_navigation_context()
    context.update(get_pending_approval_button_context())
    return render_template(
        'today_work.html',
        project_period_start=PROJECT_PERIOD_START.strftime('%Y-%m-%d'),
        project_period_end=PROJECT_PERIOD_END.strftime('%Y-%m-%d'),
        **context
    )

@app.route('/report-section')
@login_required
@permission_required('report_section')
def report_section():
    """Report Section page"""
    return render_template('report_section.html')

@app.route('/control-panel')
@login_required
@permission_required('control_panel')
def control_panel():
    """Control Panel page"""
    return render_template(
        'control_panel.html',
        report_update_dates=get_debtor_report_update_dates(),
        **get_navigation_context()
    )

@app.route('/control-panel/report-update', methods=['POST'])
@login_required
@permission_required('control_panel')
def update_debtor_report_dates():
    """Manually update debtor report navbar date labels."""
    report_as_on_date = request.form.get('report_as_on_date', '').strip()
    last_bill_update_date = request.form.get('last_bill_update_date', '').strip()
    last_receipt_update_date = request.form.get('last_receipt_update_date', '').strip()

    try:
        datetime.strptime(report_as_on_date, '%Y-%m-%d')
        datetime.strptime(last_bill_update_date, '%Y-%m-%d')
        datetime.strptime(last_receipt_update_date, '%Y-%m-%d')
    except ValueError:
        return redirect(url_for('control_panel'))

    set_debtor_report_update_dates(report_as_on_date, last_bill_update_date, last_receipt_update_date)
    return redirect(url_for('control_panel'))

@app.route('/pending-approvals')
@login_required
@permission_required('pending_approvals')
def pending_approvals():
    """Approval notification page for Work Master change requests"""
    return render_template('pending_approvals.html')

@app.route('/request-logs')
@login_required
@permission_required('request_logs')
def request_logs():
    """Task Master request log page."""
    return render_template('request_logs.html')

@app.route('/deleted-records')
@login_required
@permission_required('deleted_records')
def deleted_records():
    """Arif-only page for soft-deleted Task Master records."""
    return render_template('deleted_records.html')

@app.route('/restore-points')
@login_required
@permission_required('restore_points')
def restore_points():
    """Project restore point management page."""
    return render_template('restore_points.html')

@app.route('/api/users', methods=['GET'])
@login_required
def get_users_list():
    """Get list of users for linking with assignees"""
    conn = get_db_connection()
    users = conn.execute('SELECT id, email FROM users ORDER BY email ASC').fetchall()
    roles = conn.execute('SELECT user_id, assignee_id, access_level FROM user_roles').fetchall()
    conn.close()

    roles_dict = {}
    for role in roles:
        roles_dict.setdefault(role['user_id'], []).append({
            'assignee_id': role['assignee_id'],
            'access_level': role['access_level'] or 'edit'
        })

    return jsonify([{
        'id': user['id'],
        'email': user['email'],
        'department_access': roles_dict.get(user['id'], [])
    } for user in users])

@app.route('/api/users', methods=['POST'])
@login_required
@permission_required('user_management')
def create_user_account():
    data = request.json or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    department_access = data.get('department_access', [])

    if not email or not password:
        return jsonify({'success': False, 'error': 'Email and password are required'}), 400

    conn = get_db_connection()
    try:
        cursor = conn.execute(
            'INSERT INTO users (email, password) VALUES (?, ?)',
            (email, generate_password_hash(password))
        )
        user_id = cursor.lastrowid
        save_user_department_access(conn, user_id, department_access)
        conn.commit()
        return jsonify({'success': True, 'id': user_id}), 201
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': 'User email already exists'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
@permission_required('user_management')
def update_user_account(user_id):
    data = request.json or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    department_access = data.get('department_access')

    if not email:
        return jsonify({'success': False, 'error': 'Email is required'}), 400

    conn = get_db_connection()
    try:
        existing = conn.execute('SELECT email FROM users WHERE id = ?', (user_id,)).fetchone()
        if not existing:
            return jsonify({'success': False, 'error': 'User not found'}), 404

        if password:
            conn.execute(
                'UPDATE users SET email = ?, password = ? WHERE id = ?',
                (email, generate_password_hash(password), user_id)
            )
        else:
            conn.execute('UPDATE users SET email = ? WHERE id = ?', (email, user_id))
        if department_access is not None:
            save_user_department_access(conn, user_id, department_access)
        conn.commit()

        if session.get('user_id') == user_id:
            session['user_email'] = email

        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': 'User email already exists'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@permission_required('user_management')
def delete_user_account(user_id):
    if session.get('user_id') == user_id:
        return jsonify({'success': False, 'error': 'You cannot delete your own login'}), 400

    conn = get_db_connection()
    try:
        user = conn.execute('SELECT email FROM users WHERE id = ?', (user_id,)).fetchone()
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        if user['email'].lower() == 'arif.siddiqui@asija.in':
            return jsonify({'success': False, 'error': 'Arif user cannot be deleted'}), 400

        conn.execute('UPDATE work_master SET assigned_user_id = NULL WHERE assigned_user_id = ?', (user_id,))
        conn.execute('DELETE FROM user_roles WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM user_permissions WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/assign-to-master')
@login_required
@permission_required('assign_to_master')
def assign_to_master():
    """Assign to Master page"""
    return render_template('master.html')

@app.route('/links-and-others/static/<path:filename>')
@login_required
@arif_only_required
def links_and_others_static(filename):
    """Serve static files for the embedded links-and-others module."""
    static_dir = os.path.join(LINKS_DIR, 'static')
    file_path = os.path.join(static_dir, filename)
    if filename.endswith(('.js', '.css', '.html')) and os.path.isfile(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = rewrite_links_content(f.read())
        mimetype = 'application/javascript' if filename.endswith('.js') else None
        if filename.endswith('.css'):
            mimetype = 'text/css'
        if filename.endswith('.html'):
            mimetype = 'text/html'
        return Response(content, mimetype=mimetype)
    return send_from_directory(static_dir, filename)

@app.route('/links-and-others')
@login_required
@arif_only_required
def links_and_others():
    """Embedded MAS links/tasks dashboard from linksandothers."""
    try:
        conn = get_links_db_connection()
        status_query = request.args.get('status', 'All')
        user_query = request.args.get('user', 'All')

        db_links = conn.execute('SELECT * FROM important_links ORDER BY category').fetchall()
        categories_list = [
            row['category']
            for row in conn.execute('SELECT DISTINCT category FROM important_links ORDER BY category').fetchall()
        ]

        query = 'SELECT * FROM dailywork WHERE 1=1'
        params = []
        if status_query != 'All':
            query += ' AND status = ?'
            params.append(status_query)
        if user_query != 'All':
            query += ' AND allocated_to = ?'
            params.append(user_query)

        query += ' ORDER BY id DESC'
        tasks = conn.execute(query, params).fetchall()
        conn.close()
        return render_links_template(
            'linksandothers/templates/index.html',
            sidebar_links=db_links,
            tasks=tasks,
            categories=categories_list
        )
    except Exception as e:
        print(f"Links and others error: {e}")
        return render_links_template('linksandothers/templates/index.html', sidebar_links=[], tasks=[], categories=[])

@app.route('/links-and-others/diary')
@login_required
@arif_only_required
def links_and_others_diary():
    return render_links_template('linksandothers/templates/diary.html')

@app.route('/links-and-others/get_tasks', methods=['GET'])
@login_required
@arif_only_required
def links_get_tasks():
    conn = get_links_db_connection()
    tasks = conn.execute(
        'SELECT inflowdate, work_name, target_date, status, allocated_to, total_seconds FROM dailywork ORDER BY id DESC'
    ).fetchall()
    conn.close()
    return jsonify([dict(row) for row in tasks])

@app.route('/links-and-others/save_tasks', methods=['POST'])
@login_required
@arif_only_required
def links_save_tasks():
    data = request.json or []
    conn = get_links_db_connection()
    try:
        conn.execute('DELETE FROM dailywork')
        for item in data:
            conn.execute(
                'INSERT INTO dailywork (inflowdate, work_name, target_date, status, allocated_to, total_seconds) VALUES (?, ?, ?, ?, ?, ?)',
                (item['eDate'], item['wName'], item['tDate'], item['status'], item.get('allocatedTo', ''), item.get('totalSeconds', 0))
            )
        conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/links-and-others/get_diary', methods=['GET'])
@login_required
@arif_only_required
def links_get_diary():
    conn = get_links_db_connection()
    rows = conn.execute('SELECT * FROM daily_diary ORDER BY id ASC').fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])

@app.route('/links-and-others/save_diary', methods=['POST'])
@login_required
@arif_only_required
def links_save_diary():
    data = request.json or []
    conn = get_links_db_connection()
    try:
        conn.execute('DELETE FROM daily_diary')
        for item in data:
            conn.execute(
                'INSERT INTO daily_diary (entry_date, particulars, start_t, end_t, spend_time, remark) VALUES (?, ?, ?, ?, ?, ?)',
                (item['date'], item['particulars'], item['start_t'], item['end_t'], item['spend_time'], item['remark'])
            )
        conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/links-and-others/backup_db', methods=['GET'])
@login_required
@arif_only_required
def links_backup_db():
    return send_file(LINKS_DATABASE, as_attachment=True, download_name='mywork.db')

@app.route('/links-and-others/add_link', methods=['POST'])
@login_required
@arif_only_required
def links_add_link():
    data = request.form
    conn = get_links_db_connection()
    conn.execute(
        'INSERT INTO important_links (category, link_name, url, is_special) VALUES (?, ?, ?, ?)',
        (data.get('category'), data.get('link_name'), data.get('url'), 1 if data.get('is_special') else 0)
    )
    conn.commit()
    conn.close()
    return redirect(url_for('links_and_others'))

@app.route('/links-and-others/edit_link', methods=['POST'])
@login_required
@arif_only_required
def links_edit_link():
    data = request.form
    conn = get_links_db_connection()
    conn.execute(
        'UPDATE important_links SET category = ?, link_name = ?, url = ?, is_special = ? WHERE id = ?',
        (data.get('category'), data.get('link_name'), data.get('url'), 1 if data.get('is_special') else 0, data.get('link_id'))
    )
    conn.commit()
    conn.close()
    return redirect(url_for('links_and_others'))

@app.route('/links-and-others/delete_link/<int:link_id>', methods=['DELETE'])
@login_required
@arif_only_required
def links_delete_link(link_id):
    conn = get_links_db_connection()
    try:
        conn.execute('DELETE FROM important_links WHERE id = ?', (link_id,))
        conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/backup')
@login_required
@permission_required('control_panel')
def backup_project():
    """Zips the entire project directory for backup and download"""
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        memory_file = io.BytesIO()
        
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(base_dir):
                # Exclude environment and temporary folders
                dirs[:] = [d for d in dirs if d not in ['venv', '__pycache__', '.git', '.idea', '.vscode']]
                
                for file in files:
                    file_path = os.path.join(root, file)
                    # Create archive name relative to project root
                    arcname = os.path.relpath(file_path, base_dir)
                    zf.write(file_path, arcname)
                    
        memory_file.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'asija_app_backup_{timestamp}.zip'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/database-backups', methods=['GET'])
@login_required
@arif_only_required
def api_database_backups():
    return jsonify({
        'success': True,
        'projects': {
            key: {'label': info['label'], 'filename': info['filename']}
            for key, info in PROJECT_DATABASES.items()
        },
        'backups': list_database_backups(),
    })

@app.route('/api/database-backups', methods=['POST'])
@login_required
@arif_only_required
def api_create_database_backup():
    try:
        manifest = create_all_database_backup()
        return jsonify({
            'success': True,
            'message': 'All project databases backed up successfully.',
            'backup': manifest,
        }), 201
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/database-backups/restore', methods=['POST'])
@login_required
@arif_only_required
def api_restore_database_backup():
    data = request.json or {}
    project_key = data.get('project_key', '')
    backup_id = data.get('backup_id', '')
    confirm = data.get('confirm', '')
    if confirm != 'RESTORE':
        return jsonify({'success': False, 'error': 'Restore confirmation is required.'}), 400
    try:
        result = safe_restore_database(project_key, backup_id)
        return jsonify({
            'success': True,
            'message': f"{result['project']} database restored successfully.",
            'result': result,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/restore-points', methods=['GET'])
@login_required
@permission_required('restore_points')
def api_restore_points():
    return jsonify(list_restore_points())

@app.route('/api/restore-points', methods=['POST'])
@login_required
@permission_required('restore_points')
def api_create_restore_point():
    data = request.json or {}
    try:
        point = create_restore_point(data.get('label', 'Manual checkpoint'), reason='manual')
        return jsonify({'success': True, 'restore_point': point}), 201
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/restore-points/<path:filename>/download')
@login_required
@permission_required('restore_points')
def api_download_restore_point(filename):
    archive_path = os.path.abspath(os.path.join(RESTORE_POINTS_DIR, filename))
    restore_dir_abs = os.path.abspath(RESTORE_POINTS_DIR)
    if not archive_path.startswith(restore_dir_abs + os.sep) or not os.path.isfile(archive_path):
        return jsonify({'success': False, 'error': 'Restore point not found'}), 404
    return send_file(archive_path, as_attachment=True, download_name=filename)

@app.route('/api/restore-points/<path:filename>/restore', methods=['POST'])
@login_required
@permission_required('restore_points')
def api_restore_project(filename):
    data = request.json or {}
    if data.get('confirm') != 'RESTORE':
        return jsonify({'success': False, 'error': 'Confirmation is required'}), 400
    try:
        safety_point = restore_project_from_point(filename)
        return jsonify({
            'success': True,
            'message': 'Project restored. Restart the Flask app to load restored backend code.',
            'safety_restore_point': safety_point
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== API Endpoints ====================

# Work Master API
@app.route('/api/works', methods=['GET'])
@login_required
def get_works():
    """Get all work master records"""
    try:
        conn = get_db_connection()
        is_admin, allowed_assignee_ids = get_current_user_role_scope(conn)
        user_id = session.get('user_id')

        role_condition = ''
        role_params = []
        if not is_admin:
            role_filter = '0'
            if allowed_assignee_ids:
                placeholders = ','.join('?' for _ in allowed_assignee_ids)
                role_filter = f'(wm.assigned_user_id IS NULL AND wm.assignee_id IN ({placeholders}))'
                role_params.extend(allowed_assignee_ids)
            role_condition = f'wm.assigned_user_id = ? OR {role_filter}'
            role_params.insert(0, user_id)

        where_clause, params = build_active_work_where(role_condition, role_params)

        # Join with assignee_master to get assignee name
        works = conn.execute(f'''
            SELECT wm.*, am.name AS assignee_name, am.email AS assignee_email, u.email AS assigned_user_email
            FROM work_master wm
            LEFT JOIN assignee_master am ON wm.assignee_id = am.id
            LEFT JOIN users u ON wm.assigned_user_id = u.id
            {where_clause}
            ORDER BY wm.id DESC
        ''', params).fetchall()
        conn.close()
        # Debug log for terminal
        print(f"API: Found {len(works)} tasks in Work Master (is_admin: {is_admin}).")
        return jsonify([dict(work) for work in works])
    except Exception as e:
        print(f"CRITICAL ERROR in get_works: {str(e)}")
        return jsonify({'error': str(e)}), 500

def create_work_change_request(conn, action, payload=None, work_id=None):
    conn.execute(
        'INSERT INTO work_change_requests (requester_user_id, action, work_id, payload) VALUES (?, ?, ?, ?)',
        (session.get('user_id'), action, work_id, json.dumps(payload or {}))
    )
    conn.commit()
    request_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    return request_id

def get_pending_task_change_request(conn, work_id):
    """Find pending edit/delete request for this task, if any."""
    return conn.execute('''
        SELECT id, action, requester_user_id
        FROM work_change_requests
        WHERE work_id = ?
          AND status = 'pending'
          AND action IN ('edit', 'delete')
        ORDER BY created_at DESC
        LIMIT 1
    ''', (work_id,)).fetchone()

def insert_work_record(conn, data):
    work_name = data.get('work_name', '').strip()
    work_tat = data.get('work_tat', '')
    work_date = data.get('work_date')
    work_start_date = data.get('work_start_date')
    priority = data.get('priority', 'Normal')
    assignee_id = data.get('assignee_id')
    assigned_user_id = data.get('assigned_user_id')

    if not work_name or not work_tat or not work_date:
        raise ValueError('Work Name, Frequency, and Target Date are required')
    target_dt = datetime.strptime(work_date, '%Y-%m-%d').date()
    if target_dt < PROJECT_PERIOD_START or target_dt > PROJECT_PERIOD_END:
        raise ValueError('Target Date must be within 01.04.2026 - 31.03.2027')
    if work_start_date:
        start_dt = datetime.strptime(work_start_date, '%Y-%m-%d').date()
        if start_dt < PROJECT_PERIOD_START or start_dt > PROJECT_PERIOD_END:
            raise ValueError('Work Start Date must be within 01.04.2026 - 31.03.2027')

    conn.execute(
        'INSERT INTO work_master (work_name, day_of_month, work_date, work_start_date, work_tat, priority, assignee_id, assigned_user_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
        (work_name, None, work_date, work_start_date, work_tat, priority, assignee_id, assigned_user_id)
    )
    return conn.execute('SELECT last_insert_rowid()').fetchone()[0]

def update_work_record(conn, work_id, data):
    work_name = data.get('work_name', '').strip()
    work_tat = data.get('work_tat', '')
    work_date = data.get('work_date')
    work_start_date = data.get('work_start_date')
    priority = data.get('priority', 'Normal')
    assignee_id = data.get('assignee_id')
    assigned_user_id = data.get('assigned_user_id')

    if not work_name or not work_tat or not work_date:
        raise ValueError('Work Name, Frequency, and Target Date are required')
    target_dt = datetime.strptime(work_date, '%Y-%m-%d').date()
    if target_dt < PROJECT_PERIOD_START or target_dt > PROJECT_PERIOD_END:
        raise ValueError('Target Date must be within 01.04.2026 - 31.03.2027')
    if work_start_date:
        start_dt = datetime.strptime(work_start_date, '%Y-%m-%d').date()
        if start_dt < PROJECT_PERIOD_START or start_dt > PROJECT_PERIOD_END:
            raise ValueError('Work Start Date must be within 01.04.2026 - 31.03.2027')

    conn.execute(
        'UPDATE work_master SET work_name=?, day_of_month=?, work_date=?, work_start_date=?, work_tat=?, priority=?, assignee_id=?, assigned_user_id=? WHERE id=? AND deleted_at IS NULL',
        (work_name, None, work_date, work_start_date, work_tat, priority, assignee_id, assigned_user_id, work_id)
    )

def delete_work_record(conn, work_id):
    conn.execute('''
        UPDATE work_master
        SET deleted_at = CURRENT_TIMESTAMP,
            deleted_by_user_id = ?,
            delete_request_id = NULL,
            deleted_hidden_at = NULL,
            deleted_hidden_by_user_id = NULL
        WHERE id = ? AND deleted_at IS NULL
    ''', (session.get('user_id'), work_id))

def soft_delete_work_record(conn, work_id, request_id=None):
    conn.execute('''
        UPDATE work_master
        SET deleted_at = CURRENT_TIMESTAMP,
            deleted_by_user_id = ?,
            delete_request_id = ?,
            deleted_hidden_at = NULL,
            deleted_hidden_by_user_id = NULL
        WHERE id = ? AND deleted_at IS NULL
    ''', (session.get('user_id'), request_id, work_id))

@app.route('/api/works', methods=['POST'])
@login_required
def add_work():
    """Add new work to work master"""
    data = request.json
    conn = get_db_connection()
    
    try:
        if not has_user_permission('approve_requests'):
            ensure_current_user_can_create_work(conn, data)
            request_id = create_work_change_request(conn, 'create', data)
            conn.close()
            return jsonify({
                'success': True,
                'pending_approval': True,
                'request_id': request_id,
                'message': 'Task creation sent for approval.'
            }), 202

        last_id = insert_work_record(conn, data)
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'id': last_id}), 201
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/works/<int:work_id>', methods=['PUT'])
@login_required
def update_work(work_id):
    """Update work record"""
    data = request.json
    conn = get_db_connection()
    
    try:
        if not has_user_permission('approve_requests'):
            ensure_current_user_can_edit_work(conn, work_id, data)

            pending_request = get_pending_task_change_request(conn, work_id)
            if pending_request:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': f"This task already has a pending {pending_request['action']} request. Please wait until it is approved or rejected."
                }), 409

            request_id = create_work_change_request(conn, 'edit', data, work_id)
            conn.close()
            return jsonify({
                'success': True,
                'pending_approval': True,
                'request_id': request_id,
                'message': 'Task edit sent for approval.'
            }), 202

        update_work_record(conn, work_id, data)
        conn.commit()
        conn.close()
        
        return jsonify({'success': True}), 200
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/works/<int:work_id>', methods=['DELETE'])
@login_required
def delete_work(work_id):
    """Delete work record"""
    conn = get_db_connection()
    
    try:
        if not has_user_permission('approve_requests'):
            ensure_current_user_can_edit_work(conn, work_id)
            work = conn.execute('SELECT work_name FROM work_master WHERE id=? AND deleted_at IS NULL', (work_id,)).fetchone()
            if not work:
                conn.close()
                return jsonify({'success': False, 'error': 'Task not found or already deleted'}), 404

            pending_request = get_pending_task_change_request(conn, work_id)
            if pending_request:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': f"This task already has a pending {pending_request['action']} request. Please wait until it is approved or rejected."
                }), 409

            request_id = create_work_change_request(
                conn,
                'delete',
                {'work_name': work['work_name'] if work else f'Task #{work_id}'},
                work_id
            )
            conn.close()
            return jsonify({
                'success': True,
                'pending_approval': True,
                'request_id': request_id,
                'message': 'Task deletion sent for approval.'
            }), 202

        delete_work_record(conn, work_id)
        conn.commit()
        conn.close()
        
        return jsonify({'success': True}), 200
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/work-change-requests', methods=['GET'])
@login_required
def get_work_change_requests():
    conn = get_db_connection()
    try:
        if has_user_permission('approve_requests'):
            rows = conn.execute('''
                SELECT wcr.*, u.email AS requester_email, r.email AS reviewer_email, wm.work_name AS current_work_name
                FROM work_change_requests wcr
                LEFT JOIN users u ON u.id = wcr.requester_user_id
                LEFT JOIN users r ON r.id = wcr.reviewer_user_id
                LEFT JOIN work_master wm ON wm.id = wcr.work_id
                ORDER BY CASE WHEN wcr.status = 'pending' THEN 0 ELSE 1 END, wcr.created_at DESC
            ''').fetchall()
        else:
            rows = conn.execute('''
                SELECT wcr.*, u.email AS requester_email, r.email AS reviewer_email, wm.work_name AS current_work_name
                FROM work_change_requests wcr
                LEFT JOIN users u ON u.id = wcr.requester_user_id
                LEFT JOIN users r ON r.id = wcr.reviewer_user_id
                LEFT JOIN work_master wm ON wm.id = wcr.work_id
                WHERE wcr.requester_user_id = ?
                ORDER BY wcr.created_at DESC
            ''', (session.get('user_id'),)).fetchall()

        result = []
        for row in rows:
            item = dict(row)
            try:
                item['payload'] = json.loads(item.get('payload') or '{}')
            except Exception:
                item['payload'] = {}
            item['can_approve'] = has_user_permission('approve_requests') and item['status'] == 'pending'
            result.append(item)
        return jsonify(result)
    finally:
        conn.close()

@app.route('/api/work-change-requests/<int:request_id>/approve', methods=['POST'])
@login_required
@permission_required('approve_requests')
def approve_work_change_request(request_id):
    conn = get_db_connection()
    try:
        row = conn.execute('SELECT * FROM work_change_requests WHERE id = ?', (request_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Request not found'}), 404
        if row['status'] != 'pending':
            return jsonify({'success': False, 'error': 'Request already reviewed'}), 400

        payload = json.loads(row['payload'] or '{}')
        if row['action'] == 'create':
            insert_work_record(conn, payload)
        elif row['action'] == 'edit':
            update_work_record(conn, row['work_id'], payload)
        elif row['action'] == 'delete':
            soft_delete_work_record(conn, row['work_id'], request_id)
        else:
            return jsonify({'success': False, 'error': 'Unknown request action'}), 400

        conn.execute('''
            UPDATE work_change_requests
            SET status = 'approved', reviewer_user_id = ?, reviewed_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (session.get('user_id'), request_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/work-change-requests/<int:request_id>/reject', methods=['POST'])
@login_required
@permission_required('approve_requests')
def reject_work_change_request(request_id):
    conn = get_db_connection()
    try:
        row = conn.execute('SELECT status FROM work_change_requests WHERE id = ?', (request_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Request not found'}), 404
        if row['status'] != 'pending':
            return jsonify({'success': False, 'error': 'Request already reviewed'}), 400

        data = request.json or {}
        conn.execute('''
            UPDATE work_change_requests
            SET status = 'rejected', reviewer_user_id = ?, review_note = ?, reviewed_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (session.get('user_id'), data.get('note', ''), request_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/deleted-works', methods=['GET'])
@login_required
@permission_required('deleted_records')
def get_deleted_works():
    conn = get_db_connection()
    try:
        rows = conn.execute('''
            SELECT
                wm.*,
                am.name AS assignee_name,
                am.email AS assignee_email,
                au.email AS assigned_user_email,
                du.email AS deleted_by_email,
                wcr.requester_user_id,
                ru.email AS requested_by_email,
                wcr.created_at AS requested_at,
                wcr.reviewed_at AS approved_at
            FROM work_master wm
            LEFT JOIN assignee_master am ON wm.assignee_id = am.id
            LEFT JOIN users au ON wm.assigned_user_id = au.id
            LEFT JOIN users du ON wm.deleted_by_user_id = du.id
            LEFT JOIN work_change_requests wcr ON wm.delete_request_id = wcr.id
            LEFT JOIN users ru ON wcr.requester_user_id = ru.id
            WHERE wm.deleted_at IS NOT NULL
              AND wm.deleted_hidden_at IS NULL
            ORDER BY wm.deleted_at DESC, wm.id DESC
        ''').fetchall()
        return jsonify([dict(row) for row in rows])
    finally:
        conn.close()

@app.route('/api/deleted-works/<int:work_id>/restore', methods=['POST'])
@login_required
@permission_required('deleted_records')
def restore_deleted_work(work_id):
    conn = get_db_connection()
    try:
        row = conn.execute(
            'SELECT id FROM work_master WHERE id = ? AND deleted_at IS NOT NULL',
            (work_id,)
        ).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Deleted record not found'}), 404

        conn.execute('''
            UPDATE work_master
            SET deleted_at = NULL,
                deleted_by_user_id = NULL,
                delete_request_id = NULL,
                deleted_hidden_at = NULL,
                deleted_hidden_by_user_id = NULL
            WHERE id = ?
        ''', (work_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/deleted-works/<int:work_id>/hide', methods=['POST'])
@login_required
@permission_required('deleted_records')
def hide_deleted_work(work_id):
    conn = get_db_connection()
    try:
        row = conn.execute(
            'SELECT id FROM work_master WHERE id = ? AND deleted_at IS NOT NULL AND deleted_hidden_at IS NULL',
            (work_id,)
        ).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Deleted record not found'}), 404

        conn.execute('''
            UPDATE work_master
            SET deleted_hidden_at = CURRENT_TIMESTAMP,
                deleted_hidden_by_user_id = ?
            WHERE id = ?
        ''', (session.get('user_id'), work_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

# Assign To Master API
@app.route('/api/assignees', methods=['GET'])
@login_required
def get_assignees():
    """Get department records visible to the current user."""
    conn = get_db_connection()
    if can_manage_users_or_departments():
        assignees = conn.execute('SELECT * FROM assignee_master ORDER BY name ASC').fetchall()
    else:
        rows = conn.execute(
            "SELECT assignee_id FROM user_roles WHERE user_id = ? AND access_level IN ('view', 'edit')",
            (session.get('user_id'),)
        ).fetchall()
        allowed_ids = [row['assignee_id'] for row in rows]
        if allowed_ids:
            placeholders = ','.join('?' for _ in allowed_ids)
            assignees = conn.execute(
                f'SELECT * FROM assignee_master WHERE id IN ({placeholders}) ORDER BY name ASC',
                allowed_ids
            ).fetchall()
        else:
            assignees = []
    conn.close()
    return jsonify([dict(a) for a in assignees])

@app.route('/api/assignees', methods=['POST'])
@login_required
@permission_required('assign_to_master')
def add_assignee():
    """Add new assignee"""
    data = request.json
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Name is required'}), 400
    
    conn = get_db_connection()
    try:
        conn.execute('INSERT INTO assignee_master (name, email) VALUES (?, ?)', (name, email))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'success': False, 'error': 'Name already exists'}), 400
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/assignees/<int:a_id>', methods=['PUT'])
@login_required
@permission_required('assign_to_master')
def update_assignee(a_id):
    """Update assignee name"""
    data = request.json
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Name is required'}), 400
    
    conn = get_db_connection()
    try:
        conn.execute('UPDATE assignee_master SET name = ?, email = ? WHERE id = ?', (name, email, a_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/assignees/<int:a_id>', methods=['DELETE'])
@login_required
@permission_required('assign_to_master')
def delete_assignee(a_id):
    """Delete assignee"""
    conn = get_db_connection()
    conn.execute('DELETE FROM assignee_master WHERE id = ?', (a_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

@app.route('/api/user-roles', methods=['GET'])
@login_required
@permission_required('user_power_management')
def get_user_roles():
    """Get users and their assigned roles"""
    conn = get_db_connection()
    users = conn.execute('SELECT id, email FROM users').fetchall()
    roles = conn.execute('SELECT user_id, assignee_id, access_level FROM user_roles').fetchall()
    permissions = conn.execute('SELECT user_id, permission_key FROM user_permissions').fetchall()
    conn.close()
    
    roles_dict = {}
    for r in roles:
        if r['user_id'] not in roles_dict:
            roles_dict[r['user_id']] = []
        roles_dict[r['user_id']].append({
            'assignee_id': r['assignee_id'],
            'access_level': r['access_level'] or 'edit'
        })

    permissions_dict = {}
    for p in permissions:
        if p['user_id'] not in permissions_dict:
            permissions_dict[p['user_id']] = []
        permissions_dict[p['user_id']].append(p['permission_key'])
        
    return jsonify([{
        'id': u['id'],
        'email': u['email'],
        'assigned_ids': [role['assignee_id'] for role in roles_dict.get(u['id'], [])],
        'department_access': roles_dict.get(u['id'], []),
        'permissions': permissions_dict.get(u['id'], [])
    } for u in users])

@app.route('/api/user-roles', methods=['POST'])
@login_required
@permission_required('user_power_management')
def save_user_roles():
    """Save user-to-assignee mappings and powers."""
    data = request.json
    user_id = data.get('user_id')
    assignee_ids = data.get('assignee_ids', [])
    department_access = data.get('department_access', [])
    permissions = data.get('permissions', [])
    allowed_permissions = {p['key'] for p in PAGE_PERMISSIONS + SPECIAL_PERMISSIONS}
    
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM user_roles WHERE user_id = ?', (user_id,))
        if department_access:
            for item in department_access:
                a_id = item.get('assignee_id')
                access_level = item.get('access_level')
                if access_level in ('view', 'edit'):
                    conn.execute(
                        'INSERT INTO user_roles (user_id, assignee_id, access_level) VALUES (?, ?, ?)',
                        (user_id, a_id, access_level)
                    )
        else:
            for a_id in assignee_ids:
                conn.execute(
                    'INSERT INTO user_roles (user_id, assignee_id, access_level) VALUES (?, ?, ?)',
                    (user_id, a_id, 'edit')
                )

        conn.execute('DELETE FROM user_permissions WHERE user_id = ?', (user_id,))
        for permission_key in permissions:
            if permission_key in allowed_permissions:
                conn.execute(
                    'INSERT INTO user_permissions (user_id, permission_key) VALUES (?, ?)',
                    (user_id, permission_key)
                )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

# Today's Work API
@app.route('/api/assigned/<date>', methods=['GET'])
@login_required
def get_assigned_work(date):
    """Get all work assigned for a specific date with robust error handling and role filtering"""
    try:
        date_type = request.args.get('date_type', 'due')
        include_previous_open = request.args.get('include_previous_open') == '1'
        if date_type not in ('due', 'start'):
            return jsonify({'error': 'Invalid date_type. Use due or start.'}), 400

        try:
            selected_dt = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': f'Invalid request date format: {date}'}), 400
        if selected_dt < PROJECT_PERIOD_START or selected_dt > PROJECT_PERIOD_END:
            return jsonify({'error': 'Selected date is outside the project period'}), 400
            
        conn = get_db_connection()
        
        # Fetch allowed assignee IDs for the current user
        is_admin, allowed_assignee_ids = get_current_user_role_scope(conn)
        user_id = session.get('user_id')

        all_works = conn.execute('''
            SELECT wm.*, am.name AS assignee_name, am.email AS assignee_email, u.email AS assigned_user_email
            FROM work_master wm 
            LEFT JOIN assignee_master am ON wm.assignee_id = am.id
            LEFT JOIN users u ON wm.assigned_user_id = u.id
            WHERE wm.deleted_at IS NULL
        ''').fetchall()
        
        result = []

        def append_work_instance(work, occurrence_dt, only_unclosed=False):
            date_key = occurrence_dt.strftime('%Y-%m-%d')
            skipped = conn.execute(
                'SELECT id FROM work_skipped_dates WHERE work_id = ? AND skipped_date = ?',
                (work['id'], date_key)
            ).fetchone()
            if skipped:
                return

            assigned = conn.execute('''
                SELECT wa.*, am.name as override_name, am.email as override_email
                FROM work_assigned wa
                LEFT JOIN assignee_master am ON wa.assignee_id = am.id
                WHERE wa.work_id = ? AND wa.assigned_date = ?
            ''', (work['id'], date_key)).fetchone()

            current_status = assigned['status'] if assigned else 'Due'
            if only_unclosed and current_status in ('completed', 'final and approved'):
                return

            eff_assignee_id = assigned['assignee_id'] if (assigned and assigned['assignee_id']) else work['assignee_id']

            if not is_admin:
                if work['assigned_user_id']:
                    if work['assigned_user_id'] != user_id:
                        return
                elif eff_assignee_id not in allowed_assignee_ids:
                    return

            result.append({
                'work_id': work['id'],
                'work_name': work['work_name'],
                'work_tat': work['work_tat'],
                'priority': work['priority'],
                'work_date': work['work_date'],
                'work_start_date': work['work_start_date'] or work['work_date'],
                'date_type': date_type,
                'assigned_date': date_key,
                'is_previous_open': occurrence_dt < selected_dt,
                'status': current_status,
                'assignee_id': eff_assignee_id,
                'assignee_name': assigned['override_name'] if (assigned and assigned['override_name']) else work['assignee_name'],
                'assignee_email': assigned['override_email'] if (assigned and assigned['override_email']) else work['assignee_email'],
                'assigned_user_id': work['assigned_user_id'],
                'assigned_user_email': work['assigned_user_email'],
                'actual_start': assigned['actual_start'] if assigned else '',
                'actual_end': assigned['actual_end'] if assigned else '',
                'remarks': assigned['remarks'] if assigned else ''
            })

        for work in all_works:
            schedule_date = (work['work_start_date'] or work['work_date']) if date_type == 'start' else work['work_date']
            if not schedule_date: continue

            try:
                # Attempt to parse the date from the database
                start_dt = datetime.strptime(schedule_date, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                # If the date is in an invalid format (e.g. DD.MM.YYYY or empty),
                # we skip this specific task to prevent the whole page from crashing.
                print(f"Skipping task {work['id']} due to invalid date format: {schedule_date}")
                continue

            if include_previous_open:
                curr_dt = max(start_dt, PROJECT_PERIOD_START)
                while curr_dt < selected_dt:
                    if is_date_scheduled(start_dt, curr_dt, work['work_tat']):
                        append_work_instance(work, curr_dt, only_unclosed=True)
                    curr_dt += timedelta(days=1)

            if is_date_scheduled(start_dt, selected_dt, work['work_tat']):
                append_work_instance(work, selected_dt)
                continue

        result.sort(key=lambda item: (item['assigned_date'], item['work_name'].lower()))
        conn.close()
        print(f"API: Fetched {len(result)} tasks for {date} (is_admin: {is_admin})")
        return jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f"Server error: {str(e)}"}), 500

@app.route('/api/assigned/all', methods=['GET'])
@login_required
def get_all_assigned_work():
    """Get all work assignments for export/backup purposes"""
    try:
        conn = get_db_connection()
        assigned_works = conn.execute('''
            SELECT wa.*, am.name as assignee_name, am.email as assignee_email
            FROM work_assigned wa
            INNER JOIN work_master wm ON wm.id = wa.work_id
            LEFT JOIN assignee_master am ON wa.assignee_id = am.id
            WHERE wm.deleted_at IS NULL
        ''').fetchall()
        conn.close()
        
        result = [dict(row) for row in assigned_works]
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/works/template', methods=['GET'])
@login_required
def get_import_template():
    """Generates an Excel template with dropdown lists for bulk import"""
    wb = openpyxl.Workbook()
    
    # Create Fixed Data sheet for dropdowns
    ws_fixed = wb.create_sheet("Fixed Data")
    frequencies = ['daily', 'weekly', 'monthly', 'quarterly', 'yearly', 'once', 'ongoing']
    priorities = ['Normal', 'Medium', 'High']

    conn = get_db_connection()
    assignees = [a['name'] for a in conn.execute('SELECT name FROM assignee_master ORDER BY name ASC').fetchall()]
    conn.close()
    
    ws_fixed.append(['Frequencies', 'Priorities', 'Assignees'])
    for i in range(max(len(frequencies), len(priorities), len(assignees))):
        freq = frequencies[i] if i < len(frequencies) else ""
        prio = priorities[i] if i < len(priorities) else ""
        asgn = assignees[i] if i < len(assignees) else ""
        ws_fixed.append([freq, prio, asgn])
    
    # Hide the Fixed Data sheet to keep it clean
    ws_fixed.sheet_state = 'hidden'

    # Prepare Input Data sheet
    ws_input = wb.active
    ws_input.title = "Input Data"
    headers = ['Work Name', 'Frequency', 'Priority', 'Target Date (DD.MM.YYYY)', 'Work Start Date (DD.MM.YYYY)', 'Assign To']
    ws_input.append(headers)
    
    # Style headers
    for cell in ws_input[1]:
        cell.font = Font(bold=True)

    # Add Data Validation (Dropdowns)
    # Frequency dropdown (Column B)
    dv_freq = DataValidation(type="list", formula1=f"'Fixed Data'!$A$2:$A${len(frequencies)+1}", allow_blank=True)
    ws_input.add_data_validation(dv_freq)
    dv_freq.add("B2:B1000")
    
    # Priority dropdown (Column C)
    dv_prio = DataValidation(type="list", formula1=f"'Fixed Data'!$B$2:$B${len(priorities)+1}", allow_blank=True)
    ws_input.add_data_validation(dv_prio)
    dv_prio.add("C2:C1000")

    # Assignee dropdown (Column E)
    if assignees:
        dv_asgn = DataValidation(type="list", formula1=f"'Fixed Data'!$C$2:$C${len(assignees)+1}", allow_blank=True)
        ws_input.add_data_validation(dv_asgn)
        dv_asgn.add("F2:F1000")

    # Add a sample row
    ws_input.append(['Sample Task', 'monthly', 'High', '01.01.2024', '01.01.2024', assignees[0] if assignees else ""])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": "attachment; filename=task_template.xlsx"}
    )

@app.route('/api/works/validate-import', methods=['POST'])
@login_required
def validate_import():
    """Validates uploaded Excel data and returns a list with status"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    try:
        wb = openpyxl.load_workbook(file)
        if "Input Data" not in wb.sheetnames:
            return jsonify({'error': 'Invalid template: Missing "Input Data" sheet'}), 400
        ws = wb["Input Data"]
    except Exception as e:
        return jsonify({'error': f'Failed to read Excel file: {str(e)}'}), 400

    valid_frequencies = ['daily', 'weekly', 'monthly', 'quarterly', 'yearly', 'once', 'ongoing']
    valid_priorities = ['Normal', 'Medium', 'High']

    conn = get_db_connection()
    assignee_map = {a['name']: a['id'] for a in conn.execute('SELECT id, name FROM assignee_master').fetchall()}
    conn.close()
    
    results = []
    # Iterate through rows, skipping header
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue  # Skip empty rows
        
        errors = []
        work_name = str(row[0]).strip() if row[0] else ""
        freq = str(row[1]).strip().lower() if row[1] else ""
        priority = str(row[2]).strip() if row[2] else "Normal"
        raw_work_start = row[4]
        
        # Handle date which might be a datetime object or string from Excel
        raw_date = row[3]
        start_date = ""
        if isinstance(raw_date, datetime):
            start_date = raw_date.strftime('%Y-%m-%d')
        elif raw_date:
            date_str = str(raw_date).strip()
            try:
                # Attempt to parse DD.MM.YYYY format and convert to system format
                dt_obj = datetime.strptime(date_str, '%d.%m.%Y')
                start_date = dt_obj.strftime('%Y-%m-%d')
            except ValueError:
                try:
                    # Fallback to YYYY-MM-DD if user still uses the ISO format
                    datetime.strptime(date_str, '%Y-%m-%d')
                    start_date = date_str
                except ValueError:
                    start_date = date_str # Keep original value to show in error message
        
        work_start_date = ""
        if isinstance(raw_work_start, datetime):
            work_start_date = raw_work_start.strftime('%Y-%m-%d')
        elif raw_work_start:
            date_str = str(raw_work_start).strip()
            try:
                dt_obj = datetime.strptime(date_str, '%d.%m.%Y')
                work_start_date = dt_obj.strftime('%Y-%m-%d')
            except ValueError:
                try:
                    datetime.strptime(date_str, '%Y-%m-%d')
                    work_start_date = date_str
                except ValueError:
                    work_start_date = date_str

        if not work_name:
            errors.append("Work Name is required")
        if freq not in valid_frequencies:
            errors.append(f"Invalid Frequency (Must be: {', '.join(valid_frequencies)})")
        if priority not in valid_priorities:
            priority = 'Normal'
            
        try:
            target_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            if target_dt < PROJECT_PERIOD_START or target_dt > PROJECT_PERIOD_END:
                errors.append("Target Date outside project period (01.04.2026 - 31.03.2027)")
        except ValueError:
            errors.append(f"Invalid Target Date: {start_date} (Use DD.MM.YYYY)")
        if work_start_date:
            try:
                work_start_dt = datetime.strptime(work_start_date, '%Y-%m-%d').date()
                if work_start_dt < PROJECT_PERIOD_START or work_start_dt > PROJECT_PERIOD_END:
                    errors.append("Work Start Date outside project period (01.04.2026 - 31.03.2027)")
            except ValueError:
                errors.append(f"Invalid Work Start Date: {work_start_date} (Use DD.MM.YYYY)")
            
        assignee_name = str(row[5]).strip() if len(row) > 5 and row[5] else None
        assignee_id = None
        if assignee_name:
            if assignee_name in assignee_map:
                assignee_id = assignee_map[assignee_name]
            else:
                errors.append(f"Assignee '{assignee_name}' not found in Master")

        results.append({
            'work_name': work_name,
            'work_tat': freq,
            'priority': priority,
            'work_date': start_date,
            'work_start_date': work_start_date,
            'assignee_id': assignee_id,
            'status': 'Valid' if not errors else 'Error',
            'errors': errors
        })
        
    return jsonify(results)

@app.route('/api/works/bulk-add', methods=['POST'])
@login_required
def bulk_add_works():
    """Finalizes the import of valid tasks"""
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
        
    conn = get_db_connection()
    count = 0
    try:
        for item in data:
            # Only import items that were marked as Valid
            if item.get('status') == 'Valid':
                conn.execute(
                    'INSERT INTO work_master (work_name, day_of_month, work_date, work_start_date, work_tat, priority, assignee_id) VALUES (?, ?, ?, ?, ?, ?, ?)',
                    (item['work_name'], None, item['work_date'], item['work_start_date'], item['work_tat'], item['priority'], item.get('assignee_id'))
                )
                count += 1
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'count': count}), 201
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

# Report API
@app.route('/api/reports/status', methods=['GET'])
@login_required
def get_status_report():
    """Generate a report based on status and date range"""
    try:
        status_filter_raw = request.args.get('status')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if not start_date_str or not end_date_str or not status_filter_raw:
            return jsonify({'error': 'Missing parameters'}), 400

        status_list = status_filter_raw.split(',')

        start_dt = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        conn = get_db_connection()
        is_admin, allowed_assignee_ids = get_current_user_role_scope(conn)
        user_id = session.get('user_id')
        all_works = conn.execute('''
            SELECT wm.*, am.name AS assignee_name, am.email AS assignee_email, u.email AS assigned_user_email
            FROM work_master wm 
            LEFT JOIN assignee_master am ON wm.assignee_id = am.id
            LEFT JOIN users u ON wm.assigned_user_id = u.id
            WHERE wm.deleted_at IS NULL
        ''').fetchall()
        
        report_data = []
        curr_dt = start_dt
        while curr_dt <= end_dt:
            date_str = curr_dt.strftime('%Y-%m-%d')
            
            for work in all_works:
                if not work['work_date']: continue
                
                task_start_dt = datetime.strptime(work['work_date'], '%Y-%m-%d').date()
                if is_date_scheduled(task_start_dt, curr_dt, work['work_tat']):
                    skipped = conn.execute(
                        'SELECT id FROM work_skipped_dates WHERE work_id = ? AND skipped_date = ?',
                        (work['id'], date_str)
                    ).fetchone()
                    if skipped:
                        continue

                    assigned = conn.execute('''
                        SELECT wa.*, am.name as override_name, am.email as override_email
                        FROM work_assigned wa
                        LEFT JOIN assignee_master am ON wa.assignee_id = am.id
                        WHERE wa.work_id = ? AND wa.assigned_date = ?
                    ''', (work['id'], date_str)).fetchone()
                    
                    current_status = assigned['status'] if assigned else 'Due'
                    current_assignee_id = assigned['assignee_id'] if (assigned and assigned['assignee_id']) else work['assignee_id']
                    current_assignee_name = assigned['override_name'] if (assigned and assigned['override_name']) else work['assignee_name']
                    current_assignee_email = assigned['override_email'] if (assigned and assigned['override_email']) else work['assignee_email']

                    if not is_admin:
                        if work['assigned_user_id']:
                            if work['assigned_user_id'] != user_id:
                                continue
                        elif current_assignee_id not in allowed_assignee_ids:
                            continue
                    
                    if 'All' in status_list or current_status in status_list:
                        report_data.append({
                            'work_id': work['id'],
                            'work_name': work['work_name'],
                            'work_tat': work['work_tat'],
                            'work_date': work['work_date'],
                            'assignee_id': current_assignee_id,
                            'assigned_user_id': work['assigned_user_id'],
                            'assigned_user_email': work['assigned_user_email'],
                            'assignee_name': current_assignee_name,
                            'assignee_email': current_assignee_email,
                            'scheduled_date': date_str,
                            'status': current_status,
                            'actual_start': assigned['actual_start'] if assigned else '',
                            'actual_end': assigned['actual_end'] if assigned else '',
                            'frequency': work['work_tat'],
                            'remarks': assigned['remarks'] if assigned else ''
                        })
            curr_dt += timedelta(days=1)
            
        conn.close()
        return jsonify(report_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/assigned', methods=['POST'])
@login_required
def assign_work():
    """Save or Update specific task instance tracking details"""
    data = request.json
    work_id = data.get('work_id')
    assigned_date = data.get('assigned_date')
    status = data.get('status', 'Due')
    actual_start = data.get('actual_start')
    actual_end = data.get('actual_end')
    remarks = data.get('remarks', '')
    assignee_id = data.get('assignee_id')
    
    conn = get_db_connection()
    try:
        work = conn.execute(
            'SELECT id FROM work_master WHERE id = ? AND deleted_at IS NULL',
            (work_id,)
        ).fetchone()
        if not work:
            return jsonify({'success': False, 'error': 'Task not found or already deleted'}), 404

        conn.execute(
            'DELETE FROM work_skipped_dates WHERE work_id = ? AND skipped_date = ?',
            (work_id, assigned_date)
        )
        existing = conn.execute('SELECT id FROM work_assigned WHERE work_id = ? AND assigned_date = ?', 
                               (work_id, assigned_date)).fetchone()
        
        if existing:
            conn.execute('''UPDATE work_assigned SET status=?, actual_start=?, actual_end=?, remarks=?, assignee_id=? WHERE id=?''',
                        (status, actual_start, actual_end, remarks, assignee_id, existing['id']))
        else:
            conn.execute('''INSERT INTO work_assigned (work_id, assigned_date, status, actual_start, actual_end, remarks, assignee_id) 
                           VALUES (?, ?, ?, ?, ?, ?, ?)''', (work_id, assigned_date, status, actual_start, actual_end, remarks, assignee_id))
        
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/assigned/<int:work_id>/<date>', methods=['DELETE'])
@login_required
def skip_work_occurrence(work_id, date):
    """Hide one scheduled occurrence without deleting the recurring work series."""
    try:
        datetime.strptime(date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format'}), 400

    conn = get_db_connection()
    try:
        work = conn.execute('SELECT id FROM work_master WHERE id = ? AND deleted_at IS NULL', (work_id,)).fetchone()
        if not work:
            return jsonify({'success': False, 'error': 'Task not found'}), 404

        conn.execute(
            'INSERT OR IGNORE INTO work_skipped_dates (work_id, skipped_date) VALUES (?, ?)',
            (work_id, date)
        )
        conn.execute(
            'DELETE FROM work_assigned WHERE work_id = ? AND assigned_date = ?',
            (work_id, date)
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/export/excel', methods=['POST'])
@login_required
def export_excel():
    """Generic endpoint to export JSON data to a formatted Excel file"""
    data = request.json
    filename = data.get('filename', 'export.xlsx')
    headers = data.get('headers', [])
    rows = data.get('rows', [])
    title = data.get('title', '')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    current_row = 1
    if title:
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.font = Font(bold=True, size=14, color="4E54C8")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max(len(headers), 1))
        current_row += 1

    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=i, value=header)
        cell.font = Font(bold=True, color="4E54C8")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    data_start_row = current_row + 1
    for r_idx, row in enumerate(rows, data_start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if c_idx == 1 or c_idx == 2 else "center", vertical="center")
            
    # Auto-adjust column width
    for col in ws.columns:
        # Start checking length from the header row to prevent title from stretching column A
        relevant_cells = col[current_row-1:] if len(col) >= current_row else col
        lengths = [len(str(cell.value or "")) for cell in relevant_cells]
        max_length = max(lengths) if lengths else 10
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 3, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)

@app.route('/api/assigned/<int:assignment_id>/status', methods=['PUT'])
def update_status(assignment_id):
    """This endpoint is kept for compatibility but not used in new flow"""
    return jsonify({'success': True, 'message': 'Status tracking not needed in new flow'}), 200

@app.route('/api/assigned/<int:assignment_id>', methods=['DELETE'])
def remove_assignment(assignment_id):
    """This endpoint is kept for compatibility but not used in new flow"""
    return jsonify({'success': True, 'message': 'No need to remove auto-assigned tasks'}), 200

def mount_daily_debtor_report():
    from debtorapp import app as debtor_report_module

    debtor_app = debtor_report_module.app
    debtor_app.secret_key = app.secret_key
    try:
        debtor_report_module.init_db()
    except sqlite3.OperationalError as exc:
        print(f"Daily Debtor Report database init skipped: {exc}")

    @debtor_app.before_request
    def require_daily_debtor_report_access():
        if request.endpoint == 'static':
            return None
        if not session.get('user_id'):
            return redirect('/login')
        if not has_user_permission('daily_debtor_report'):
            return redirect('/')
        if (request.path.startswith('/control-panel') or request.path.startswith('/deleted-records')) and not is_arif_user():
            return redirect('/daily-debtor-report/')
        return None

    app.wsgi_app = DispatcherMiddleware(app.wsgi_app, {
        '/daily-debtor-report': debtor_app,
    })

def mount_client_analytic():
    from client_analytic import app as client_analytic_module

    client_app = client_analytic_module.app
    client_app.secret_key = app.secret_key

    @client_app.before_request
    def require_client_analytic_access():
        if request.endpoint == 'static':
            return None
        if not session.get('user_id'):
            return redirect('/login')
        if not has_user_permission('client_analytic'):
            return redirect('/')
        return None

    app.wsgi_app = DispatcherMiddleware(app.wsgi_app, {
        '/client-analytic': client_app,
    })

mount_daily_debtor_report()
mount_client_analytic()

if __name__ == '__main__':
    try:
        from waitress import serve
        serve(app, host='0.0.0.0', port=5001, threads=12, connection_limit=200)
    except ImportError:
        app.run(host='0.0.0.0', debug=False, use_reloader=False, threaded=True, port=5001)
