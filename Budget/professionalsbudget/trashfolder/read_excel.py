from openpyxl import load_workbook

# Read Excel file
excel_file = r'd:\01software backup\Asijaapp\Budget\professionalsbudget\Book1.xlsx'
wb = load_workbook(excel_file)
ws = wb.active

print("Sheet Name:", ws.title)
print("\nExcel Data:")
print("="*60)

for row in ws.iter_rows(values_only=True):
    print(row)
