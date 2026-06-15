import os
import sqlite3
from datetime import datetime

from flask import Flask, render_template, url_for
from openpyxl import load_workbook


app = Flask(
    __name__,
    template_folder='templates',
    static_folder='static',
)
app.secret_key = 'client-analytic-dev-secret'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'database.db')
INPUT_FILE_PATH = os.path.join(BASE_DIR, 'input file.xlsx')


def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def get_meta(cursor, key, default=''):
    cursor.execute('SELECT value FROM app_meta WHERE key = ?', (key,))
    row = cursor.fetchone()
    return row['value'] if row else default


def set_meta(cursor, key, value):
    cursor.execute('''
        INSERT INTO app_meta (key, value)
        VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    ''', (key, str(value)))


def format_excel_value(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()


def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS app_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_year_report (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_name TEXT,
            row_no INTEGER,
            year TEXT,
            client_group_name TEXT,
            type TEXT,
            client_sourced_by TEXT,
            doj TEXT,
            imported_at TEXT
        )
    ''')
    conn.commit()
    conn.close()
    import_input_file_if_needed()


def import_input_file_if_needed(force=False):
    if not os.path.exists(INPUT_FILE_PATH):
        return {'imported': False, 'row_count': 0, 'message': 'input file.xlsx not found'}

    file_mtime = str(os.path.getmtime(INPUT_FILE_PATH))
    conn = get_db_connection()
    cursor = conn.cursor()
    previous_mtime = get_meta(cursor, 'input_file_mtime')
    if not force and previous_mtime == file_mtime:
        cursor.execute('SELECT COUNT(*) AS count FROM client_year_report')
        row_count = cursor.fetchone()['count']
        conn.close()
        return {'imported': False, 'row_count': row_count, 'message': 'already current'}

    workbook = load_workbook(INPUT_FILE_PATH, data_only=True, read_only=True)
    imported_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    rows_to_insert = []

    for sheet in workbook.worksheets:
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if index == 1:
                continue
            values = list(row[:5])
            values.extend([None] * (5 - len(values)))
            if not any(value not in (None, '') for value in values):
                continue
            rows_to_insert.append((
                sheet.title,
                index,
                format_excel_value(values[0]),
                format_excel_value(values[1]),
                format_excel_value(values[2]),
                format_excel_value(values[3]),
                format_excel_value(values[4]),
                imported_at,
            ))

    workbook.close()
    cursor.execute('DELETE FROM client_year_report')
    cursor.executemany('''
        INSERT INTO client_year_report (
            sheet_name, row_no, year, client_group_name, type,
            client_sourced_by, doj, imported_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', rows_to_insert)
    set_meta(cursor, 'input_file_mtime', file_mtime)
    set_meta(cursor, 'input_file_name', os.path.basename(INPUT_FILE_PATH))
    set_meta(cursor, 'input_file_imported_at', imported_at)
    conn.commit()
    conn.close()
    return {'imported': True, 'row_count': len(rows_to_insert), 'message': 'imported'}


def client_analytic_url(path='/'):
    path = '/' + str(path or '/').lstrip('/')
    if path == '/':
        return url_for('dashboard')
    return path


@app.context_processor
def inject_client_analytic_helpers():
    return {
        'client_analytic_url': client_analytic_url,
        'report_as_on_label': f"Client Analytic as on {datetime.now().strftime('%d.%m.%Y')}",
    }


def get_sample_client_rows():
    return [
        {
            'client_name': 'B K Saraf Private Limited',
            'client_group': 'BKS Group',
            'category': 'Private',
            'crp': 'RohitS',
            'followup_partner': 'SahilD',
            'total_bills': 8,
            'outstanding': 186540.0,
            'received': 92500.0,
            'overdue_bills': 3,
            'last_activity': '12-06-26',
            'status': 'Watch',
        },
        {
            'client_name': 'Karam Safety Private Limited',
            'client_group': 'Karam',
            'category': 'Corporate',
            'crp': 'KamalF',
            'followup_partner': 'RohitS',
            'total_bills': 11,
            'outstanding': 472000.0,
            'received': 118000.0,
            'overdue_bills': 5,
            'last_activity': '10-06-26',
            'status': 'High',
        },
        {
            'client_name': 'La Martiniere College, Lucknow',
            'client_group': 'Education',
            'category': 'Trust',
            'crp': 'AkashA',
            'followup_partner': 'RohitS',
            'total_bills': 4,
            'outstanding': 82400.0,
            'received': 54000.0,
            'overdue_bills': 1,
            'last_activity': '08-06-26',
            'status': 'Normal',
        },
    ]


def get_year_wise_report_rows():
    import_input_file_if_needed()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT year, client_group_name, type, client_sourced_by, doj
        FROM client_year_report
        ORDER BY row_no
    ''')
    rows = [dict(row) for row in cursor.fetchall()]
    cursor.execute('''
        SELECT
            COUNT(*) AS total_rows,
            COUNT(DISTINCT year) AS year_count,
            COUNT(DISTINCT type) AS type_count,
            COUNT(DISTINCT client_sourced_by) AS source_count
        FROM client_year_report
    ''')
    summary = dict(cursor.fetchone())
    filter_options = {
        'years': sorted({row['year'] for row in rows if row.get('year')}),
        'types': sorted({row['type'] for row in rows if row.get('type')}),
        'sources': sorted({row['client_sourced_by'] for row in rows if row.get('client_sourced_by')}),
    }
    meta = {
        'file_name': get_meta(cursor, 'input_file_name', 'input file.xlsx'),
        'imported_at': get_meta(cursor, 'input_file_imported_at', ''),
    }
    conn.close()
    return rows, summary, meta, filter_options


@app.route('/')
def dashboard():
    rows = get_sample_client_rows()
    summary = {
        'clients': len(rows),
        'bills': sum(row['total_bills'] for row in rows),
        'outstanding': sum(row['outstanding'] for row in rows),
        'received': sum(row['received'] for row in rows),
        'overdue': sum(row['overdue_bills'] for row in rows),
    }
    return render_template(
        'dashboard.html',
        active_page='client_dashboard',
        page_title='Client Analytic',
        rows=rows,
        summary=summary,
    )


@app.route('/year-wise-report')
def year_wise_report():
    rows, summary, meta, filter_options = get_year_wise_report_rows()
    return render_template(
        'year_wise_report.html',
        active_page='year_wise_report',
        page_title='Year Wise Client Report',
        rows=rows,
        summary=summary,
        meta=meta,
        filter_options=filter_options,
    )


init_db()


if __name__ == '__main__':
    app.run(debug=True, port=5006)
