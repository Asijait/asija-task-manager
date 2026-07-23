import os
import re
import sys
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

import pg8000.dbapi
from urllib.parse import unquote, urlparse
from dotenv import dotenv_values
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
LOCAL_ENV = dotenv_values(ROOT / ".env")
app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv(
    "WORKREPORT_SECRET_KEY", LOCAL_ENV.get("SECRET_KEY", "dev-only-secret")
)
DATABASE_URL = os.getenv(
    "WORKREPORT_DATABASE_URL",
    LOCAL_ENV.get("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/workreport"),
)


class Result:
    def __init__(self, cursor):
        self.cursor = cursor
        self.rowcount = cursor.rowcount
        self.columns = [item[0] for item in cursor.description] if cursor.description else []

    def fetchone(self):
        row = self.cursor.fetchone()
        return dict(zip(self.columns, row)) if row else None

    def fetchall(self):
        return [dict(zip(self.columns, row)) for row in self.cursor.fetchall()]


class Database:
    def __init__(self):
        url = urlparse(DATABASE_URL)
        self.conn = pg8000.dbapi.connect(
            user=unquote(url.username or "postgres"),
            password=unquote(url.password or ""),
            host=url.hostname or "localhost",
            port=url.port or 5432,
            database=(url.path or "/postgres").lstrip("/"),
        )

    def execute(self, sql, params=None):
        cursor = self.conn.cursor()
        cursor.execute(sql, params or ())
        return Result(cursor)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, *_):
        if exc_type:
            self.conn.rollback()
        else:
            self.conn.commit()
        self.conn.close()


def db():
    return Database()


def ensure_database():
    """Create the configured local database when it does not exist yet."""
    url = urlparse(DATABASE_URL)
    database_name = (url.path or "/workreport").lstrip("/")
    if not re.fullmatch(r"[A-Za-z0-9_]+", database_name):
        raise ValueError("Database name may contain only letters, numbers and underscores")
    admin = pg8000.dbapi.connect(
        user=unquote(url.username or "postgres"), password=unquote(url.password or ""),
        host=url.hostname or "localhost", port=url.port or 5432, database="postgres"
    )
    admin.autocommit = True
    cursor = admin.cursor()
    cursor.execute("SELECT 1 FROM pg_database WHERE datname = %s", [database_name])
    if not cursor.fetchone():
        cursor.execute(f'CREATE DATABASE "{database_name}"')
    admin.close()


def iso(value):
    return value.isoformat() if isinstance(value, (date, datetime)) else value


def clean_row(row):
    return {key: iso(value) for key, value in row.items()}


def current_user():
    provider = app.config.get("CURRENT_USER_PROVIDER")
    return provider() if provider else None


def init_db():
    with db() as conn:
        for statement in (ROOT / "schema.sql").read_text(encoding="utf-8").split(";"):
            if statement.strip():
                conn.execute(statement)


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


FIELDS = ["work_name", "section", "status", "work_inflow", "next_scheduled",
          "reschedule_scheduled", "target_date", "actual_completion_date", "remark", "allotted_to"]
DATE_FIELDS = {"work_inflow", "next_scheduled", "reschedule_scheduled", "target_date", "actual_completion_date"}


def payload_values(data):
    values = []
    for field in FIELDS:
        value = data.get(field)
        values.append(parse_date(value) if field in DATE_FIELDS else (str(value).strip() if value else None))
    if not values[0]:
        raise ValueError("Work name is required")
    values[2] = values[2] or "Not Started"
    if values[2] == "Done" and not values[7]:
        values[7] = date.today()
    elif values[2] != "Done":
        values[7] = None
    return values


def work_filters(args):
    status = args.get("status", "all")
    search = args.get("search", "").strip()
    scope = args.get("scope", "all")
    where, params = ["deleted_at IS NULL"], []
    user = current_user()
    if user:
        visible_names = user.get("visible_names") or [user["name"]]
        where.append("LOWER(BTRIM(allotted_to)) IN (" + ",".join(["LOWER(%s)"] * len(visible_names)) + ")")
        params.extend(visible_names)
        requested_team = args.get("team", user["name"]).strip()
        if requested_team.lower() != "all":
            selected_team = next((name for name in visible_names if name.lower() == requested_team.lower()), user["name"])
            where.append("LOWER(BTRIM(allotted_to)) = LOWER(%s)")
            params.append(selected_team)
    if status != "all":
        where.append("status = %s")
        params.append(status)
    if search:
        where.append("(work_name ILIKE %s OR allotted_to ILIKE %s OR remark ILIKE %s)")
        params.extend([f"%{search}%"] * 3)
    if scope == "pending":
        where.append("status <> 'Done'")
    elif scope == "overdue":
        where.append("status <> 'Done' AND target_date < CURRENT_DATE")
    elif scope == "today":
        where.append("status IN ('Not Started', 'WIP') AND COALESCE(reschedule_scheduled, next_scheduled, target_date) <= CURRENT_DATE")
    elif scope == "start_today":
        where.append("status IN ('Not Started', 'WIP') AND COALESCE(reschedule_scheduled, next_scheduled) <= CURRENT_DATE")
    elif scope == "upcoming":
        where.append("status <> 'Done' AND target_date > CURRENT_DATE AND target_date <= CURRENT_DATE + 7")
    elif scope == "new_requests" and user:
        where.append("id IN (SELECT work_item_id FROM work_notifications WHERE LOWER(recipient_alias) = LOWER(%s))")
        params.append(user["name"])
    clause = " WHERE " + " AND ".join(where) if where else ""
    return clause, params, scope


@app.get("/")
def index():
    return render_template("index.html")


@app.get("/api/work")
def list_work():
    clause, params, _ = work_filters(request.args)
    sql = f"SELECT * FROM work_items{clause} ORDER BY CASE WHEN status = 'Done' THEN 1 ELSE 0 END, target_date NULLS LAST, id DESC"
    with db() as conn:
        rows = conn.execute(sql, params).fetchall()
    return jsonify([clean_row(row) for row in rows])


@app.get("/api/export")
def export_work():
    clause, params, scope = work_filters(request.args)
    with db() as conn:
        rows = conn.execute(f"""SELECT excel_sl, work_name, section, status, work_inflow,
            next_scheduled, reschedule_scheduled, target_date, actual_completion_date,
            remark, allotted_to FROM work_items{clause}
            ORDER BY CASE WHEN status = 'Done' THEN 1 ELSE 0 END, target_date NULLS LAST, id DESC""", params).fetchall()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Work Report"
    headers = ["SL", "Work Name", "Work Allotted By", "Status", "Work Inflow", "Next Scheduled",
               "Reschedule Scheduled", "Target Date", "Actual Completion Date", "Remark", "Work Alloted to"]
    for col, heading in enumerate(headers, 1):
        cell = sheet.cell(1, col, heading)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="34495E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    keys = ["excel_sl", "work_name", "section", "status", "work_inflow", "next_scheduled",
            "reschedule_scheduled", "target_date", "actual_completion_date", "remark", "allotted_to"]
    for number, row in enumerate(rows, 1):
        sheet.append([row[key] if key != "excel_sl" else (row[key] or number) for key in keys])
    for row_number in range(2, sheet.max_row + 1):
        for column_number in range(5, 10):
            sheet.cell(row_number, column_number).number_format = "DD-MM-YYYY"
    widths = [8, 42, 20, 15, 15, 18, 22, 15, 22, 42, 20]
    for column_number, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column_number)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    safe_scope = scope if scope in {"all", "pending", "overdue", "today", "start_today", "upcoming"} else "report"
    return send_file(output, as_attachment=True,
                     download_name=f"work_{safe_scope}_{date.today().isoformat()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/summary")
def summary():
    user = current_user()
    visible_names = (user.get("visible_names") or [user["name"]]) if user else []
    conditions, params = ["deleted_at IS NULL"], []
    if user:
        conditions.append("LOWER(BTRIM(allotted_to)) IN (" + ",".join(["LOWER(%s)"] * len(visible_names)) + ")")
        params.extend(visible_names)
        requested_team = request.args.get("team", user["name"]).strip()
        if requested_team.lower() != "all":
            selected_team = next((name for name in visible_names if name.lower() == requested_team.lower()), user["name"])
            conditions.append("LOWER(BTRIM(allotted_to)) = LOWER(%s)")
            params.append(selected_team)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    with db() as conn:
        row = conn.execute(f"""
            SELECT COUNT(*) AS total,
              COUNT(*) FILTER (WHERE status = 'Done') AS done,
              COUNT(*) FILTER (WHERE status <> 'Done') AS pending,
              COUNT(*) FILTER (WHERE status <> 'Done' AND target_date < CURRENT_DATE) AS overdue,
              COUNT(*) FILTER (WHERE status IN ('Not Started', 'WIP')
                               AND COALESCE(reschedule_scheduled, next_scheduled, target_date) <= CURRENT_DATE) AS today,
              COUNT(*) FILTER (WHERE status <> 'Done' AND target_date > CURRENT_DATE
                               AND target_date <= CURRENT_DATE + 7) AS upcoming
            FROM work_items
            {where}
        """, params).fetchone()
    return jsonify(row)


@app.get("/api/allottees")
def list_allottees():
    provider = app.config.get("ALLOTTEE_PROVIDER")
    users = provider() if provider else []
    with db() as conn:
        rows = conn.execute("""
            SELECT LOWER(BTRIM(allotted_to)) AS alias_key, COUNT(*) AS active_count
            FROM work_items
            WHERE deleted_at IS NULL AND status IN ('WIP', 'Not Started') AND allotted_to IS NOT NULL
            GROUP BY LOWER(BTRIM(allotted_to))
        """).fetchall()
    counts = {row["alias_key"]: row["active_count"] for row in rows}
    return jsonify([
        {**user, "active_count": counts.get(user["name"].strip().lower(), 0)}
        for user in users
    ])


@app.get("/api/current-user")
def get_current_user():
    return jsonify(current_user() or {})


@app.get("/api/notifications")
def get_notifications():
    user = current_user()
    if not user:
        return jsonify({"unread": 0})
    with db() as conn:
        row = conn.execute("""
            SELECT COUNT(*) AS unread
            FROM work_notifications notification
            JOIN work_items work ON work.id = notification.work_item_id
            WHERE work.deleted_at IS NULL
              AND LOWER(notification.recipient_alias) = LOWER(%s)
              AND notification.is_read = FALSE
        """, [user["name"]]).fetchone()
    return jsonify(row)


@app.get("/api/todays-report")
def todays_work_report():
    user = current_user()
    if not user:
        return jsonify(error="User not found"), 401
    alias = user["name"]
    with db() as conn:
        old_completed = conn.execute("""
            SELECT * FROM work_items
            WHERE deleted_at IS NULL
              AND LOWER(BTRIM(allotted_to)) = LOWER(%s)
              AND actual_completion_date = CURRENT_DATE
              AND (created_at::date < CURRENT_DATE OR created_by_alias IS NULL OR BTRIM(created_by_alias) = '')
            ORDER BY updated_at DESC, id DESC
        """, [alias]).fetchall()
        new_added = conn.execute("""
            SELECT * FROM work_items
            WHERE deleted_at IS NULL
              AND LOWER(BTRIM(created_by_alias)) = LOWER(%s)
              AND created_at::date = CURRENT_DATE
            ORDER BY created_at DESC, id DESC
        """, [alias]).fetchall()
        tomorrow_plan = conn.execute("""
            SELECT * FROM work_items
            WHERE deleted_at IS NULL
              AND LOWER(BTRIM(allotted_to)) = LOWER(%s)
              AND status <> 'Done'
              AND (next_scheduled = CURRENT_DATE + 1
                   OR reschedule_scheduled = CURRENT_DATE + 1
                   OR target_date = CURRENT_DATE + 1)
            ORDER BY target_date NULLS LAST, id DESC
        """, [alias]).fetchall()
        report_date = conn.execute("SELECT CURRENT_DATE AS report_date").fetchone()["report_date"]
    return jsonify({
        "user": user,
        "report_date": iso(report_date),
        "old_completed": [clean_row(row) for row in old_completed],
        "new_added": [clean_row(row) for row in new_added],
        "tomorrow_plan": [clean_row(row) for row in tomorrow_plan],
    })


@app.patch("/api/notifications/read")
def read_notifications():
    user = current_user()
    if not user:
        return jsonify({"success": True, "updated": 0})
    with db() as conn:
        result = conn.execute("""
            UPDATE work_notifications SET is_read = TRUE
            WHERE LOWER(recipient_alias) = LOWER(%s) AND is_read = FALSE
        """, [user["name"]])
    return jsonify({"success": True, "updated": result.rowcount})


@app.post("/api/work")
def create_work():
    try:
        values = payload_values(request.get_json() or {})
        user = current_user()
        if user:
            values[1] = user["name"]
        if user and not values[9]:
            values[9] = user["name"]
        user = current_user()
        columns = ", ".join([*FIELDS, "created_by_alias"])
        placeholders = ", ".join(["%s"] * (len(FIELDS) + 1))
        with db() as conn:
            row = conn.execute(
                f"INSERT INTO work_items ({columns}) VALUES ({placeholders}) RETURNING *",
                [*values, user["name"] if user else None],
            ).fetchone()
            provider = app.config.get("ALLOTTEE_PROVIDER")
            valid_names = {item["name"].lower(): item["name"] for item in (provider() if provider else [])}
            recipient = valid_names.get(str(values[9] or "").lower())
            if user and recipient and recipient.lower() != user["name"].lower():
                conn.execute("""
                    INSERT INTO work_notifications (work_item_id, recipient_alias, sender_alias)
                    VALUES (%s, %s, %s)
                """, [row["id"], recipient, user["name"]])
        return jsonify(clean_row(row)), 201
    except ValueError as exc:
        return jsonify(error=str(exc)), 400


@app.put("/api/work/<int:item_id>")
def update_work(item_id):
    try:
        values = payload_values(request.get_json() or {})
        assignments = ", ".join(f"{field} = %s" for field in FIELDS)
        user = current_user()
        owner_clause = " AND deleted_at IS NULL" + (" AND LOWER(BTRIM(allotted_to)) = LOWER(%s)" if user else "")
        owner_params = [user["name"]] if user else []
        with db() as conn:
            existing = conn.execute(
                "SELECT section FROM work_items WHERE id = %s" + owner_clause,
                [item_id, *owner_params],
            ).fetchone()
            if not existing:
                return jsonify(error="Work not found"), 404
            values[1] = existing["section"] or (user["name"] if user else None)
            row = conn.execute(
                f"UPDATE work_items SET {assignments}, updated_at = NOW() WHERE id = %s{owner_clause} RETURNING *",
                [*values, item_id, *owner_params],
            ).fetchone()
        return (jsonify(clean_row(row)), 200) if row else (jsonify(error="Work not found"), 404)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400


@app.patch("/api/work/<int:item_id>/complete")
def complete_work(item_id):
    user = current_user()
    owner_clause = " AND deleted_at IS NULL" + (" AND LOWER(BTRIM(allotted_to)) = LOWER(%s)" if user else "")
    params = [item_id, user["name"]] if user else [item_id]
    with db() as conn:
        row = conn.execute("""UPDATE work_items SET status='Done', actual_completion_date=CURRENT_DATE,
                            updated_at=NOW() WHERE id=%s""" + owner_clause + " RETURNING *", params).fetchone()
    return (jsonify(clean_row(row)), 200) if row else (jsonify(error="Work not found"), 404)


@app.delete("/api/work/<int:item_id>")
def delete_work(item_id):
    user = current_user()
    owner_clause = " AND deleted_at IS NULL" + (" AND LOWER(BTRIM(allotted_to)) = LOWER(%s)" if user else "")
    params = [item_id, user["name"]] if user else [item_id]
    with db() as conn:
        result = conn.execute(
            "UPDATE work_items SET deleted_at=NOW(), deleted_by_alias=%s, updated_at=NOW() WHERE id=%s" + owner_clause,
            [user["name"] if user else "Unknown", *params],
        )
    return ("", 204) if result.rowcount else (jsonify(error="Work not found"), 404)


def import_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    count = 0
    sql = """INSERT INTO work_items
        (excel_sl, work_name, section, status, work_inflow, next_scheduled,
         reschedule_scheduled, target_date, actual_completion_date, remark, allotted_to)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (excel_sl) WHERE excel_sl IS NOT NULL DO UPDATE SET
        work_name=EXCLUDED.work_name, section=EXCLUDED.section, status=EXCLUDED.status,
        work_inflow=EXCLUDED.work_inflow, next_scheduled=EXCLUDED.next_scheduled,
        reschedule_scheduled=EXCLUDED.reschedule_scheduled, target_date=EXCLUDED.target_date,
        actual_completion_date=EXCLUDED.actual_completion_date, remark=EXCLUDED.remark,
        allotted_to=EXCLUDED.allotted_to, updated_at=NOW()"""
    with db() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue
            status = row[3] if row[3] in {"Not Started", "WIP", "Done", "On Hold"} else "Not Started"
            conn.execute(sql, [row[0], row[1], row[2], status, *row[4:11]])
            count += 1
    return count


def import_default_data_if_empty():
    """Seed a new integrated database from the bundled workbook once."""
    source = ROOT / "workdetail.xlsx"
    if not source.exists():
        return 0
    with db() as conn:
        row = conn.execute("SELECT COUNT(*) AS total FROM work_items").fetchone()
    return import_excel(source) if row["total"] == 0 else 0


if __name__ == "__main__":
    ensure_database()
    init_db()
    import_default_data_if_empty()
    if len(sys.argv) > 1 and sys.argv[1] == "--import":
        source = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "workdetail.xlsx"
        print(f"Imported/updated {import_excel(source)} work items from {source.name}")
    else:
        app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=True)
